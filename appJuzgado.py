from appApiModel import ApiModel
from appGenerales import AppGenerales
from appLog import Logger
import sys, time, os, json, re
from selenium.webdriver.common.by import By   
from selenium.webdriver.common.keys import Keys 
from os.path import join as pjoin, exists, dirname, abspath
from urllib.error import HTTPError
from PyPDF2 import PdfReader, PdfWriter
import tabula
import zipfile
import requests
import fitz
import pandas as pd
from pdfminer.high_level import extract_text
import datetime
import openpyxl

import pdfx 
from selenium import webdriver

class JuzgadoJSON:
    def __init__(self, url, xpath_pest, xpath_tab):
        self.url = url
        self.xpath_pest = xpath_pest
        self.xpath_tab = xpath_tab
    #__init__
#JuzgadoJSON

        
class DatosJuzgados:
    def __init__(self):
        self.juzgadosJSON = {}
    #__init__

    def agregar_juzgado(self, nombre, url, xpath_pest = '', xpath_tab = ''):
        self.juzgadosJSON[nombre] = JuzgadoJSON(url, xpath_pest, xpath_tab)
    #agregar_juzgado
#JuzgadoJSON

class JuzgadoBase:
    def __init__(self ):
        self.api_model      = ApiModel()
        self.app_general    = AppGenerales()
        self.datos_juzgados = DatosJuzgados()
        self.app_log        = Logger('logJuzgadoMunicipal')
        
        self.app_config  = self.app_general.appConfig
        self.app_config.wRutaAlmPdfDrive = self.app_config.wRutaAlmPdfDrive.replace('ANIO',str(self.app_general.fnConfigFecha('AA')))
        
        self.fnInicializarJuzgados()
                 
        #xpath pestaña mes
        self.xpath_pest_mes  = {
             'ruta-1' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/ul/li'
            ,'ruta-2' : '/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/div/div/div[1]/div[2]/div/ul/li'  
            ,'ruta-3' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div/ul/li'  
            ,'ruta-4' : '/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/ul/li'
        }   
        
        self.mapeo_rutas_tab = {
            '2022': 'ruta-6',
            '2021': 'ruta-11',
            '2020': 'ruta-19'
            # Añade más años y rutas según sea necesario
        }

    # __init__
    
    def fnInicializarJuzgados(self):
        objEnlaceMicrositio = self.api_model.fnEjecutarApiPost('BusquedaRadicacion','ListadoEnlaceMicrositioRamaJudicial')
        for data in objEnlaceMicrositio['Data']:
            self.datos_juzgados.agregar_juzgado(data['CODIGO'],data['ENLACE'], data['XPATH_PEST'],  data['XPATH_TAB'])
        #Fin for
    #fnInicializarJuzgados
    
    def fnRetornarDatosAgrupados( self, wDatosGenerales, wTipoOpcion = 'EE' ):
        """ 
            wDatosGenerales : Conjunto de datos
            wTipoOpcion : AU,EE,TR -> AUTOS, ESTADOS ELECTRONICOS, TRASLADOS
            
            NOTA: 
               La variable datos['Ciudad']: es el codigo -> [tipoJuzgado-subTipoJuzgado-Ciudad-Juzgado])
        """
        wArrDatosGenerales = {}
        for datos in wDatosGenerales:
            # variables 
            codigo = wTipoOpcion + '-' + datos['Ciudad'] + '-' + datos['Anio']
            if codigo not in wArrDatosGenerales:
                wArrDatosGenerales[codigo] = []
            #fin if 
            
            wArrDatosGenerales[codigo].append(datos)
        #fin for
        
        return wArrDatosGenerales
    #fnRetornarDatosAgrupados
     
    def fnProcesoPaginaInicio( self, dv, enlace_pagina ):
        try:
            continuar = False
            
            if dv == None:
                self.app_general.fnImprimir(f"Ingresando a la pagina",9)
                dv, wait = self.app_general.fnAbrirPagina( enlace_pagina , self.app_config.wRutaDescPdf )
                
                if dv == None:
                    self.app_general.fnImprimir(f"Se excedio el tiempo de espera de la pagina.",11)
                    self.app_log.error( f"ERROR: Se excedio el tiempo de espera de la pagina." )
                    continuar = True
                #fin if 
            else:
                if  len(dv.window_handles) == 2:
                    dv.close()
                    dv.switch_to.window(dv.window_handles[0])
                #fin if

                self.app_general.fnLimpiarCache(dv)
                # dv.execute_script("window.location.reload(true);")
            #fin if
        except HTTPError as e:
            if "502" in str(e):
                if dv != None:
                    dv.quit()
                    dv = None
                #fin if 
                
                continuar = True
                
                self.app_general.fnImprimir(f"RPTA: No termino de cargar la pagina.",11)
                self.app_log.error( f"ERROR: No termino de cargar la pagina." )
                
                pass
            #fin if  
        except:
            if dv != None:
                dv.quit()
                dv = None
            #fin if 
            continuar = True
            
            self.app_general.fnImprimir(f"RPTA: Ocurrio un incoveniente al cargar la pagina, favor de revisar el log.",11)
            self.app_log.error( f"ERROR al cargar la pagina: {sys.exc_info()[1]}" )
            
            pass
        #fin try
        
        return dv, continuar
    #fnProcesoPaginaInicio
      
    def fnRetornarXpathDescarga( self, wRuta, wMes = '' ):
        xpath_descarga  = {
             'ruta-1' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table/tbody/tr'
            ,'ruta-2' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table'
            ,'ruta-3' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table/thead/tr'
            ,'ruta-4' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/table'
            ,'ruta-5' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div/div/table/tbody/tr' 
            ,'ruta-6' : '/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table/tbody/tr'
            ,'ruta-7' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table[2]/tbody/tr'
            ,'ruta-8' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/table/tbody/tr'
            ,'ruta-9' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div/table/tbody/tr'
            ,'ruta-10' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div/div/div['+str(wMes)+']/div[1]/table/tbody/tr'
            ,'ruta-11' : '/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table/tbody/tr'
            ,'ruta-12' : '/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table'
            ,'ruta-13' : '/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table[1]/tbody/tr'
            ,'ruta-14' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table[1]/tbody/tr'
            ,'ruta-15' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[2]/div[2]/div/div/div['+str(wMes)+']/table/tbody/tr'
            ,'ruta-16' : '/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table'
            ,'ruta-17' : '/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table[1]/tbody/tr'
            ,'ruta-18' : '/html/body/div[2]/div[1]/div[4]/div/div[2]/div/div[2]/div/section/div[2]/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table'
            ,'ruta-19' : '/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/div/div/div[1]/div[2]/div/div/div['+str(wMes)+']/table[1]/tbody/tr'
        }
        
        return xpath_descarga[wRuta]
    #fnRetornarXpathDescarga
    
    def fnDatosBusqueda( self, wNroRadicacion ): 
        #11001400304320220089600
        jsonBusqueda = { 
             'tbusq-001' : (str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ))                                                    #2022-00896
            ,'tbusq-002' : (str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ) + '-' + str( wNroRadicacion[-2:] ))                 #2022-00896-00
            ,'tbusq-003' :(str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[17:21] ))                                                     #2022-0896-00
            ,'tbusq-004' : (str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[17:21] ) + '-' + str( wNroRadicacion[-2:] ))                 #2022-00896-00
            ,'tbusq-005' : (str( wNroRadicacion[9:12] ) + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ) )               #043-2022-00896
            ,'tbusq-006' : (str( int(wNroRadicacion[9:12] )) + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ) )          #43-2022-00896
            ,'tbusq-007' : (str( int(wNroRadicacion[9:12] )) + str( wNroRadicacion[12:16] ) + str( wNroRadicacion[16:21] ) )                      #43202200896
            ,'tbusq-008' : (str( wNroRadicacion[12:16] ) + '-' + str( int(wNroRadicacion[16:21]) ))                                               #2022-896
            ,'tbusq-009' : ('('+str( int(wNroRadicacion[9:12] )) + ') ' + str( wNroRadicacion[12:16] ) + '' + str( int(wNroRadicacion[16:21]) ))  #(78) 2022-896
            ,'tbusq-010' : (str( wNroRadicacion[12:16] ) + ' ' + str( wNroRadicacion[16:21] ))                                                    #2022 00896
            ,'tbusq-011' : (str( wNroRadicacion[12:16] ) + str( wNroRadicacion[16:21] ))                                                          #202200896
            ,'tbusq-012' : (str( wNroRadicacion[12:16] ) + str( wNroRadicacion[17:21] ))                                                          #20220896
            ,'tbusq-013' : (str( wNroRadicacion[0:12] ) + ' ' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] )  )             #110014003066 2020-00605
            ,'tbusq-014' : (str( wNroRadicacion[0:12] ) + ' ' + str( wNroRadicacion[12:16] ) + ' ' + str( wNroRadicacion[16:21] )+ ' ' + str( wNroRadicacion[-2:] )  )      #110014003066 2020 00605 00
            ,'tbusq-015' : wNroRadicacion                                                                                                 #110014003066202000605
            ,'tbusq-016' : (str( wNroRadicacion[0:12] ) + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ) + '-' + str( wNroRadicacion[-2:] ) )   #110014003043-2022-00896-00
            ,'tbusq-017' : (str( wNroRadicacion[0:5] ) + '-' + str( wNroRadicacion[5:7] ) + '-' + str( wNroRadicacion[7:9] )  + '-' + str( wNroRadicacion[9:12] )  + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ) + '-' + str( wNroRadicacion[-2:] ) ) #11001-40-03-043-2022-00896-00
            ,'tbusq-018' : str(wNroRadicacion[0:5]) + str(wNroRadicacion[12:16]) + str(wNroRadicacion[5:7]) + ' ' + str(wNroRadicacion[7:9]) + ' ' + str(wNroRadicacion[9:12]) + str(wNroRadicacion[16:21]) #11001201931 03 00100489
            ,'tbusq-019' : (str( wNroRadicacion[9:12] )  + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ) + '-' + str( wNroRadicacion[-2:] ) ) #043-2022-00896-00
            ,'tbusq-020' : (str( wNroRadicacion[10:12] )  + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[17:21] ) + '-' + str( wNroRadicacion[-2:] ) ) #043-2022-0896-00
            ,'tbusq-021' : ( str( wNroRadicacion[0:5] ) + '-' + str( wNroRadicacion[5:7] )  + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[7:9] )  + '-' + str( wNroRadicacion[9:12] ) + '-' + str( wNroRadicacion[16:21] ) ) #11001-40-2023-03-052-01005
            ,'tbusq-022' : ( str( wNroRadicacion[0:5] ) + ' ' + str( wNroRadicacion[5:7] )  + ' ' + str( wNroRadicacion[12:16] ) + ' ' + str( wNroRadicacion[7:9] )  + ' ' + str( wNroRadicacion[9:12] ) + ' ' + str( wNroRadicacion[16:21] ) ) #11001 40 2023 03 052 01005
            ,'tbusq-023' : ( str( wNroRadicacion[0:5] ) + ' ' + str( wNroRadicacion[5:7] ) + ' ' + str( wNroRadicacion[7:9] ) + ' ' + str( wNroRadicacion[9:12] )  + ' ' + str( wNroRadicacion[12:16] )  + ' ' + str( wNroRadicacion[16:21] )) #11001-40-03-052-2023-01005
            ,'tbusq-024' : ( str( wNroRadicacion[0:5] ) + ' ' + str( wNroRadicacion[5:7] ) + str( wNroRadicacion[12:16] ) + str( wNroRadicacion[7:9] ) + ' ' + str( wNroRadicacion[9:12] )  + str( wNroRadicacion[16:21] ) ) #11001 40201703 05200942
            ,'tbusq-025' : (str( wNroRadicacion[0:16] )  + '-' + str( wNroRadicacion[16:21] ) + '-' + str( wNroRadicacion[-2:] ) )  #1100131030382022-00292-00
            ,'tbusq-026' : wNroRadicacion[0:9] + '-' + wNroRadicacion[9:12] + '-' + wNroRadicacion[12:21] + '-' + wNroRadicacion[-2:]
            ,'tbusq-027' : (str( wNroRadicacion[12:16] ) + ' - ' + str( wNroRadicacion[16:21] ))                                                    #2022 - 00896
        } 
        
        return jsonBusqueda
    #fnDatosBusqueda

    def fnDatosBusquedaPdf( self, wNroRadicacion ): 
        #11001400304320220089600
        jsonBusqueda = { 
             'tbusq-1' : wNroRadicacion                                                                                                                                 #11001400304320220089600
            ,'tbusq-2' : (str( wNroRadicacion[0:12] ) + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ) + '-' + str( wNroRadicacion[-2:] ) )    #110014003043-2022-00896-00
            ,'tbusq-3' : (str( wNroRadicacion[0:12] ) + ' ' + str( wNroRadicacion[12:16] ) + ' ' + str( wNroRadicacion[16:21] ) + ' ' + str( wNroRadicacion[-2:] ) )    #110014003043 2022 00896 00
            ,'tbusq-4' : (str( wNroRadicacion[0:5] ) + '-' + str( wNroRadicacion[5:7] ) + '-' + str( wNroRadicacion[7:9] )  + '-' + str( wNroRadicacion[9:12] )  + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ) + '-' + str( wNroRadicacion[-2:] ) )    #11001-40-03-043-2022-00896-00
            ,'tbusq-5' : (str( wNroRadicacion[0:5] ) + ' ' + str( wNroRadicacion[5:7] ) + ' ' + str( wNroRadicacion[7:9] )  + ' ' + str( wNroRadicacion[9:12] )  + ' ' + str( wNroRadicacion[12:16] ) + ' ' + str( wNroRadicacion[16:21] ) )    #11001 40 03 043 2022 00896
            ,'tbusq-6' : (str( wNroRadicacion[9:12] )  + '-' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] ) + '-' + str( wNroRadicacion[-2:] ))    #043-2022-00896-00
            ,'tbusq-7' : (str( wNroRadicacion[0:12] ) + ' ' + str( wNroRadicacion[12:16] ) + '-' + str( wNroRadicacion[16:21] )  )                                      #110014003043 2022-00896
        } 
        
        return jsonBusqueda
    #fnDatosBusquedaPdf

    def fnRetornarCodigoIncidencia( self, modelo, cadena ):
        if modelo == 1: 
            if len( cadena.split('-') ) == 2: #{YYYY}-{5d}
                anio,nro2 = cadena.split('-')
                nro2 = nro2.zfill(5)
                return str(anio)+'-'+str(nro2)
            #fin if
            
            if len( cadena.split('-') ) == 3: #{d}-{YYYY}-{5d}
                nro1,anio,nro2 = cadena.split('-')
                nro1 = nro1.zfill(3)
                nro2 = nro2.zfill(5)
                return str(nro1)+'-'+str(anio)+'-'+str(nro2)
            #fin if 
        #fin if 
    #fin fnRetornarCodigoIncidencia
    
    def fnSeleccionarMesv2( self, driver, datos, ruta_xpath, dar_click = True):
        try:  
            # Variables
            posicion_li = -1
            elemento_li = False 
            
            # Moviendo scroll al título para evitar errores al hacer clic en el mes
            driver.find_element(By.XPATH, '/html/body/div[2]/div[1]/div[2]').location_once_scrolled_into_view

            # Retorna el elemento si encuentra el mes
            cont = 1
            for li in driver.find_elements(By.XPATH, ruta_xpath):
                if str(li.text).strip().split(' ')[0].upper().replace('JULIIO', 'JULIO') == str(datos['Mes']).upper() or str(datos['Mes']).upper() in str(li.text).strip():
                    posicion_li = cont
                    elemento_li = li
                #fin if
                cont+=1
            #fin for 
            
            if not elemento_li:
                self.app_general.fnImprimir(f"No se encontró el [Mes: {str(datos['Mes']).upper()}]", 9)
                self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ ERROR: No se encontró el [Mes: {str(datos['Mes']).upper()}]" )
                return True,posicion_li
            #fin if 

            # Seleccionamos la pestaña
            if dar_click:
                elemento_li.click() 
        except:
            self.app_general.fnImprimir(f"Ocurrió un inconveniente con la página al seleccionar la pestania.", 9)
            self.app_log.error(f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ ERROR: {str(sys.exc_info()[1])}")
            return True,posicion_li
        #fin try
        
        return False,posicion_li
    #fnSeleccionarMesV2
    

    def fnSeleccionarMes(self, dv, data_registro, ruta_xpath=None):
        try:
            if ruta_xpath is None:
                ruta_xpath = self.xpath_pest_mes
            #fin if 
            
            # Moviendo scroll al título para evitar errores al hacer clic en el mes
            dv.find_element(By.XPATH, '/html/body/div[2]/div[1]/div[2]').location_once_scrolled_into_view

            # Retorna el elemento si encuentra el mes
            wArrElementoMes = [
                                li 
                                for li in dv.find_elements(By.XPATH, ruta_xpath) 
                                if str(li.text).strip().split(' ')[0].upper().replace('JULIIO', 'JULIO') == data_registro['Mes'].upper()
                            ]
            
            if len(wArrElementoMes) == 0:
                self.app_general.fnImprimir(f"No se encontró el [Mes: {data_registro['Mes'].upper()}]", 9)
                self.app_log.error( f"NRO RADICACION: {data_registro['NumeroRadicacion']} - MSJ ERROR: No se encontró el [Mes: {data_registro['Mes'].upper()}]" )
                return True
            #fin if 

            # Seleccionamos la pestaña
            wArrElementoMes[0].click() 
        except:
            self.app_general.fnImprimir(f"Ocurrió un inconveniente con la página", 9)
            self.app_log.error(f"NRO RADICACION: {data_registro['NumeroRadicacion']} - MSJ ERROR: {str(sys.exc_info()[1])}")
            return True
        #fin try
        
        return False
    #fnSeleccionarMes
    
    def fnObtenerXpathClick( self, dv, click_estado, click_provid, ruta_xpath, tipo_col = 'td' ): 
        xpath_estado, xpath_provid  = '',''
        for tr in dv.find_elements(By.XPATH, ruta_xpath)[:2]:
            cont = 1
            for td in tr.find_elements(By.TAG_NAME,tipo_col):
                
                if click_estado != '':
                    if click_estado.split('-')[1] != '00':
                        if len(click_estado.split('-')) > 2:
                            xpath_estado = tipo_col+'['+str(click_estado.split('-')[2])+']'
                        else:
                            if str(td.text).strip() == click_estado.split('-')[0]:
                                xpath_estado = tipo_col+'['+str(cont)+']'
                            #fin if
                        #fin if 
                    #fin if 
                #fin if 
                
                if click_provid != '':
                    if click_provid.split('-')[1] != '00':
                        if len(click_provid.split('-')) > 2:
                            xpath_provid = tipo_col+'['+str(click_provid.split('-')[2])+']'
                        else:
                            if click_provid.split('-')[0] in str(td.text).strip() :
                                xpath_provid = tipo_col+'['+str(cont)+']'
                            #fin if
                        #fin if 
                    #fin if 
                #fin if
                
                cont+=1 
            #fin for
            
            if xpath_estado != '' or xpath_provid != '':
                break
            #fin if 
        #fin for
        return xpath_estado, xpath_provid
    #fnObtenerXpathClick
    
    def fnSeleccionarPdfZip(self, wNroDocumento, data_registro, json_tbusqueda ):
        wExiste      = False
        wRutaZip     = self.app_config.wRutaDescPdfFinZip + wNroDocumento + '.zip'
        wRutaCarpeta = self.app_config.wRutaDescPdfFinZip+wNroDocumento+'\\'
        
        if self.app_general.fnExisteArchivo( wRutaZip):
                
            with zipfile.ZipFile(wRutaZip, "r") as zip_ref:
                zip_ref.extractall(wRutaCarpeta)
            #fin with
            
            for directorio_raiz, directorios, archivos in os.walk(wRutaCarpeta):
                for archivo in archivos:
                    if archivo.lower().endswith('.pdf'):
                        ruta_archivo = os.path.join(directorio_raiz, archivo)
                        
                        for clave, valor in json_tbusqueda.items():
                            if valor in archivo.split(' ')[0].strip():
                                wExiste = True
                                self.app_general.fnMoverArchivo( ruta_archivo, (self.app_config.wRutaDescPdfFin + data_registro['NumeroRadicacion'] + '.pdf') )
                                time.sleep(3)
                            #fin if 
                        #fin for 
                        
                        if wExiste:
                            break
                        #fin if 
                    #fin if
                #fin for
            #fin for  
            
            self.app_general.fnEliminarArchivo( wRutaZip )
            self.app_general.fnEliminarCarpeta( wRutaCarpeta )
        #fin if
        
        return wExiste
    #fnSeleccionarPdfZip
    
    def fnSeleccionandoOneDrive( self, dv, datos, wNroDocumento, json_tbusqueda, tipo_dcto = 'EE-' ):
        # Cambiando de pestaña, 
        wTiempo = 0 
        wExiste = False
        wTipoDescarga = 'Zip'
        while wTiempo <= 180:
            try:
                if len(dv.window_handles)>1:
                    dv.switch_to.window(dv.window_handles[1])
                    
                    if len(dv.find_elements(By.XPATH,"//span[text()='Descargar']")) == 1:
                        time.sleep(1)
                        dv.find_elements(By.XPATH,"//span[text()='Descargar']")[0].click() 
                        
                        # Validando tipo de documento
                        if len(dv.find_elements(By.XPATH,"//span[text()='Información']")) == 1:
                            wTipoDescarga = 'Pdf' 
                        #fin if 
                        
                        wExiste = True 
                        break
                    #fin if 
                #fin if
            except:
                pass
            finally:
                time.sleep(1)
                wTiempo+=1
            #fin try
        #fin while
        
        if not wExiste:
            self.app_general.fnImprimir(f"Excedio el tiempo de espera de OneDrive", 13)
            self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ ERROR: Excedio el tiempo de espera de OneDrive")
            return wExiste 
        #fin if  
        
        if wTipoDescarga == 'Zip':
            wExisteZip = wExiste = self.fnValidandoExisteDctoZip( wNroDocumento )
            time.sleep(2)
            
            # Extraer Zip
            if wExisteZip:
                wExiste = self.fnSeleccionarPdfZip( wNroDocumento, datos, json_tbusqueda ) 
            #fin if 
        elif wTipoDescarga == 'Pdf':
            wExiste = self.fnValidandoExisteDcto( wNroDocumento, '', '', tipo_dcto )
        #fin if 

        dv.close()
        dv.switch_to.window(dv.window_handles[0]) 
        
        return wExiste, wTipoDescarga
    #fnSeleccionandoOneDrive 
    
    def fnSeleccionandoOneDriveEstado( self, dv, datos, wNroDocumento ):
        # Cambiando de pestaña, 
        wTiempo = 0 
        wExiste = False
        wTipoDescarga = 'Zip'
        while wTiempo <= 180:
            try:
                if len(dv.window_handles)>1:
                    dv.switch_to.window(dv.window_handles[1])
                    
                    if len(dv.find_elements(By.XPATH,"//span[text()='Descargar']")) == 1:
                        time.sleep(1)
                        dv.find_elements(By.XPATH,"//span[text()='Descargar']")[0].click() 
                        
                        # Validando tipo de documento
                        if len(dv.find_elements(By.XPATH,"//span[text()='Información']")) == 1:
                            wTipoDescarga = 'Pdf'
                        #fin if 
                        
                        wExiste = True 
                        break
                    #fin if 
                #fin if
            except:
                pass
            finally:
                time.sleep(1)
                wTiempo+=1
            #fin try
        #fin while
        
        if not wExiste:
            self.app_general.fnImprimir(f"Excedio el tiempo de espera de OneDrive", 13)
            self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ ERROR: Excedio el tiempo de espera de OneDrive")
            return wExiste 
        #fin if  
        
        if wTipoDescarga == 'Zip':
            wExisteZip = wExiste = self.fnValidandoExisteDctoZip( wNroDocumento )
            time.sleep(2)
            
            # Extraer Zip
            if wExisteZip:
                wExiste = self.fnSeleccionarPdfZip( wNroDocumento, datos, json_tbusqueda ) 
            #fin if 
        elif wTipoDescarga == 'Pdf':
            wExiste = self.fnValidandoExisteDcto( wNroDocumento, '' ,'' , 'ES-' )
        #fin if 

        dv.close()
        dv.switch_to.window(dv.window_handles[0]) 
        
        return wExiste
    #fnSeleccionandoOneDrive 
            
    def fnTipoBusquedaFecha( self, data_registro ):
        wArrFechaBusqueda = []
        wFechaIniTermino = self.app_general.fnConvertirFecha(data_registro['FechaIniciaTermino'], '-', '/')  
        
        wFechaIniTerminov2 = wFechaIniTermino.replace( wFechaIniTermino.split('/')[0] + '/', str(int( wFechaIniTermino.split('/')[0])) + '/' )
        wFechaIniTerminov3 = wFechaIniTerminov2.replace( wFechaIniTerminov2.split('/')[2], wFechaIniTerminov2.split('/')[2][-3:] )
        
        wFechaIniTerminov4 = wFechaIniTermino.replace( wFechaIniTermino.split('/')[2], wFechaIniTermino.split('/')[2][-3:] )
        
        wFechaTexto   = str(self.app_general.fnConvertirFechaTextoMes(wFechaIniTermino)).lower() 
        wFechaTextov2   = wFechaTexto.replace( wFechaTexto.split(' ')[0]+' ', str(int(wFechaTexto.split(' ')[0])) + ' ' )
        
        wMesAnio      = wFechaTexto.replace(' ' + data_registro['Anio'], '') 
        wMesAniov2    = wMesAnio.replace( wMesAnio.split(' ')[0]+' ', str(int(wMesAnio.split(' ')[0])) + ' ' )
        
        wMesAnioGuion = wMesAnio.replace(' ', '-') 
        wMesAnioGuionv2 = wMesAniov2.replace(' ', '-') 
        
        wMesInicio    = wFechaTexto.split(' ')[1] + ' ' + wFechaTexto.split(' ')[0] + ' ' + wFechaTexto.split(' ')[2]
        
        
        # Obtener el nombre del dia 
        fecha_objeto = datetime.datetime.strptime(data_registro['FechaIniciaTermino'], '%Y-%m-%d')
        nombre_del_dia = fecha_objeto.strftime("%A").encode('latin-1').decode('utf-8')
        wDiaMasFechaTexto = nombre_del_dia + ',' + wFechaTexto
        
        wArrFechaBusqueda = [
            wFechaIniTermino           # 01/11/2023
            ,wFechaIniTerminov2        #  1/11/2023
            ,wFechaIniTerminov3        #  1/11/023
            ,wFechaIniTerminov4        # 01/11/023
            ,wFechaIniTermino.replace('/','-') #  01-11-2023
            ,wFechaIniTerminov2.replace('/','-') #  1-11-2023
            ,wFechaTexto               # 01 noviembre 2023
            ,wFechaTextov2             #  1 noviembre 2023
            ,wMesAnio                  # 01 noviembre
            ,wMesAniov2                #  1 noviembre
            ,wMesAnioGuion             # 01-noviembre
            ,wMesAnioGuionv2           #  1-noviembre
            ,wMesInicio                # noviembre 01 2023
            ,wDiaMasFechaTexto         # domingo, 01 enero 2023
        ]
        
        return wArrFechaBusqueda
    #fnTipoBusquedaFecha
    
    def fnValidandoSiPdfEsImagen( self, wNroDocumento ):
        documento = fitz.open( self.app_config.wRutaDescPdfFin + 'ES-' + wNroDocumento + '.pdf' )
        primera_pagina = documento.load_page(0)
        texto_primera_pagina = primera_pagina.get_text("text")
        documento.close()
        
        return False if texto_primera_pagina else True 
    #fin fnValidandoSiPdfEsImagen
    
    def fnDescargarPdfPorFecha(self, dv, data_registro, ruta_xpath, ruta_xpath_columna_busqueda, ruta_xpath_columna_click, ruta_xpath_columna_info, tipo_separador='/', empezar=0, 
                               buscarRadicacionEnPdf = False, wGuardarDrive = False, tipoFnSeleccionarPdf = 1, wDesOneDrive = False ):
        try:
            self.app_general.fnImprimir(f"Buscando [Fecha: {self.app_general.fnConvertirFecha(data_registro['FechaIniciaTermino'], '-', '/')}]", 11)
            # variables 
            wArrFechaBusqueda =  self.fnTipoBusquedaFecha(  data_registro )
            # print(5)
            # obtnemos la posicion de estado y providencia donde se van a presionar click
            click_estado, click_provid  = ruta_xpath_columna_click.split('|')
            # print(6)
            tipo_columna =  'td' if 'td' in ruta_xpath_columna_busqueda else 'th'
            xpath_estado, xpath_provid = self.fnObtenerXpathClick( dv, click_estado, click_provid, ruta_xpath, tipo_columna  ) 
            
            ## buscamos a nivel de fecha.
            # print(7)
            # Almacenar elementos buscados repetidamente
            elementos = dv.find_elements(By.XPATH, ruta_xpath)[empezar:]
            # print(8)
            xpath_busqueda = ruta_xpath_columna_busqueda
            # print(9)
            # Buscar solo una vez el método
            fn_retornar_fecha = self.app_general.fnRetornarFechaLimpia
            # print(10)
            # Lista para almacenar resultados
            arr_elemento_fecha = []
            # print(11)
            # Bucle optimizado
            for tr in elementos:
                # Almacenar el elemento para evitar búsquedas repetitivas
                if xpath_estado!= '':
                    if len(tr.find_elements(By.TAG_NAME, tipo_columna)) < int(xpath_estado.split('[')[-1].replace(']','')):
                        continue
                    #fin if 
                #fin if 
                if xpath_provid!= '':
                    if len(tr.find_elements(By.TAG_NAME, tipo_columna)) < int(xpath_provid.split('[')[-1].replace(']','')):
                        continue
                    #fin if 
                #fin if 
                xpath_click_estado = tr.find_element(By.XPATH, xpath_estado) if xpath_estado != '' else False
                xpath_click_provid = tr.find_element(By.XPATH, xpath_provid) if xpath_provid != '' else False
                xpath_info = tr.find_element(By.XPATH, ruta_xpath_columna_info)
                
                xpath_busqueda_elemento = tr.find_element(By.XPATH, xpath_busqueda)
                
                # Operaciones optimizadas dentro del bucle
                if (
                    len(tr.find_elements(By.XPATH, tipo_columna)) >= int(xpath_busqueda.replace(']','').replace(tipo_columna+'[','')) and 
                    fn_retornar_fecha(xpath_busqueda_elemento.text) in wArrFechaBusqueda
                ):
                    arr_elemento_fecha.append([xpath_click_estado, xpath_click_provid, xpath_info])
                    
                    # break # salir por que ya encontro
                #fin if 
                
                
                if (
                    len(tr.find_elements(By.XPATH, tipo_columna)) >= int(xpath_busqueda.replace(']','').replace(tipo_columna+'[','')) and 
                    fn_retornar_fecha(xpath_busqueda_elemento.text) not in wArrFechaBusqueda and
                    len(arr_elemento_fecha) > 0
                ):
                    break
                #fin if 
                
            #fin for 
            # print(12)
            if len(arr_elemento_fecha) == 0:
                self.app_general.fnImprimir(f"No se encontró la [Fecha: {self.app_general.fnConvertirFecha(data_registro['FechaIniciaTermino'], '-', '/')}]", 13)
                self.app_log.error( f"NRO RADICACION: {data_registro['NumeroRadicacion']} - MSJ ERROR: No se encontró la [Fecha: {self.app_general.fnConvertirFecha(data_registro['FechaIniciaTermino'], '-', tipo_separador)}]")
                return True, ''
            # fin if
            
            # print(13)
            # Bajando scroll
            if xpath_estado != '':
                arr_elemento_fecha[0][0].location_once_scrolled_into_view
            elif xpath_provid != '':
                arr_elemento_fecha[0][1].location_once_scrolled_into_view
            dv.execute_script("window.scrollBy(0, -100);")
            
            # print(14)
            # Presionando para descargar
            wContador = 0
            # print(15)
            json_tbusqueda = self.fnDatosBusqueda( data_registro['NumeroRadicacion'] )

            # print(16)
            #Presionando Descargar estado 
            if xpath_estado != '':
                for elementos in arr_elemento_fecha:
                    if elementos[0]:
                        for a in elementos[0].find_elements(By.TAG_NAME, 'a'):
                            # Presionamos click 
                            if ('PROVIDENCIA' in str(a.text).upper() or 'AUTOS' in str(a.text).upper()) :
                                continue
                            #fin if 
                            
                            a.click()
                            time.sleep(1)
                            
                            wNroDocumento = data_registro['NumeroRadicacion'] 
                            wExiste, wExisteEnlace, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( click_estado.split('-')[1], a, wNroDocumento, json_tbusqueda, data_registro, dv ) 
                             
                            if wExiste == True and wExisteEnlace == True : # si existe enlace va a guardar la providencia
                                #Guardando Estado
                                self.fnGuardarDocumento( ('ES-' + data_registro['NumeroRadicacion'] + '.pdf') , wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive ) 
                                
                                #Guardando Providencia 
                                self.fnGuardarDocumento( ('EE-' + data_registro['NumeroRadicacion'] + '.pdf') , wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive ) 
                                
                            elif wExiste == True and wExisteEnlace == False:
                                self.fnGuardarDocumento( ('ES-' + data_registro['NumeroRadicacion'] + '.pdf') , wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )                                
                            else:
                                if click_estado.split('-')[1] != '01':
                                    print()
                                #fin if 
                                self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {data_registro['NumeroRadicacion']} - en el documento pdf.",13)
                                self.app_log.error( f"No se encontro el Nro Radicación: {data_registro['NumeroRadicacion']} - en el documento pdf.")
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + "ES-" + wNroDocumento + '.pdf' ):
                                    self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + "ES-" + wNroDocumento + '.pdf' )
                                #fin if 
                            #fin if
 
                            wContador+=1 
                        #fin if 
                    #fin if 
                #fin for
            #fin if 
            
            #Presionando Descargar Providencia 
            if xpath_provid != '':
                for elementos in arr_elemento_fecha:
                    if elementos[1]:
                        for a in elementos[1].find_elements(By.TAG_NAME, 'a'):
                            # Presionamos click    
                            
                            if ('ESTADO' in str(a.text).upper() and 'PROVIDENCIA' not in str(a.text).upper()  ) or 'VER COMPRIMIDO' in str(a.text).upper() :
                                continue
                            #fin if 
                            
                            a.click()
                            time.sleep(1)
                            
                            # wNroDocumento = data_registro['NumeroRadicacion'] + ( '_'+str(wContador) if wContador > 0 else '' )   
                            wNroDocumento = data_registro['NumeroRadicacion'] 
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( dv, '01', a, data_registro['NumeroRadicacion'], json_tbusqueda, data_registro ) 
                              
                            if wExiste : 
                                if wGuardarDrive:
                                    self.fnGuardarDocumento( ('EE-' + data_registro['NumeroRadicacion'] + '.pdf') , wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )
                                #fin if 
                                
                                break
                            else:
                                self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {data_registro['NumeroRadicacion']} - en el documento pdf.",13)
                                self.app_log.error( f"No se encontro el Nro Radicación: {data_registro['NumeroRadicacion']} - en el documento pdf.")
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + 'EE-' + wNroDocumento + '.pdf' ):
                                    self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + 'EE-' + wNroDocumento + '.pdf' )
                                #fin if 
                            #fin if
                             
                            wContador+=1 
                        #fin if 
                    #fin if 
                #fin for
            #fin if 
             
            
        except:
            self.app_general.fnImprimir(f"Ocurrió un inconveniente con la página", 13)
            self.app_log.error( f"NRO RADICACION: {data_registro['NumeroRadicacion']} - FN: fnDescargarPdfPorFecha - MSJ ERROR: {str(sys.exc_info()[1])}]")
            return True, ''
        #fin if 

        return False, ''
    #fnDescargarPdfPorFecha
    
    def fnDescargarPdfPorFechaProc24( self, dv, data_registro, ruta_xpath, xpath_busqueda, xpath_estado_provid, modelo_dctos = '00|00', tipo_separador='/', empezar=0, wGuardarDrive = False  ):
        try:
            self.app_general.fnImprimir(f"Buscando [Fecha: {self.app_general.fnConvertirFecha(data_registro['FechaIniciaTermino'], '-', '/')}]", 11)
            # variables 
            wArrFechaBusqueda =  self.fnTipoBusquedaFecha(  data_registro )  
            
            ## buscamos a nivel de fecha. 
            # Almacenar elementos buscados repetidamente
            elementos = dv.find_elements(By.XPATH, ruta_xpath)[empezar:]  
            fn_retornar_fecha = self.app_general.fnRetornarFechaLimpia 
            
            # Lista para almacenar resultados
            arr_elemento_fecha = []
            
            # Bucle optimizado
            for tr in elementos:
                # Almacenar el elemento para evitar búsquedas repetitivas
                if len(tr.find_elements(By.TAG_NAME, 'td')) < int(xpath_estado_provid.split('[')[-1].replace(']','')):
                    continue
                #fin if  
                
                xpath_click_estprovid   = tr.find_element(By.XPATH, xpath_estado_provid) 
                xpath_busqueda_elemento = tr.find_element(By.XPATH, xpath_busqueda)
                
                # Operaciones optimizadas dentro del bucle
                if (
                    len(tr.find_elements(By.XPATH, 'td')) >= int(xpath_busqueda.replace(']','').replace('td[','')) and 
                    fn_retornar_fecha(xpath_busqueda_elemento.text) in wArrFechaBusqueda
                ):
                    arr_elemento_fecha.append(xpath_click_estprovid)
                #fin if 
                
                if (
                    len(tr.find_elements(By.XPATH, 'td')) >= int(xpath_busqueda.replace(']','').replace('td[','')) and 
                    fn_retornar_fecha(xpath_busqueda_elemento.text) not in wArrFechaBusqueda and
                    len(arr_elemento_fecha) > 0
                ):
                    break
                #fin if 
            #fin for 
             
            if len(arr_elemento_fecha) == 0:
                self.app_general.fnImprimir(f"No se encontró la [Fecha: {self.app_general.fnConvertirFecha(data_registro['FechaIniciaTermino'], '-', '/')}]", 13)
                self.app_log.error( f"NRO RADICACION: {data_registro['NumeroRadicacion']} - MSJ ERROR: No se encontró la [Fecha: {self.app_general.fnConvertirFecha(data_registro['FechaIniciaTermino'], '-', tipo_separador)}]")
                return True, ''
            # fin if
             
            # Bajando scroll
            arr_elemento_fecha[0].location_once_scrolled_into_view
            dv.execute_script("window.scrollBy(0, -100);")
            time.sleep(0.5)
            
            # Presionando para descargar
            wContador      = 0
            wExiste        = False 
            wRutaDctoPdf   = self.app_config.wRutaDescPdfFin
            json_tbusqueda = self.fnDatosBusqueda( data_registro['NumeroRadicacion'] )
            
            for elementos in arr_elemento_fecha:
                for a in elementos.find_elements(By.TAG_NAME, 'a'):
                    a.click()
                    time.sleep(0.5)
                    
                    if 'estado' in a.get_attribute('href'):
                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( modelo_dctos.split('-')[0], a, data_registro['NumeroRadicacion'], json_tbusqueda, data_registro, dv )
                    else:
                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( dv, modelo_dctos.split('-')[1], a, data_registro['NumeroRadicacion'], json_tbusqueda, data_registro ) 
                    #fin if 
                    
                    if wExiste : 
                        if wGuardarDrive:
                            if wRutaDctoPdf != '':
                                self.app_general.fnCopiarArchivo( wRutaDctoPdf + data_registro['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + data_registro['NumeroRadicacion'] + '.pdf' )
                                
                                while True:
                                    if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + data_registro['NumeroRadicacion'] + '.pdf'  ):
                                        self.app_general.fnEliminarArchivo( wRutaDctoPdf + data_registro['NumeroRadicacion'] + '.pdf' )
                                        self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                        break
                                    #fin if 
                                    time.sleep(1)
                                #fin while
                            #fin if 
                        #fin if 
                        break
                    else:
                        self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {data_registro['NumeroRadicacion']} - en el documento pdf.",13)
                        self.app_log.error( f"No se encontro el Nro Radicación: {data_registro['NumeroRadicacion']} - en el documento pdf.")
                        if self.app_general.fnExisteArchivo( wRutaDctoPdf + data_registro['NumeroRadicacion'] + '.pdf' ):
                            self.app_general.fnEliminarArchivo( wRutaDctoPdf + data_registro['NumeroRadicacion'] + '.pdf' )
                        #fin if 
                    #fin if
                #fin for
                
                if wExiste:
                    break
                #fin if 
            #fin for   
        except:
            self.app_general.fnImprimir(f"Ocurrió un inconveniente con la página", 13)
            self.app_log.error( f"NRO RADICACION: {data_registro['NumeroRadicacion']} - FN: fnDescargarPdfPorFecha - MSJ ERROR: {str(sys.exc_info()[1])}]")
            return True, ''
        #fin if 

        return False, ''
    #fnDescargarPdfPorFechaProc24
    
    def fnValidandoExisteDcto( self, wNroRadicacion, wRutaDescargaPdfIniTmp = '', wRutaDescargaPdfDesTmp = '', wTipoDcto = 'EE-' ):
        wRutaDescargaPdfIni = wRutaDescargaPdfIniTmp if wRutaDescargaPdfIniTmp != '' else self.app_config.wRutaDescPdf
        wRutaDescargaPdfDes = wRutaDescargaPdfDesTmp if wRutaDescargaPdfDesTmp != '' else self.app_config.wRutaDescPdfFin
        wTipoDocumento      = 'EE-' if wTipoDcto == '' else wTipoDcto
        self.app_general.fnImprimir(f"Esperando descarga del documento",11)
        wExiste = False
        wTiempo = 0
        while wTiempo<=300:
            try:
                array_pdf = [ archivo for archivo in os.listdir(wRutaDescargaPdfIni) if archivo.lower().endswith('.pdf') ]
                if len(array_pdf) > 0:
                    self.app_general.fnImprimir(f"Se descargo exitosamente el archivo",13)
                    
                    # moviendo archivo
                    self.app_general.fnMoverArchivo( (wRutaDescargaPdfIni + array_pdf[0]), (wRutaDescargaPdfDes + wTipoDocumento + wNroRadicacion+'.pdf') )
                    wExiste = True
                    time.sleep(5)
                    break
                #fin if
            except:
                pass
            finally:
                time.sleep(1)
                wTiempo+=1
            #fin try
        #fin while 
        
        return wExiste
    #fnValidandoExisteDcto 
    
    def fnValidandoExisteDctoXlsx( self, wNroRadicacion, wRutaDescargaPdfIniTmp = '', wRutaDescargaPdfDesTmp = '' ):
        wRutaDescargaPdfIni = wRutaDescargaPdfIniTmp if wRutaDescargaPdfIniTmp != '' else self.app_config.wRutaDescPdf
        wRutaDescargaPdfDes = wRutaDescargaPdfDesTmp if wRutaDescargaPdfDesTmp != '' else self.app_config.wRutaDescPdfFin
        self.app_general.fnImprimir(f"Esperando descarga del documento",11)
        wExiste = False
        wTiempo = 0
        while wTiempo<=300:
            try:
                array_pdf = [ archivo for archivo in os.listdir(wRutaDescargaPdfIni) if archivo.lower().endswith('.xlsx') ]
                if len(array_pdf) > 0:
                    self.app_general.fnImprimir(f"Se descargo exitosamente el archivo",13)
                    
                    # moviendo archivo
                    self.app_general.fnMoverArchivo( (wRutaDescargaPdfIni + array_pdf[0]), (wRutaDescargaPdfDes + wNroRadicacion+'.xlsx') )
                    wExiste = True
                    time.sleep(5)
                    break
                #fin if
            except:
                pass
            finally:
                time.sleep(1)
                wTiempo+=1
            #fin try
        #fin while 
        
        return wExiste
    #fnValidandoExisteDctoXlsx 
    
    def fnValidandoExisteDctoZip( self, wNroRadicacion, wRutaDescargaPdfIniTmp = '', wRutaDescargaPdfDesTmp = '' ):
        wRutaDescargaPdfIni = wRutaDescargaPdfIniTmp if wRutaDescargaPdfIniTmp != '' else self.app_config.wRutaDescPdf
        wRutaDescargaPdfDes = wRutaDescargaPdfDesTmp if wRutaDescargaPdfDesTmp != '' else self.app_config.wRutaDescPdfFinZip
        self.app_general.fnImprimir(f"Esperando descarga del documento",11)
        wTiempo = 0
        wExiste = False
        while wTiempo<=300:
            try:
                array_pdf = [ archivo for archivo in os.listdir(wRutaDescargaPdfIni) if archivo.lower().endswith('.zip') ]
                if len(array_pdf) > 0:
                    self.app_general.fnImprimir(f"Se descargo exitosamente el archivo",13)
                    wExiste = True
                    # moviendo archivo
                    self.app_general.fnMoverArchivo( (wRutaDescargaPdfIni + array_pdf[0]), (wRutaDescargaPdfDes + wNroRadicacion+'.zip') )
                    time.sleep(5)
                    break
                #fin if
            except:
                pass
            finally:
                time.sleep(1)
                wTiempo+=1
            #fin try
        #fin while 
        
        return wExiste
    #fnValidandoExisteDctoZip
    
    def fnRetornarRadicadoBusqueda( self, wNroRadicacion ):
        wArreglo = [ wNroRadicacion ]
        wArreglo.append( wNroRadicacion[:10] + ' ' + wNroRadicacion[10:12] + ' ' + wNroRadicacion[12:16] + ' ' + wNroRadicacion[16:21] + ' ' + wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[:10] + '-' + wNroRadicacion[10:12] + '-' + wNroRadicacion[12:16] + '-' + wNroRadicacion[16:21] + '-' + wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[0:12] + ' ' + wNroRadicacion[12:16]+ '-' + wNroRadicacion[16:] )
        wArreglo.append( wNroRadicacion[0:12] + '-' + wNroRadicacion[12:16]+ '-' + wNroRadicacion[16:] )
        wArreglo.append( wNroRadicacion[0:12] + ' ' + wNroRadicacion[12:16]+ '-' + wNroRadicacion[16:21]+'-'+wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[0:12] + '-' + wNroRadicacion[12:16]+ '-' + wNroRadicacion[16:21]+'-'+wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[0:12] + '-' + wNroRadicacion[12:16]+ '-' + wNroRadicacion[16:21]+' '+wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[0:12] + ' ' + wNroRadicacion[12:16]+ ' ' + wNroRadicacion[16:21]+' '+wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[12:16]+ '-' + wNroRadicacion[16:21]+'-'+wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[12:16]+ '-' + wNroRadicacion[16:21]+' '+wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[12:16]+ '-' + wNroRadicacion[16:21] )
        wArreglo.append( wNroRadicacion[12:16]+ '-' + wNroRadicacion[17:21] )
        wArreglo.append( wNroRadicacion[0:2]  + ' ' + wNroRadicacion[2:5] +  ' ' + wNroRadicacion[5:7] +  ' ' + wNroRadicacion[7:9] +  ' ' + wNroRadicacion[9:12] +  ' ' + wNroRadicacion[12:16] +  ' ' + wNroRadicacion[16:21] +  ' ' + wNroRadicacion[21:23]  )
        wArreglo.append( wNroRadicacion[0:2]  + ' ' + wNroRadicacion[2:5] +  ' ' + wNroRadicacion[5:7] +  ' ' + wNroRadicacion[7:9] +  ' ' + wNroRadicacion[9:12] +  ' ' + wNroRadicacion[12:16] +  ' ' + wNroRadicacion[16:19] +  ' ' + wNroRadicacion[19:21] +  ' ' + wNroRadicacion[21:23]  )
        wArreglo.append( wNroRadicacion[0:2]  + ' ' + wNroRadicacion[2:5] +  '-' + wNroRadicacion[5:7] +  '-' + wNroRadicacion[7:9] +  '-' + wNroRadicacion[9:12] +  '-' + wNroRadicacion[12:16] +  '-' + wNroRadicacion[16:21] +  '-' + wNroRadicacion[21:23]  )
        wArreglo.append( wNroRadicacion[0:2]  + '-' + wNroRadicacion[2:5] +  '-' + wNroRadicacion[5:7] +  '-' + wNroRadicacion[7:9] +  '-' + wNroRadicacion[9:12] +  '-' + wNroRadicacion[12:16] +  '-' + wNroRadicacion[16:21] +  '-' + wNroRadicacion[21:23]  )
        wArreglo.append( wNroRadicacion[:5] + ' ' + wNroRadicacion[5:7] + ' ' + wNroRadicacion[7:9] + ' ' + wNroRadicacion[9:16] + ' ' + wNroRadicacion[16:21] + ' ' + wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[:5] + ' ' + wNroRadicacion[5:7] + ' ' + wNroRadicacion[7:9] + ' ' + wNroRadicacion[9:12] + ' ' + wNroRadicacion[12:16] + ' ' + wNroRadicacion[16:21]   )
        wArreglo.append( wNroRadicacion[:5] + ' ' + wNroRadicacion[5:7] + ' ' + wNroRadicacion[7:9] + ' ' + wNroRadicacion[9:12] + ' ' + wNroRadicacion[12:16] + ' ' + wNroRadicacion[16:21] + ' ' + wNroRadicacion[-2:]  )
        wArreglo.append( wNroRadicacion[:5] + '-' + wNroRadicacion[5:7] + '-' + wNroRadicacion[7:9] + '-' + wNroRadicacion[9:16] + '-' + wNroRadicacion[16:21] + '-' + wNroRadicacion[-2:] )
        wArreglo.append( wNroRadicacion[:5] + '-' + wNroRadicacion[5:7] + '-' + wNroRadicacion[7:9] + '-' + wNroRadicacion[9:12] + '-' + wNroRadicacion[12:16] + '-' + wNroRadicacion[16:21]   )
        wArreglo.append( wNroRadicacion[:5] + '-' + wNroRadicacion[5:7] + '-' + wNroRadicacion[7:9] + '-' + wNroRadicacion[9:12] + '-' + wNroRadicacion[12:16] + '-' + wNroRadicacion[16:21] + '-' + wNroRadicacion[-2:]  )
        wArreglo.append( wNroRadicacion[0:12]+' '+wNroRadicacion[12:21] )
        wArreglo.append( wNroRadicacion[0:12]+' '+wNroRadicacion[12:16]+' '+wNroRadicacion[16:23] )
        wArreglo.append( wNroRadicacion[0:5] + ' ' + wNroRadicacion[5:21] )
        wArreglo.append( wNroRadicacion[12:16]+' '+wNroRadicacion[17:21] )
        wArreglo.append( wNroRadicacion[12:14]+' '+wNroRadicacion[14:16]+' '+wNroRadicacion[17:21] )
        return wArreglo
    #fnRetornarRadicadoBusqueda

    def fnSeleccionarPdf( self, wNombreDocumento ):
        # variables
        wExiste = False
        wNroRadicacion    = str(wNroRadicacion).split('_')[0]
        wRutaArchivo      = self.app_config.wRutaDescPdfFin + wNombreDocumento + '.pdf'
        wRutaArchivoFinal = self.app_config.wRutaDescPdfSel + wNroRadicacion + '.pdf'
        wRutaArchivoDrive = self.app_config.wRutaAlmPdfDrive + wNroRadicacion + '.pdf'
        
        if self.app_general.fnExisteArchivo( wRutaArchivo ): 
            wArrDiffNroRadicacion = self.fnRetornarRadicadoBusqueda( wNroRadicacion )
            
            # Abre el archivo original y crea un escritor para el nuevo PDF
            with open(wRutaArchivo, 'rb') as file:
                pdf_reader = PdfReader(file)
                pdf_writer = PdfWriter()
                num_pages  = len(pdf_reader.pages)
                wNroRadicacionEncontrado = ''

                # Crear un nuevo PDF con las páginas deseadas
                for page_num in range(num_pages):
                    page = pdf_reader.pages[page_num]
                    
                    for buscar in wArrDiffNroRadicacion:
                        if buscar in page.extract_text().replace('  ',' '):
                            pdf_writer.add_page(page)
                            wExiste = True
                            wNroRadicacionEncontrado = buscar
                            break
                        #fin if
                    
                    if wExiste: 
                        if 'Rad. No.' in page.extract_text().replace('  ',' ') :
                            if wNroRadicacionEncontrado not in page.extract_text().replace('  ',' '):
                                break
                            #fin if 
                        else:
                            pdf_writer.add_page(page)
                        #fin if
                    #fin if 
                #fin for 

                # Guardar el nuevo PDF si se encontraron páginas que cumplen con el criterio
                if wExiste:
                    with open(wRutaArchivoFinal, 'wb') as new_file:
                        pdf_writer.write(new_file)
                    #fin with
                    
                    self.app_general.fnCopiarArchivo( wRutaArchivoFinal, wRutaArchivoDrive)
                else:
                    self.app_general.fnImprimir("No se encontraron páginas que cumplan con el criterio especificado.")
                #fin if
            #fin with
        else:
            self.app_general.fnImprimir("No se pudo encontrar el archivo")
        #fin if
        return wExiste
    #fnSeleccionarPdf
       
    def fnSeleccionarPdfv2( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo   = self.app_config.wRutaDescPdfFin + wNombreDocumento + '.pdf'
        wNroRadicacion = str(wNombreDocumento).split('_')[0]
        wExiste        = False
        self.app_general.fnImprimir("Leyendo Pdf ",11)
        wDatoBusqueda = wNroRadicacion[0:7]+' '+wNroRadicacion[7:9]+' '+wNroRadicacion[9:12]+' \n'+wNroRadicacion[12:16]+ ' '+wNroRadicacion[16:21]
        wDatoBusqueda2 = wNroRadicacion[0:5]+' '+wNroRadicacion[5:7]+' '+wNroRadicacion[7:9]+' '+wNroRadicacion[9:12]+' \n'+wNroRadicacion[12:16]+ ' '+wNroRadicacion[16:21]
        texto_pdf = extract_text(wRutaArchivo)
        
        if wDatoBusqueda in str(texto_pdf).replace("  "," ").replace("  "," ") or wDatoBusqueda2 in str(texto_pdf).replace("  "," ").replace("  "," "):
            wExiste = True
        #fin if
        
        ####### Proceso de busqueda de URL
        if wExiste :
            pdf = pdfx.PDFx(self.app_config.wRutaDescPdfFin + wNombreDocumento + '.pdf') 
            #Obtenemos el enlace
            enlacePdf = pdf.get_references_as_dict()
            
            for wEnlace in enlacePdf['url']:
                try:
                    driverPdf, waitPdf = self.app_general.fnAbrirPagina(wEnlace, self.app_config.wRutaDescPdfFinV2 ) 
                    wUrl = driverPdf.current_url 
                    
                    wAnio = wNroRadicacion[12:16]
                    wCodigo = str(int(wNroRadicacion[16:21]))
                    wBuscar =  wAnio + '%2D' + wCodigo 
                    
                    if wBuscar in wUrl:
                        driverPdf.find_element(By.XPATH,'/html/body/div/div/div[2]/div/div/div[2]/div/div/div/div/div/div/div/div/div[1]/div/button/span').click()
                        time.sleep(1.5) 
                        try:
                            driverPdf.find_element(By.XPATH,'/html/body/div[2]/div/div/div/div/div/div/ul/li[1]/button/div/span').click()
                        except:
                            pass
                        time.sleep(0.5)
                        
                        wExiste = True
                        break
                    #fin if
                except:
                    pass
                finally:
                    driverPdf.quit()
                #fin try
            #fin for
             
            self.fnValidandoExisteDcto(  wNroRadicacion, self.app_config.wRutaDescPdfFinV2, self.app_config.wRutaDescPdfSel )  
        #fin if     
        
        return wExiste
    ######## Procesos 
    
    def fnSeleccionarPdfv3( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo   = self.app_config.wRutaDescPdfFin + wNombreDocumento + '.pdf'
        wNroRadicacion = str(wNombreDocumento).split('_')[0]
        wExiste        = False
        self.app_general.fnImprimir("Leyendo Pdf ",11)
        wDatoBusqueda =  wNroRadicacion[12:16]+ '-'+wNroRadicacion[18:21]
        wDatoBusqueda2 = wNroRadicacion[12:16]+ '-'+wNroRadicacion[17:21]
        texto_pdf = extract_text(wRutaArchivo)
        
        if wDatoBusqueda in str(texto_pdf).replace("  "," ").replace("  "," ") or wDatoBusqueda2 in str(texto_pdf).replace("  "," ").replace("  "," "):
            wExiste = True
        #fin if
        
        ####### Proceso de busqueda de URL
        if wExiste :
            pdf = pdfx.PDFx(self.app_config.wRutaDescPdfFin + wNombreDocumento + '.pdf') 
            #Obtenemos el enlace
            enlacePdf = pdf.get_references_as_dict()
            
            for wEnlace in enlacePdf['url']:
                try:
                    driverPdf, waitPdf = self.app_general.fnAbrirPagina(wEnlace, self.app_config.wRutaDescPdfFinV2 ) 
                    wUrl = driverPdf.current_url 
                    
                    wAnio = wNroRadicacion[12:16]
                    wCodigo = str(int(wNroRadicacion[16:21]))
                    wBuscar =  wAnio + '%2D' + wCodigo 
                    
                    if wBuscar in wUrl:
                        driverPdf.find_element(By.XPATH,'/html/body/div/div/div[2]/div/div/div[2]/div/div/div/div/div/div/div/div/div[1]/div/button/span').click()
                        time.sleep(1.5) 
                        try:
                            driverPdf.find_element(By.XPATH,'/html/body/div[2]/div/div/div/div/div/div/ul/li[1]/button/div/span').click()
                        except:
                            pass
                        time.sleep(0.5)
                        
                        wExiste = True
                        break
                    #fin if
                except:
                    pass
                finally:
                    driverPdf.quit()
                #fin try
            #fin for
             
            self.fnValidandoExisteDcto(  wNroRadicacion, self.app_config.wRutaDescPdfFinV2, self.app_config.wRutaDescPdfSel )  
        #fin if     
        
        return wExiste
    #fnSeleccionarPdfv3
    
    def fnSeleccionarPdfv4( self, wNombreDocumento ):
        # variables
        wExiste = False
        wNroRadicacion    = str(wNroRadicacion).split('_')[0]
        wRutaArchivo      = self.app_config.wRutaDescPdfFin + wNombreDocumento + '.pdf'
        wRutaArchivoFinal = self.app_config.wRutaDescPdfSel + wNroRadicacion + '.pdf'
        wRutaArchivoDrive = self.app_config.wRutaAlmPdfDrive + wNroRadicacion + '.pdf'
        
        if self.app_general.fnExisteArchivo( wRutaArchivo ): 
            wArrDiffNroRadicacion = self.fnRetornarRadicadoBusqueda( wNroRadicacion )
            
            # Abre el archivo original y crea un escritor para el nuevo PDF
            with open(wRutaArchivo, 'rb') as file:
                pdf_reader = PdfReader(file)
                pdf_writer = PdfWriter()
                num_pages  = len(pdf_reader.pages)
                wNroRadicacionEncontrado = ''

                # Crear un nuevo PDF con las páginas deseadas
                for page_num in range(num_pages):
                    page = pdf_reader.pages[page_num]
                    
                    for buscar in wArrDiffNroRadicacion:
                        if buscar in page.extract_text().replace('  ',' '):
                            pdf_writer.add_page(page)
                            wExiste = True
                            wNroRadicacionEncontrado = buscar
                            break
                        #fin if
                    
                    if wExiste: 
                        if 'Rad. No.' in page.extract_text().replace('  ',' ') :
                            if wNroRadicacionEncontrado not in page.extract_text().replace('  ',' '):
                                break
                            #fin if 
                        else:
                            pdf_writer.add_page(page)
                        #fin if
                    #fin if 
                #fin for 

                # Guardar el nuevo PDF si se encontraron páginas que cumplen con el criterio
                if wExiste:
                    with open(wRutaArchivoFinal, 'wb') as new_file:
                        pdf_writer.write(new_file)
                    #fin with
                    
                    self.app_general.fnCopiarArchivo( wRutaArchivoFinal, wRutaArchivoDrive)
                else:
                    self.app_general.fnImprimir("No se encontraron páginas que cumplan con el criterio especificado.")
                #fin if
            #fin with
        else:
            self.app_general.fnImprimir("No se pudo encontrar el archivo")
        #fin if
        return wExiste
    #fnSeleccionarPdfv4
    
    
    def fnBuscarEstadoXlsxModelo01( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExiste          = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfSel
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Xlsx ",11)
        libro_trabajo = openpyxl.load_workbook(wRutaArchivo)

        # Obtener las hojas de trabajo (pestañas)
        hojas = libro_trabajo.sheetnames

        # Recorrer cada hoja y obtener su información
        for nombre_hoja in hojas:
            hoja = libro_trabajo[nombre_hoja]
            print(f'Nombre de la hoja: {nombre_hoja}')
            
            # Obtener información de la hoja
            cont = 1
            for fila in hoja.iter_rows(values_only=True):
                try:
                    if cont < 4:
                        continue
                    #fin if 
                    
                    wRadicado = fila[1]
                    if wRadicado in wArrDatoBusqueda:
                        wClase         = fila[2]
                        wDemandante    = fila[4]
                        wDemandado     = fila[5]
                        wFechaAuto     = fila[7].strftime('%Y-%m-%d')
                        wAutoAnotacion = fila[8] 
                        
                        wExiste = True
                        wRutaDctoPdf = ''
                        break
                    #fin if 
                except:
                    pass
                finally:
                    cont+=1
                #fin try
            #fin for
            
            if wExiste:
                break
            #fin if
        #fin for
        return wExiste, wRutaDctoPdf
    ######## fnBuscarEstadoXlsxModelo01 
    
    #" Solo sirve para poder buscar la providencia  
    # "
    
    
    def fnVariableBusquedaUrlPdf(self, wNroRadicacion): 
        wArrDatoBusquedaUrl = []
        wArrDatoBusquedaUrl.append( wNroRadicacion[12:16] + '%2D' + str(int(wNroRadicacion[16:21])) ) # 2021%2D853
        wArrDatoBusquedaUrl.append( wNroRadicacion[12:16] + '-' + str(wNroRadicacion[16:21]) ) # 2021-00853
        wArrDatoBusquedaUrl.append( wNroRadicacion[12:16] + '-' + str(wNroRadicacion[17:21]) ) # 2021-0853
        wArrDatoBusquedaUrl.append( wNroRadicacion[12:16] + '-' + str(int(wNroRadicacion[17:21])) ) # 2021-0853
        
        return wArrDatoBusquedaUrl
    #fnVariableBusquedaUrlPdf
    
    def fnRadicadosBusquedaPdf(self, wNroRadicacion): 
        wArrDatoBusqueda = [] 
        wArrDatoBusqueda.append( wNroRadicacion )                                                                                                                                              #110014003028202100853
        wArrDatoBusqueda.append( wNroRadicacion[0:7]+' '+wNroRadicacion[7:9]+' '+wNroRadicacion[9:12]+' \n'+wNroRadicacion[12:16]+ ' '+wNroRadicacion[16:21] )                                 #1100140 03 028 \n2021 00853
        wArrDatoBusqueda.append( wNroRadicacion[0:7]+' '+wNroRadicacion[7:9]+' '+wNroRadicacion[9:12]+' \n'+wNroRadicacion[12:16]+ ' '+wNroRadicacion[16:21] + ' ' + wNroRadicacion[-2:] )     #1100140 03 028 \n2021 00853 00
        wArrDatoBusqueda.append( wNroRadicacion[0:5]+' '+wNroRadicacion[5:7]+' '+wNroRadicacion[7:9]+' '+wNroRadicacion[9:12]+' \n'+wNroRadicacion[12:16]+ ' '+wNroRadicacion[16:21] )         #11001 40 03 028 \n2021 00853
        wArrDatoBusqueda.append( wNroRadicacion[0:5]+' '+wNroRadicacion[5:7]+' '+wNroRadicacion[7:9]+' '+wNroRadicacion[9:12]+' \n\n'+wNroRadicacion[12:16]+ ' '+wNroRadicacion[16:21] )       #11001 40 03 028 \n\n2022 00962
        wArrDatoBusqueda.append( wNroRadicacion[0:5]+' '+wNroRadicacion[5:7]+' \n'+wNroRadicacion[12:16]+' \n\n'+wNroRadicacion[7:9]+ ' '+wNroRadicacion[9:12]+' \n'+wNroRadicacion[16:21] )   #11001 40 03 028 \n\n2022 00962 
        wArrDatoBusqueda.append( wNroRadicacion[17:21] + wNroRadicacion[14:16] )                                                                                                               #0096222                           
        wArrDatoBusqueda.append( wNroRadicacion[12:16] + wNroRadicacion[16:21] )                                                                                                               #202200962                           
        wArrDatoBusqueda.append( wNroRadicacion[12:16] + '-' + wNroRadicacion[16:21] )                                                                                                         #2022-00962                          
        wArrDatoBusqueda.append( wNroRadicacion[12:16] + '-' + wNroRadicacion[16:21] + '-' + wNroRadicacion[-2:] )                                                                             #2022-00962-00
        wArrDatoBusqueda.append( wNroRadicacion[0:5]+'\n\n'+wNroRadicacion[5:7]+'\n'+wNroRadicacion[12:16]+'\n\n'+wNroRadicacion[7:9]+ '\n'+wNroRadicacion[9:12]+'\n'+wNroRadicacion[16:21] )  #11001\n\n41\n2020\n\n89\n037\n01582
        wArrDatoBusqueda.append( wNroRadicacion[0:5]+' '+wNroRadicacion[5:7]+' '+wNroRadicacion[7:9]+' '+wNroRadicacion[9:12]+'\n'+wNroRadicacion[12:16]+' '+wNroRadicacion[16:21] )           #11001 31 03 031\n2008 00293
        wArrDatoBusqueda.append( (wNroRadicacion[0:12]+' '+wNroRadicacion[12:16] + ' ' + wNroRadicacion[16:21]+' ' + wNroRadicacion[-2:]) )                                                    #110014189035 2021 00412 00
        wArrDatoBusqueda.append( wNroRadicacion[0:5]+'\n'+wNroRadicacion[12:16]+'\n\n'+wNroRadicacion[5:7]+'\n\n'+wNroRadicacion[9:12]+'\n\n'+wNroRadicacion[7:9]+'\n'+wNroRadicacion[16:21] ) #11001\n2023\n\n31\n\n008\n\n03\n00486
        wArrDatoBusqueda.append( wNroRadicacion[0:5] + ' \n'+ wNroRadicacion[12:16] + ' \n\n'+ wNroRadicacion[5:7] + ' ' + wNroRadicacion[7:9] + ' ' + wNroRadicacion[9:12] + ' \n'+ wNroRadicacion[16:21] ) #11001 \n2022 \n\n31  03  030 \n00535
        return wArrDatoBusqueda
    #fnRadicadosBusquedaPdf
    
    def fnPresionandoClickEnlacePdf( self, wNombreDocumento, wNroRadicacion ):
        self.app_general.fnImprimir("Se encontro Nro de Radicación",13)
        llave_enlace = ''
        wExiste = False
        
        #Obtenemos el enlace
        doc = fitz.open(self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf')
        enlacePdfOri = []
        enlacePdfBus = {'url': []}
        for pagina_num in range(doc.page_count):
            pagina = doc[pagina_num]
            enlaces_pagina = pagina.get_links()
            enlacePdfOri.extend(enlaces_pagina)
        #fin for
        doc.close() 

        # pdf = pdfx.PDFx(self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf') 
        # try:
        #     enlacePdf = pdf.get_references_as_dict()
        # finally:
        #     print()
        #     # pdf.close()
        # #fin try
         
        if not enlacePdfOri :
            self.app_general.fnImprimir("No se encontraron enlaces de descargas.",13)
            return False
        else:
            enlacePdfBus['url'] = [ enlace.get('file', '') for enlace in enlacePdfOri  if enlace.get('file', '') not in ('https://procesojudicial.ramajudicial.gov.co/FirmaElectronica','ramajudicial.gov.co')]
            if len(enlacePdfBus['url']) == 0:
                self.app_general.fnImprimir("No se encontraron enlaces de descargas.",13)
                return False
            #fin if   
        #fin if
            
        # Variable
        wArrDatoBusquedaUrl = self.fnVariableBusquedaUrlPdf(wNroRadicacion) 
        
        wExisteProv = False
        wDescargarOneDrive = False
        
        self.app_general.fnImprimir("Oteniendo Pdf de providencias",11)
        for wEnlace in enlacePdfBus['url']:
            try: 
                driverPdf = None
                wExisteDcto = False
                if  wEnlace[:11] == '/documents/':
                    wEnlace = "https://www.ramajudicial.gov.co" + wEnlace
                #fin if
                
                if 'ramajudicial.gov.co' in wEnlace:
                    for buscar in wArrDatoBusquedaUrl:
                        if buscar in wEnlace:
                            wExisteDcto = True
                            break
                        #fin if 
                    #fin for 
                else:
                    if wEnlace[:2] != '..': 
                        response = requests.get(wEnlace, allow_redirects=False, verify=False)  
                        for buscar in wArrDatoBusquedaUrl:
                            if buscar in response.headers['Location']:
                                wExisteDcto = True
                                wDescargarOneDrive = True
                                break
                            #fin if 
                        #fin for 
                    #fin if
                #fin if 
                
                if wExisteDcto :
                    break
                #fin if
            except:
                pass 
            #fin try
        #fin for
        
        if wExisteDcto:
            try:
                driverPdf, waitPdf = self.app_general.fnAbrirPagina(wEnlace, self.app_config.wRutaDescPdfFinV2 ) 
                time.sleep(2)
                
                if wDescargarOneDrive:
                    wTiempo = 0
                    while wTiempo <= 60:
                        try:
                            if len(driverPdf.find_elements(By.XPATH,"//span[text()='Descargar']")) > 0:
                                # Descargando
                                time.sleep(1)
                                driverPdf.find_elements(By.XPATH,"//span[text()='Descargar']")[0].click()
                                
                                wExisteProv = True
                                break
                            #fin if  
                        except:
                            pass
                        finally:
                            wTiempo+=1
                            time.sleep(1)
                        #fin try
                    #fin while 
                else:
                    wExiste = self.fnValidandoExisteDcto(  wNroRadicacion, self.app_config.wRutaDescPdfFinV2, self.app_config.wRutaDescPdfSel, 'ES-' ) 
                    
                    wExisteProv = True
                #fin if 
            except:
                pass
            finally:
                if driverPdf != None:
                    driverPdf.quit()
                #fin if 
            #fin try 
        #fin if
        
        if wDescargarOneDrive:
            if wExisteProv :
                wExiste = self.fnValidandoExisteDcto(  wNroRadicacion, self.app_config.wRutaDescPdfFinV2, self.app_config.wRutaDescPdfSel, 'ES-' )  
            else: 
                self.app_general.fnImprimir("Excedio el tiempo de espera de descarga",13)
            #fin if
        #fin if  
        
        return wExiste 
    #fnPresionandoClickEnlacePdf
     
    
    def fnSeleccionandoEnlace1raPagina( self, driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda ):
        # variables
        wExiste      = False
        wRutaDctoPdf = ''
        
        try:
            # Descargando unico enlace
            pdf_document = fitz.open(wRutaArchivo)
            page   = pdf_document.load_page(0)
            links  = page.get_links()
            enlace = links[0]['uri']
            
            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', False, wNroRadicacion, json_tbusqueda, datos, enlace )  
        except:
            pass 
        #fin try
        
        return wExiste, wRutaDctoPdf
    #fnSeleccionandoEnlace1raPagina   
    
    def verificar_tipo_enlace(self, enlace):
        try:
            patron_onedrive = r'sharepoint\.com'
            if re.search(patron_onedrive, enlace):
                return 'OneDrive'
            else:
                wPagina = 'https://www.ramajudicial.gov.co'
                wEnlace = enlace
                if enlace[0:10] == '/documents':
                    wEnlace = wPagina+enlace
                #fin if 
                response = requests.head(wEnlace, verify=False)
                content_type = response.headers.get('Content-Type')

                if content_type and 'application/pdf' in content_type:
                    return 'Pdf'
                elif content_type and 'application/x-zip' in content_type:
                    return 'Zip'
                elif content_type and 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
                    return 'Xlsx'
                else:
                    return 'Error'
                #fin if
            #fin if
            
          
        except requests.exceptions.RequestException as e:
            return 'Error'
        #fin if
    #verificar_tipo_enlace
    
    def fnProcesoDescargaDcto(self, driver, modelo_providencia, eti_a, wNroDocumento, json_tbusqueda, datos, wEnlace = False ):
        tipo_dcto     = self.verificar_tipo_enlace( eti_a.get_attribute('href') if not wEnlace else wEnlace ) 
        wExiste       = False
        wRutaDctoPdf  = self.app_config.wRutaDescPdfFin 
        if tipo_dcto == 'Pdf': 
            wExiste = self.fnValidandoExisteDcto( wNroDocumento )
            
            if wExiste:
                if not wEnlace:
                    if not eti_a.text[:3].isdigit() :
                        if modelo_providencia in ('01','02') :  
                            wExiste, wRutaDctoPdf = self.fnBuscarProvidenciaModelo01( wNroDocumento )  
                        #fin if 
                    #fin if
                else:
                    if modelo_providencia in ('01','02') :  
                        wExiste, wRutaDctoPdf = self.fnBuscarProvidenciaModelo01( wNroDocumento )  
                    #fin if 
                #fin if   
            #fin if  
        elif tipo_dcto == 'Zip':
            wExisteZip = self.fnValidandoExisteDctoZip( wNroDocumento )
            if wExisteZip:
                wExiste = self.fnSeleccionarPdfZip( wNroDocumento, datos, json_tbusqueda )  
            #fin if
        elif tipo_dcto == 'OneDrive':
            wExiste, tipo_dcto = self.fnSeleccionandoOneDrive( driver, datos, wNroDocumento, json_tbusqueda )
            if tipo_dcto == 'Pdf':
                if modelo_providencia in ('01','02') :  
                    wExiste, wRutaDctoPdf = self.fnBuscarProvidenciaModelo01( wNroDocumento )  
                #fin if 
            #finif 
        #fin if
        
        return wExiste, wRutaDctoPdf
    #fnProcesoDescargaDcto
    
    def fnProcesoDescargaDctoEstado(self, modelo_estado, eti_a, wNroDocumento, json_tbusqueda, datos, driver ):
        tipo_dcto     = self.verificar_tipo_enlace(eti_a.get_attribute('href'))
        wExiste       = False
        wExisteEnlace = False
        wRutaDctoPdf  = self.app_config.wRutaDescPdfFin 
        if tipo_dcto == 'Pdf': 
            wExiste = self.fnValidandoExisteDcto( wNroDocumento, '','', 'ES-' )
            
            if wExiste:  
                if not self.fnValidandoSiPdfEsImagen( wNroDocumento ): 
                    wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoModeloEstado( modelo_estado, wNroDocumento, driver, datos, json_tbusqueda ) 
                    if not wExisteRadicado:
                        wExiste  = False
                    #fin if 
                    
                    if not wExisteEnlace:
                        wRutaDctoPdf  = self.app_config.wRutaDescPdfFin 
                    #fin if 
                else: 
                    self.app_general.fnImprimir(f"Pdf descargado es una imagen.", 13)
                #fin if 
            else:
                self.app_general.fnImprimir( f"{datos['NumeroRadicacion']} - Excedio el tiempo de espera en la descarga del dcto.",13)
            #fin if  
        elif tipo_dcto == 'Xlsx':
            wExiste = self.fnValidandoExisteDctoXlsx( wNroDocumento ) 
            if wExiste:
                if modelo_estado == '01'   : wExiste, wRutaDctoPdf = self.fnBuscarEstadoXlsxModelo01( wNroDocumento )
            else:
                self.app_general.fnImprimir( f"{datos['NumeroRadicacion']} - Excedio el tiempo de espera en la descarga del dcto.",13)
            #fin if 
        elif tipo_dcto == 'OneDrive':
            wExiste = self.fnSeleccionandoOneDrive( driver, datos, datos['NumeroRadicacion'], json_tbusqueda, 'ES-' ) 
            if wExiste:  
                if not self.fnValidandoSiPdfEsImagen( wNroDocumento ): 
                    wExiste, wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoModeloEstado( modelo_estado, wNroDocumento, driver, datos, json_tbusqueda ) 
                    if not wExiste:
                        wRutaDctoPdf  = self.app_config.wRutaDescPdfFin 
                    #fin if 
                else: 
                    self.app_general.fnImprimir(f"Pdf descargado es una imagen.", 13)
                #fin if 
            else:
                self.app_general.fnImprimir( f"{datos['NumeroRadicacion']} - Excedio el tiempo de espera en la descarga del dcto.",13)
            #fin if  
        else:
            self.app_general.fnImprimir( f"{datos['NumeroRadicacion']} - Otro tipo de dcto.",13)
        #fin if
        
        return wExiste, wExisteEnlace, wRutaDctoPdf
    #fnProcesoDescargaDctoEstado 
    
    def fnProcesoDescargaDctoEstadoDirecto(self, modelo_estado, eti_a, wNroDocumento, json_tbusqueda, datos, driver ):
        tipo_dcto     = self.verificar_tipo_enlace(eti_a.get_attribute('href'))
        wExiste       = False
        wRutaDctoPdf  = self.app_config.wRutaDescPdfFin 
        if tipo_dcto == 'Pdf': 
            wExiste = self.fnValidandoExisteDcto( wNroDocumento, self.app_config.wRutaDescPdf, self.app_config.wRutaDescPdfFin, 'ES-' )
            
            if not wExiste:  
                self.app_general.fnImprimir( f"{datos['NumeroRadicacion']} - Excedio el tiempo de espera en la descarga del dcto.",13)
            #fin if  
            
        elif tipo_dcto == 'Xlsx':
            wExiste = self.fnValidandoExisteDctoXlsx( wNroDocumento ) 
            if wExiste:
                if modelo_estado == '01'   : wExiste, wRutaDctoPdf = self.fnBuscarEstadoXlsxModelo01( wNroDocumento )
            else:
                self.app_general.fnImprimir( f"{datos['NumeroRadicacion']} - Excedio el tiempo de espera en la descarga del dcto de estado.",13)
            #fin if 
        elif tipo_dcto == 'OneDrive':
            wExiste = self.fnSeleccionandoOneDriveEstado( driver, datos, datos['NumeroRadicacion'] )
            
            if not wExiste:  
                self.app_general.fnImprimir( f"{datos['NumeroRadicacion']} - Excedio el tiempo de espera en la descarga del dcto de estado.",13)
            #fin if  
        else:
            self.app_general.fnImprimir( f"{datos['NumeroRadicacion']} - Otro tipo de dcto.",13)
        #fin if
        
        return wExiste, wRutaDctoPdf
    #fnProcesoDescargaDctoEstado
    
    def fnSeleccionandoModeloEstado( self, wTipoModelo, wNroRadicacion, driver, datos, json_tbusqueda ):
        wExisteRadicado = False
        wExisteEnlace   = False
        wRutaDctoPdf = ''
        
        if wTipoModelo == '01'    : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo01( wNroRadicacion )
        elif wTipoModelo == '03'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo03( wNroRadicacion )
        elif wTipoModelo == '06'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo06( wNroRadicacion ) 
        elif wTipoModelo == '08'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo08( wNroRadicacion ) 
        elif wTipoModelo == '11'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo11( wNroRadicacion )
        elif wTipoModelo == '12'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo12( wNroRadicacion )
        elif wTipoModelo == '16'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo16( wNroRadicacion )
        elif wTipoModelo == '18'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo18( wNroRadicacion )
        elif wTipoModelo == '19'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo19( wNroRadicacion )
        elif wTipoModelo == '20'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo20( wNroRadicacion )
        elif wTipoModelo == '22'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo22( wNroRadicacion )
        elif wTipoModelo == '23'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo23( wNroRadicacion )
        elif wTipoModelo == '25'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo25( wNroRadicacion )
        elif wTipoModelo == '27'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo27( wNroRadicacion )
        elif wTipoModelo == '29'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo29( wNroRadicacion )
        elif wTipoModelo == '30'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo30( wNroRadicacion )
        elif wTipoModelo == '32'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo32( wNroRadicacion )
        elif wTipoModelo == '34'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo34( wNroRadicacion )
        elif wTipoModelo == '37'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo37( wNroRadicacion )
        elif wTipoModelo == '39'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo39( wNroRadicacion )
        elif wTipoModelo == '42'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo42( wNroRadicacion )
        elif wTipoModelo == '43'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo43( wNroRadicacion )
        elif wTipoModelo == '47'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo47( wNroRadicacion )
        elif wTipoModelo == '52'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo52( wNroRadicacion )
        elif wTipoModelo == '57'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo57( wNroRadicacion )
        elif wTipoModelo == '69'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo69( wNroRadicacion )
        elif wTipoModelo == '70'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo70( wNroRadicacion )
        elif wTipoModelo == '72'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo72( wNroRadicacion )
        elif wTipoModelo == '88'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo88( wNroRadicacion )
        elif wTipoModelo == '91'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo91( wNroRadicacion )
        elif wTipoModelo == '95'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo95( wNroRadicacion )
        elif wTipoModelo == '98'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo98( wNroRadicacion )
        elif wTipoModelo == '102' : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo102( wNroRadicacion )
        elif wTipoModelo == '104' : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo104( wNroRadicacion )
        elif wTipoModelo == '105' : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo105( wNroRadicacion )
        elif wTipoModelo == '107' : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo107( wNroRadicacion )
        elif wTipoModelo == '112' : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo112( wNroRadicacion )
        elif wTipoModelo == '113' : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo113( wNroRadicacion )
        
        
        elif wTipoModelo == '07'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo07( wNroRadicacion, driver, datos, json_tbusqueda ) 
        elif wTipoModelo == '40'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo40( wNroRadicacion, driver, datos, json_tbusqueda )
        elif wTipoModelo == '41'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo41( wNroRadicacion, driver, datos, json_tbusqueda )
        elif wTipoModelo == '38'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo38( wNroRadicacion, driver, datos, json_tbusqueda )
        elif wTipoModelo == '31'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo31( wNroRadicacion, driver, datos, json_tbusqueda )
        elif wTipoModelo == '28'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo28( wNroRadicacion, driver, datos, json_tbusqueda )
        elif wTipoModelo == '26'  : wExisteRadicado, wExisteEnlace, wRutaDctoPdf = self.fnBuscarEstadoModelo26( wNroRadicacion, driver, datos, json_tbusqueda )
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    #fnSeleccionandoModeloEstado
    
    def fmLimpiandoCarpeta( self, wRutaCarpeta ):
        # Obtener la lista de archivos en la carpeta
        archivos = os.listdir(wRutaCarpeta)

        # Eliminar cada archivo en la carpeta
        for archivo in archivos:
            try:
                ruta_archivo = os.path.join(wRutaCarpeta, archivo)
                if os.path.isfile(ruta_archivo):
                    os.remove(ruta_archivo)
                #fin if 
            except:
                pass
            #fin try 
        #fin for 
    #fnLimpiandoCarpeta
    
    def fnRetornarElementoTabla( self, xpath_tab, driver ):
        if xpath_tab.split('/')[-1] == 'table':
            elemento_table = driver.find_elements(By.XPATH, xpath_tab)
        else:
            elemento_tmp = driver.find_element(By.XPATH, xpath_tab)
            elemento_table = elemento_tmp.find_elements(By.TAG_NAME,'table')
        #fin if 
        
        return elemento_table
    #fnRetornarElementoTabla
    
    def fnRetornarRadicadoLimpio( self, radicado_web ):
        
        def secuencia(form: str, inf: int = 1, sup: int = 4):
            return [form.format(num) for num in range(inf, sup + 1)]
        
        valores_a_reemplazar = {
            "/", "auto.pdf"'consulte aquí la providencia del proceso', 'auto proceso', 'ver providencias adjuntas',
            ':', 'automandamientodepago', 'autoniegamedidacautelar', 'autoniegaembargorequieredemandante',
            'autocorretrasladodemandante', 'autoliquidacioncostasponeconocimiento', 'agregamemorialsinconsideracion',
            'autodecretamedidaordenacomision', 'autoliquidacioncostas', 'autoordenacomision', 'autoniegamedidacautelar',
            'automandamientodepago', '/', 'sentencia.pdf', 'providencia', *secuencia("cd {}"), *secuencia("c {}"),
            '(ver)', 'ver', *secuencia("({})"), *secuencia("no. {}"), "rad"
        }
        
        for valor in valores_a_reemplazar:
            radicado_web = str(radicado_web).replace(valor, "").strip()
        
        return radicado_web
    #fnRetornarRadicadoLimpio
    
    ######## Procesos 
    def fnProcesos( self, wProceso, juzgado, datosProceso, ejecutar_pestania_mes, driver, datos, wMesDiv, wMes ):
        try:
            # Variables
            wContinuar = False
            wPosicionLi = -1
            wMesDiv = int(wMesDiv)
            
            json_tbusqueda    = self.fnDatosBusqueda( datos['NumeroRadicacion'] )
            wArrFechaBusqueda = self.fnTipoBusquedaFecha(  datos )
            
            xpath_pest = str(self.datos_juzgados.juzgadosJSON.get(juzgado).xpath_pest)
            xpath_tab  = str(self.datos_juzgados.juzgadosJSON.get(juzgado).xpath_tab)
            
            if ejecutar_pestania_mes:
                # Seleccionando Mes
                self.app_general.fnImprimir(f"Seleccionando [Mes: {str(datos['Mes']).upper()}]",9)
                if wMes != datos['Mes']: 
                    wArrMes = xpath_pest.split('|')
                    wExiste = False
                    for ruta_xpath in wArrMes:
                        continuar, wPosicionLi = self.fnSeleccionarMesv2( driver, datos, ruta_xpath ) 
                        # continuar = self.fnSeleccionarMes( driver, datos, self.xpath_pest_mes[ruta_xpath] )
                        if not continuar:
                            xpath_tab = xpath_tab.replace('MES',str(wPosicionLi))
                            wExiste = True
                            break
                        #fin if
                    #fin for
                    
                    if not wExiste:
                        return True
                    #fin if 
                else:
                    wArrMes = xpath_pest.split('|')
                    wExiste = False
                    for ruta_xpath in wArrMes:
                        continuar, wPosicionLi = self.fnSeleccionarMesv2( driver, datos, ruta_xpath, False ) # False para que no le de click                     
                        if not continuar:
                            xpath_tab = xpath_tab.replace('MES',str(wPosicionLi))
                            wExiste = True
                            break
                        #fin if 
                    #fin if 
                #fin if
            else:
                xpath_tab = xpath_tab.replace('MES',str(1))
            #fin if 
            
            if wProceso == 1:   
                """ [ BOGOTA-001, BOGOTA-003, BOGOTA-028, BOGOTA-029, BOGOTA-030, BOGOTA-033 ]
                    datosProceso[0] => columna donde se va a realizar la busqueda
                    datosProceso[1] => columna donde se va a realizar el click
                    datosProceso[2] => columna la cual deseas que retorne
                    datosProceso[3] => Separador para la fecha
                    datosProceso[4] => Numero la cual desea que empiece el for a buscar
                    datosProceso[5] => Sumar al mes ( motivo ) juzgado 56 la tabla es 1 nro mas 
                    datosProceso[6] => True si desea q busque el radicado en el pdf, False no realiza busqueda
                    datosProceso[7] => True guardar directo al drive, False no
                    datosProceso[8] => 1 Seleccionar descarga pdf 
                    datosProceso[9] => True Descargar desde Drive, False no Descarga desde Drive
                """
                
                #'/',0,0,False, False, 1, False
                wSeparador    = datosProceso[3] if len( datosProceso ) > 3 else '/'    # Separador
                wInicioFor    = datosProceso[4] if len( datosProceso ) > 4 else 0      # Inicio de For
                wSumarMesDiv  = datosProceso[5] if len( datosProceso ) > 5 else 0      #X No se utiliza 
                wBuscarPdf    = datosProceso[6] if len( datosProceso ) > 6 else False  # Buscar Pdf
                wGuardarDrive = datosProceso[7] if len( datosProceso ) > 7 else False  # Guardar en la carpeta del Drive
                wSelDescPdf   = datosProceso[8] if len( datosProceso ) > 8 else False  # Buscar por la descripcion en el pdf ( no se utiliza )
                wDesOneDrive  = datosProceso[9] if len( datosProceso ) > 9 else False  # Descarga de one drive ( no se utiliza )

                # Descargando Segun fecha
                self.app_general.fnImprimir(f"Descargando Documento",9)
                #continuar, nro_estado = self.fnDescargarPdfPorFecha( driver, datos, self.fnRetornarXpathDescarga(datosProceso[1], (wMesDiv+wSumarMesDiv)), datosProceso[2], datosProceso[3], datosProceso[4], wSeparador, wInicioFor, wBuscarPdf, wGuardarDrive, wSelDescPdf )
                continuar, nro_estado = self.fnDescargarPdfPorFecha( driver, datos, xpath_tab, datosProceso[0], datosProceso[1], datosProceso[2], wSeparador, wInicioFor, wBuscarPdf, wGuardarDrive, wSelDescPdf, wDesOneDrive )
                if continuar: wContinuar = True
                
            elif wProceso == 2:
                try:
                    """ [ BOGOTA-002 ] 
                        datosProceso[0] => columna donde se va a realizar la busqueda 
                        datosProceso[1] => Guarda en drive
                    """ 
                    xpath_colum_busq   = datosProceso[0]
                    wGuardarDrive      = datosProceso[1] if len(datosProceso) > 1 else False 
                    xpath_colum_estado = datosProceso[2] if len(datosProceso) > 2 else False  #si este campo se encuentra lleno quiere decir que procedera a guardar el estado
                    
                    wArrEleBusCodIncidencia = []
                    elemento_estado = None
                    wRutaDctoPdf = self.app_config.wRutaDescPdfFin
                    
                    for tr in driver.find_elements(By.XPATH, xpath_tab):
                        if  len(tr.find_elements(By.XPATH,'td')) >= int(re.findall(r'\[(\d+)\]', xpath_colum_busq)[0]) :
                            if ( self.app_general.fnRetornarFechaLimpia( str(tr.find_element(By.XPATH, xpath_colum_busq).text).strip() ) in wArrFechaBusqueda ):
                                #Moviendo scroll
                                tr.location_once_scrolled_into_view 
                                driver.execute_script("window.scrollBy(0, -100);")
                                time.sleep(1)
                                
                                for a in tr.find_elements(By.TAG_NAME,'a'):
                                    wdato_limpio =  self.fnRetornarRadicadoLimpio( a.text )   
                                    if wdato_limpio == "":
                                        continue
                                    #fin if 

                                    for clave, valor in json_tbusqueda.items():
                                        if valor in wdato_limpio:
                                            wArrEleBusCodIncidencia.append(a)
                                            
                                            if xpath_colum_estado:
                                                elemento_estado = tr.find_element(By.XPATH, xpath_colum_estado)
                                            #fin if 
                                            
                                            break
                                        #fin if
                                    #fin for
        
                                    # if len(wArrEleBusCodIncidencia) > 0:
                                    #     break
                                    # #fin if
                                #fin for
                                
                                if len(wArrEleBusCodIncidencia) > 0:
                                    break
                                #fin if 
                            #fin if 
                        #fin if 
                    #fin for
                    
                    if len( wArrEleBusCodIncidencia ) == 0:
                        self.app_general.fnImprimir(f"{datos['NumeroRadicacion']} No se encontro Radicado con fecha {datos['FechaIniciaTermino']}",13)
                        self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro Radicado con fecha [ {datos['FechaIniciaTermino']}] " )
                        return True
                    #fini if

                    # DEscargando pdf estado
                    if xpath_colum_estado:
                        if elemento_estado:  
                            if len(elemento_estado.find_elements(By.TAG_NAME, 'a')) > 0:
                                elemento_eti_a  = elemento_estado.find_element(By.TAG_NAME, 'a')
                                time.sleep(0.5)

                                elemento_eti_a.click()
                                time.sleep(2)

                                # Validamos la descarga del documento
                                wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstadoDirecto( '01', elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver )
                                # Variables para el guardado en el drive
                                wNombreDocumento  = 'ES-' + datos['NumeroRadicacion'] + '.pdf'
                                wRutaDescargaDcto = wRutaDctoPdf +  wNombreDocumento
                                wRutaAlmacenamientoDriveDcto = self.app_config.wRutaAlmPdfDrive + wNombreDocumento
                                
                                if wExiste:
                                    if self.app_general.fnExisteArchivo( wRutaDctoPdf + wNombreDocumento ): 
                                        if wGuardarDrive:
                                            self.fnGuardarDocumento( wNombreDocumento, wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )   
                                        #fin if
                                    else:
                                        self.app_general.fnImprimir(f"{datos['NumeroRadicacion']} No se completo la descarga del estado",13)
                                    #fin if 
                                else:
                                    self.app_general.fnImprimir( f"No se pudo descargar el estado pdf.",13)
                                    self.app_log.error( f"Nro Radicación: {datos['NumeroRadicacion']} - No se pudo descargar el estado pdf.")
                                    if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + wNombreDocumento ):
                                        self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + wNombreDocumento )
                                    #fin if  
                                #fin if 
                        
                            #fin if   
                        else:
                            self.app_general.fnImprimir(f"{datos['NumeroRadicacion']} No se encontro enlace para la descarga de Estado",13)    
                            self.app_log.warning( f"{datos['NumeroRadicacion']} - No se encontro enlace para la descarga de Estado - COD: {datos['Ciudad']} ")  
                        #fin if 
                    #fin if 
                    
                    # Descargando Dcto Pdf
                    contDcto = 1
                    for eleA in wArrEleBusCodIncidencia:
                        try:
                            self.app_general.fnImprimir(f"Descargando Documento {eleA.text}",13)
                            eleA.location_once_scrolled_into_view
                            driver.execute_script("window.scrollBy(0, -100);")
                            time.sleep(0.5)
                            eleA.click()
                            
                            wExiste = self.fnValidandoExisteDcto(  datos['NumeroRadicacion']  )
                            
                            if wExiste:
                                wNombreDocumento       = 'EE-' + datos['NumeroRadicacion'] + '.pdf'
                                wNombreDocumentoSalida = 'EE-' + (datos['NumeroRadicacion'] + ( '' if len(wArrEleBusCodIncidencia) == 1 else '_'+str(contDcto).zfill(2) ) ) + '.pdf'
                                  
                                if self.app_general.fnExisteArchivo( wRutaDctoPdf + wNombreDocumento ): 
                                    if wGuardarDrive:
                                        self.fnGuardarDocumento( wNombreDocumento, wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive, wNombreDocumentoSalida )    
                                    #fin if
                                else:
                                    self.app_general.fnImprimir(f"{datos['NumeroRadicacion']} No se completo la descarga del radicado",13)
                                #fin if 
                                time.sleep(1)
                            else:
                                self.app_general.fnImprimir(f"Excedio el tiempo de espera con la descarga del dcto.",13)
                            #fin if 
                        except:
                            pass
                        finally:
                            contDcto+=1
                        
                    #fin for 
                except:
                    pass
            elif wProceso == 3: 
                """ [ BOGOTA-004 ]
                    datosProceso[0] => columna busqueda
                    datosProceso[1] => columna click
                    datosProceso[2] => Guardar en drive 
                    datosProceso[3] => columna busqueda 2 -> la cual se va a concatenar con la "columna busqueda 1"
                    datosProceso[4] => indicar que debe darle click a la primera pagina de la tabla
                """ 
                self.app_general.fnImprimir(f"Buscando por Nro de Radicado.",11)
                xpath_colum_busq    = datosProceso[0]
                xpath_colum_click   = datosProceso[1]
                wGuardarDrive       = datosProceso[2] if len(datosProceso) > 2 else False 
                xpath_colum_busq2   = datosProceso[3] if len(datosProceso) > 3 else False 
                seleccionar_1raFila = datosProceso[4] if len(datosProceso) > 4 else True 
                cuenta_con_subtabla = datosProceso[5] if len(datosProceso) > 5 else False 
                
                ## Descargar segun Fecha y seleccionando si es par o impar
                self.app_general.fnImprimir(f"Buscando Nro Radicado",11)
                elemento         = []
                existe           = False 
                tipo_dcto        = 'Pdf'
                wExiste          = False
                wRutaDctoPdf     = self.app_config.wRutaDescPdfFin 
                wBuscarDentroPdf = False
                  
                for table in self.fnRetornarElementoTabla( xpath_tab, driver ): 
                    elementos_tr = table.find_elements(By.TAG_NAME,'tr') 
                    if cuenta_con_subtabla:
                        elemento_subTable = table.find_element(By.TAG_NAME,'table')
                        elementos_tr = elemento_subTable.find_elements(By.TAG_NAME,'tr') 
                    #fin if 
                    
                    elemento = [tr 
                                    for tr in elementos_tr
                                    if  len(tr.find_elements(By.XPATH,'td')) >= int(re.findall(r'\[(\d+)\]', xpath_colum_busq)[0]) and 
                                        (   datos['NumeroRadicacion']  in str(tr.find_element(By.XPATH, xpath_colum_busq).text).strip().replace('-','') or 
                                            (str(tr.find_element(By.XPATH, xpath_colum_busq).text).strip().replace('-','').replace('\n','') + (  str(tr.find_element(By.XPATH, xpath_colum_busq2).text).strip().replace('-','').replace('\n','') if xpath_colum_busq2 else '' )) in json_tbusqueda.values()
                                        ) 
                    ]
                    if len( elemento ) > 0:  
                        try:
                            if elemento[0].find_element(By.XPATH, xpath_colum_click).text != '':
                                elemento_td = elemento[0].find_element(By.XPATH, xpath_colum_click)
                                eti_a       = elemento_td.find_element(By.TAG_NAME, 'a' )
                                elemento_td.find_element(By.TAG_NAME, 'a').click()
                                
                                tipo_dcto     = self.verificar_tipo_enlace( eti_a.get_attribute('href')  ) 
                                
                                existe = True 
                                break
                            #fin if 
                        except:
                            pass
                        #fin try
                        
                        if not existe and seleccionar_1raFila:
                            existe = False 
                            try:
                                element_tr = table.find_element(By.XPATH,'tbody/tr[1]')
                                if element_tr.find_element(By.TAG_NAME,'a') != '':
                                    # element_tr.find_element(By.TAG_NAME,'a').click()
                                    # elemento_td.find_element(By.TAG_NAME, 'a').click()
                                    eti_a       = element_tr.find_element(By.TAG_NAME,'a')
                                    eti_a.click()
                                    
                                    tipo_dcto     = self.verificar_tipo_enlace( eti_a.get_attribute('href')  ) 
                                    existe = True 
                                    wBuscarDentroPdf = True
                                    break
                                #fin if 
                            except:
                                pass
                            #fin try 
                        #fin if 
                        
                        if not existe:
                            self.app_general.fnImprimir(f"No se encontro enlace para el nro de Radicado.",13)
                        #fin if 
                        break
                    #fin if
                #fin for
                
                if existe: 
                    if tipo_dcto == 'Pdf': 
                        wExiste = self.fnValidandoExisteDcto( datos['NumeroRadicacion'] ) 
                        if wExiste == True and wBuscarDentroPdf == True: 
                            wExiste, wRutaDctoPdf = self.fnBuscarProvidenciaModelo01( datos['NumeroRadicacion'] )  
                        #fin if  
                    elif tipo_dcto == 'Zip':
                        wExisteZip = self.fnValidandoExisteDctoZip( datos['NumeroRadicacion'] )
                        if wExisteZip:
                            wExiste = self.fnSeleccionarPdfZip( datos['NumeroRadicacion'], datos, json_tbusqueda )  
                        #fin if
                    elif tipo_dcto == 'OneDrive':
                        wExiste = self.fnSeleccionandoOneDrive( driver, datos, datos['NumeroRadicacion'], json_tbusqueda )
                    #fin if
             
                    if wExiste:
                        if wGuardarDrive:
                            self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                            
                            while True:
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                    self.app_general.fnEliminarArchivo( wRutaDctoPdf + datos['NumeroRadicacion']  + '.pdf' )
                                    self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                    break
                                #fin if 
                                time.sleep(1)
                            #fin while
                        #fin if 
                    else:
                        self.app_general.fnImprimir(f"No se encontro Nro de Radicado, o Excedio el tiempo de descarga del dcto.",13)
                    #fin if
                else:
                    self.app_general.fnImprimir(f"No se encontro Nro de Radicado.",13)
                #fin if
            elif wProceso == 4: 
                """ [ BOGOTA-004 ]
                    datosProceso[0] => Campo a buscar
                    datosProceso[1] => Campo hacer click
                    datosProceso[2] => Guardar directo al drive
                """ 
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                xpath_colum_busq  = datosProceso[0]
                xpath_colum_click = datosProceso[1].split('-')[0]
                wGuardarDrive     = datosProceso[2] if len(datosProceso) > 2 else False 
                wPosicionFecha    = datosProceso[3] if len(datosProceso) > 3 else False
                elemento = []
                ## Descargar segun Fecha y seleccionando si es par o impar
                 
                for table in self.fnRetornarElementoTabla( xpath_tab, driver ):
                    elemento = [
                        tr.find_element(By.XPATH, xpath_colum_click) 
                        for tr in table.find_elements(By.XPATH,'tbody/tr') 
                        # if self.app_general.fnRetornarFechaLimpia( str(tr.find_element(By.XPATH, xpath_colum_busq).text).strip().split(' ')[wPosicionFecha] if wPosicionFecha else str(tr.find_element(By.XPATH, xpath_colum_busq).text).strip() ) == self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/') 
                        if self.app_general.fnRetornarFechaLimpia( str(tr.find_element(By.XPATH, xpath_colum_busq).text).strip().split(' ')[wPosicionFecha] if wPosicionFecha else str(tr.find_element(By.XPATH, xpath_colum_busq).text).strip() ) in wArrFechaBusqueda
                    ]
                    
                    if len( elemento ) > 0:
                        wContador = 0
                        for a in elemento[0].find_elements(By.TAG_NAME, 'a'): 
                            time.sleep(1)
                            a.click()
                            
                            wNroDocumento = datos['NumeroRadicacion'] + ( '_'+str(wContador) if wContador > 0 else '' ) 
                            tipo_dcto     = self.verificar_tipo_enlace(a.get_attribute('href'))
                            wExiste       = False
                            wRutaDctoPdf  = self.app_config.wRutaDescPdfFin
                            if tipo_dcto == 'Pdf': 
                                wExiste = self.fnValidandoExisteDcto( wNroDocumento )
                                
                                if wExiste:
                                    if  len(datosProceso[1]) > 2:
                                        wExiste, wRutaDctoPdf = self.fnSeleccionandoModeloEstado( datosProceso[1].split('-')[2], wNroDocumento, driver, datos, json_tbusqueda  ) 
                                    #fin if 
                                    
                                    # Descarga a nivel de providencia 
                                    if len(datosProceso[1]) > 1:
                                        if datosProceso[1].split('-')[1] == '01':
                                            wExiste, wRutaDctoPdf = self.fnBuscarProvidenciaModelo01( wNroDocumento )
                                        #fin if
                                    #fin if 
                                #fin if  
                            elif tipo_dcto == 'Zip':
                                wExisteZip = self.fnValidandoExisteDctoZip( wNroDocumento )
                                if wExisteZip:
                                    wExiste = self.fnSeleccionarPdfZip( wNroDocumento, datos, json_tbusqueda )  
                                #fin if
                            elif tipo_dcto == 'OneDrive':
                                wExiste = self.fnSeleccionandoOneDrive( driver, datos, wNroDocumento, json_tbusqueda )
                            #fin if
                             
                            if wExiste :
                                if wGuardarDrive:
                                    self.app_general.fnCopiarArchivo( wRutaDctoPdf + wNroDocumento + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                                    
                                    while True:
                                        if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                            self.app_general.fnEliminarArchivo( wRutaDctoPdf + datos['NumeroRadicacion']  + '.pdf' )
                                            self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                            break
                                        #fin if 
                                        time.sleep(1)
                                    #fin while
                                #fin if 
                            else:
                                self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                                self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' ):
                                    self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' )
                                #fin if 
                            #fin if 
                            
                            wContador+=1
                        #fin if 
                        break
                    #fin if
                #fin for
                
                if len( elemento ) == 0: 
                    self.app_general.fnImprimir(f"No se encontro Nro de Radicado.",13)
                #fin if
            elif wProceso == 5:
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => Campo a buscar  
                    datosProceso[1] => Campo hacer click 
                    datosProceso[2] => GUardar al drive
                """ 
                
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                
                xpath_colum_busq  = datosProceso[0]
                xpath_colum_click = datosProceso[1] 
                wGuardarDrive     = datosProceso[2] if len(datosProceso) > 2 else False 
                
                xpath_tabTmp = xpath_tab + ( '/tbody/tr'  if xpath_tab.split('/')[-1] == 'table' else '/table[1]/tbody/tr' )
                click_estado, click_provid = xpath_colum_click.split('|')
                xpath_estado, xpath_provid = self.fnObtenerXpathClick( driver, click_estado, click_provid, xpath_tabTmp ) 
                wNroModelo   = click_provid.split('-')[1]
                wRutaDctoPdf = self.app_config.wRutaDescPdfFin
    
                for table in self.fnRetornarElementoTabla( xpath_tab, driver ):
                    arr_elemento_fecha = [    
                        ( tr.find_element(By.XPATH, xpath_estado) if xpath_estado != '' else False,  tr.find_element(By.XPATH, xpath_provid)  if xpath_provid != '' else False )
                        for tr in table.find_elements(By.TAG_NAME,'tr') 
                        if  self.app_general.fnRetornarFechaLimpia( tr.find_element(By.XPATH, xpath_colum_busq).text ) in wArrFechaBusqueda
                    ]
                    if len( arr_elemento_fecha ) > 0:  
                        #Presionando Descargar Providencia 
                        if xpath_provid != '':
                            for elementos in arr_elemento_fecha:
                                if elementos[1]:
                                    wContador = 0
                                    Existe = False
                                    for a in elementos[1].find_elements(By.TAG_NAME, 'a'):
                                        try:
                                            if 'estado' in a.get_attribute('href'):
                                                continue
                                            #fin if 
                                            
                                            # Presionamos click    
                                            a.click()
                                            time.sleep(1)
                                            
                                            wNroDocumento = datos['NumeroRadicacion'] + ( '_'+str(wContador) if wContador > 0 else '' )   
                                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, a, wNroDocumento, json_tbusqueda, datos ) 
                                              
                                            if wExiste : 
                                                if wGuardarDrive:
                                                    self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                                                    
                                                    while True:
                                                        if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                                            self.app_general.fnEliminarArchivo( wRutaDctoPdf + wNroDocumento + '.pdf' )
                                                            self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                                            break
                                                        #fin if 
                                                        time.sleep(1)
                                                    #fin while
                                                    
                                                    break
                                                #fin if 
                                            else:
                                                self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                                                self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                                                if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' ):
                                                    self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' )
                                                #fin if 
                                            #fin if
                                            
                                        except:
                                            pass
                                        finally:
                                            wContador+=1
                                        #fin try                                            
                                    #fin if 
                                    
                                    if Existe:
                                        break
                                    #fin if 
                                #fin if 
                            #fin for
                        #fin if  
                        break
                    #fin if
                #fin if 
            elif wProceso == 6: # recorriendo pestaña
                """ [ BOGOTA-013 ] 
                    datosProceso[0] => Campo a buscar  
                    datosProceso[1] => condicion a buscar
                    datosProceso[2] => Campo hacer click
                    datosProceso[3] => GUardar al drive
                """ 
                
                xpath_colum_busq   = datosProceso[0]
                dato_buscar        = datosProceso[1]
                xpath_colum_click  = datosProceso[2]
                wGuardarDrive      = datosProceso[3] if len(datosProceso) > 3 else False 
                wPosicionSeparador = datosProceso[4] if len(datosProceso) > 4 else 19  # esta es la posición del Div la cual se va actualizar por el "nro de pestaña"
                existe             = False
                wBuscarDentroPdf   = False
                for item in range(len([ _li for _li in driver.find_elements(By.XPATH, xpath_pest)]) - 1, -1, -1):
                    # Variables 
                    xpath_pestania = xpath_pest+'['+str(item+1)+']'  
                    xpath_descarga = xpath_tab[:xpath_tab.rfind("/")] + '/div['+str(item+1)+']' 
                    
                    busqueda_radicado = dato_buscar if dato_buscar != '' else json_tbusqueda.values()
                    busqueda_fecha    = ''
                    
                    column_busq01 = xpath_colum_busq
                    column_busq02 = ''
                    
                    if '|' in dato_buscar:
                        arrBusqueda = dato_buscar.split('|')
                        busqueda_radicado = arrBusqueda[0]
                        busqueda_fecha    = arrBusqueda[1]
                    #fin if
                    
                    if '|' in dato_buscar:
                        arrBusqueda = xpath_colum_busq.split('|')
                        column_busq01 = arrBusqueda[0]
                        column_busq02 = arrBusqueda[1]
                    #fin if
                    
                    try:
                        elemento_pest = driver.find_element(By.XPATH, xpath_pestania)
                        elemento_pest.location_once_scrolled_into_view
                        driver.execute_script("window.scrollBy(0, -100);")
                        time.sleep(0.5)
                        elemento_pest.click()
                        time.sleep(2)
                    except:
                        pass
                    #fin try
                    
                    if xpath_descarga.split('/')[-1] == 'table':
                        elemento_table = driver.find_elements(By.XPATH, xpath_descarga)
                    else:
                        elemento_tmp = driver.find_element(By.XPATH, xpath_descarga)
                        elemento_table = elemento_tmp.find_elements(By.TAG_NAME,'table')
                    #fin if 
                    
                    for table in elemento_table:
                        for fila in table.find_elements(By.TAG_NAME, 'tr'):
                            
                            if (fila.find_element(By.XPATH,column_busq01).text).replace('"',"").replace('\n'," ").replace('  '," ").replace('.pdf','') in busqueda_radicado: 
                                elemento_pdf = fila.find_elements(By.XPATH,'td')[xpath_colum_click] 
                                if len(elemento_pdf.find_elements(By.TAG_NAME, 'a')) > 0:
                                    elemento_pdf.click()
                                    existe = True
                                #fin if 
                                break
                            #fin if 
                            
                            if busqueda_fecha != '':
                                if self.app_general.fnRetornarFechaLimpia(  (fila.find_element(By.XPATH,column_busq02).text).replace('"',"") ) == busqueda_fecha: 
                                    elemento_pdf = fila.find_elements(By.XPATH,'td')[xpath_colum_click] 
                                    if len(elemento_pdf.find_elements(By.TAG_NAME, 'a')) > 0:
                                        elemento_pdf.click()
                                        existe = True
                                        wBuscarDentroPdf = True
                                    #fin if 
                                    break
                                #fin if 
                            #fin if 
                        #fin for
                        
                        if existe:
                            break
                        #fin if 
                        
                    #fin for
                       
                    if existe:
                        break
                    #fin if 
                #fin for   
                
                if existe:  
                    # Validando si se descargo el documento
                    wExiste      = self.fnValidandoExisteDcto(  datos['NumeroRadicacion']  )
                    wRutaDctoPdf = self.app_config.wRutaDescPdfFin
                    if wExiste:
                        if wBuscarDentroPdf:
                            wExiste, wRutaDctoPdf = self.fnBuscarProvidenciaModelo01( datos['NumeroRadicacion'] ) 
                        #fin if 
                        
                        if wExiste:
                            if wGuardarDrive:
                                self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion']  + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                                
                                while True:
                                    if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                        self.app_general.fnEliminarArchivo( wRutaDctoPdf + datos['NumeroRadicacion']  + '.pdf' )
                                        self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                        break
                                    #fin if 
                                    time.sleep(1)
                                #fin while
                            #fin if 
                        #fin if 
                    else:
                        self.app_general.fnImprimir(f"Excedio el tiempo en espera del documento.",13)
                    #fin if 
                else:
                    self.app_general.fnImprimir(f"No se encontro documento.",13)
                #fin if  
            elif wProceso == 7:
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => columna donde se va a realizar la busqueda
                    datosProceso[1] => Guarda en drive
                    datosProceso[2] => click columna 
                    datosProceso[3] => buscar en pdf
                """ 
                xpath_colum_busq  = datosProceso[0]
                wGuardarDrive     = datosProceso[1] if len(datosProceso) > 1 else False 
                wClickProvidencia = datosProceso[2] if len(datosProceso) > 2 else False 
                wBuscarEnPdf      = datosProceso[3] if len(datosProceso) > 3 else False 
                wArrEleBusCodIncidencia = []
                wArrFechaBusqueda =  self.fnTipoBusquedaFecha( datos )
                
                for table in self.fnRetornarElementoTabla( xpath_tab, driver ):
                    wFila = 0
                    for tr in table.find_elements(By.XPATH, 'tbody/tr'):
                        if ( self.app_general.fnRetornarFechaLimpia( str(tr.find_element(By.XPATH, xpath_colum_busq).text).strip() ) in wArrFechaBusqueda ):
                            if wClickProvidencia:
                                if 'tr' in wClickProvidencia:
                                    wArr = wClickProvidencia.split('-')
                                    elementoTr = 'tr['+str( wFila + int( wArr[0].replace('tr+','') ) )+']'
                                    elementoTd = wArr[1]
                                    
                                    elemento_td = table.find_elements(By.XPATH, 'tbody/'+elementoTr+'/'+elementoTd) 
                                else:
                                    elemento_td = tr.find_element(By.XPATH, wClickProvidencia)
                                #fin if 
                                wArrEleBusCodIncidencia.append( elemento_td.find_element(By.TAG_NAME, 'a') ) 
                            else:
                                for a in tr.find_elements(By.TAG_NAME,'a'): 
                                    wdato = str(a.text.split(' ')[0])
                                    wdato_limpio = wdato.replace(wdato.split('-')[-1],wdato.split('-')[-1].zfill(5))
                                    
                                    for clave, valor in json_tbusqueda.items():
                                        if valor in wdato_limpio:
                                            wArrEleBusCodIncidencia.append(a) 
                                        #fin if 
                                    #fin for  
                                    
                                    if len(wArrEleBusCodIncidencia) > 0:
                                        break
                                    #fin if
                                #fin for
                            #fin if
                              
                            if len(wArrEleBusCodIncidencia) > 0:
                                break
                            #fin if
                        #fin if 
                        
                        wFila+=1
                    #fin for
                    
                    # temporal hasta indicar que se pueden descargar mas de 1 dcto.
                    if len(wArrEleBusCodIncidencia) > 0:
                        break
                    #fin if
                #fin if 
                 
                if len( wArrEleBusCodIncidencia ) == 0:
                    self.app_general.fnImprimir(f"ERROR: No se encontro providencia [{json_tbusqueda['tbusq-1']}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro providencia [{json_tbusqueda['tbusq-1']}] " )
                    return True
                #fini if 
                
                # Descargando Dcto Pdf
                for eleA in wArrEleBusCodIncidencia:
                    self.app_general.fnImprimir(f"Descargando Documento {eleA.text}",13)
                    eleA.location_once_scrolled_into_view
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.5)
                    eleA.click()
                    time.sleep(2)
                    
                    wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', eleA, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                    
                    if wExiste:  
                        if wGuardarDrive:
                            if wExiste:
                                self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                                
                                while True:
                                    if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                        self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion']  + '.pdf' )
                                        self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                        break
                                    #fin if 
                                    time.sleep(1)
                                #fin while
                                
                                break
                            #fin if 
                        #fin if 
                    time.sleep(1)
                #fin for  
            elif wProceso == 8: # no se usa
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => columna donde se va a realizar la busqueda
                    datosProceso[1] => Guarda en drive
                """ 
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                # Variables segun parametro
                xpath_colum_busq  = datosProceso[0]
                xpath_colum_click = datosProceso[1]
                empieza           = datosProceso[2] if len(datosProceso) > 2 else 0 
                wGuardarDrive     = datosProceso[3] if len(datosProceso) > 3 else False 
                
                # Variables independientes
                wExiste = False
                wBuscarPdf = False
                for tr in driver.find_elements(By.XPATH, xpath_tab)[empieza:]:
                    if len(tr.find_elements(By.XPATH,'td')) >= int(re.findall(r'\[(\d+)\]', xpath_colum_busq)[0]) :
                        if ( self.app_general.fnRetornarFechaLimpia( str(tr.find_element(By.XPATH, xpath_colum_busq).text).strip() ) in wArrFechaBusqueda ):
                            tr.find_element(By.XPATH, xpath_colum_click).click()
                            if not tr.find_element(By.XPATH, xpath_colum_click).text[:3].isdigit() :
                                wBuscarPdf = True
                            #fin if 
                            wExiste = True
                            break
                        #fin if 
                    #fin if 
                #fin for
                 
                if not wExiste:
                    self.app_general.fnImprimir(f"ERROR: No se encontro providencia [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro providencia [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}] " )
                    return True
                #fini if 
                
                # wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                self.fnValidandoExisteDcto(  datos['NumeroRadicacion']  )
                wExiste = True
                wRutaDctoPdf =  self.app_config.wRutaDescPdfFin
                if wBuscarPdf:
                    wExiste, wRutaDctoPdf = self.fnBuscarProvidenciaModelo01( datos['NumeroRadicacion'], datos['NombreoRazonSocial']  ) 
                #fin if 
                wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                    
                if wExiste:
                    if wGuardarDrive:
                        self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                        
                        while True:
                            if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                self.app_general.fnEliminarArchivo( wRutaDctoPdf + datos['NumeroRadicacion']  + '.pdf' )
                                self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                break
                            #fin if 
                            time.sleep(1)
                        #fin while
                    #fin if  
                #fin if  
            elif wProceso == 9:
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => Campo a buscar    
                    datosProceso[1] => GUardar al drive
                    datosProceso[2] => Enlace en el Radicado
                """ 
                
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                xpath_colum_busq   = datosProceso[0] 
                wGuardarDrive      = datosProceso[1] if len(datosProceso) > 1 else False 
                wClickEnlace       = datosProceso[2] if len(datosProceso) > 2 else False 
                xpath_colum_estado = datosProceso[3] if len(datosProceso) > 3 else False  #si este campo se encuentra lleno quiere decir que procedera a guardar el estado
                utilizar_elemento_a = datosProceso[4] if len(datosProceso) > 4 else False  #Utilizar elemento que se obtiene mientras recorre aun no se usa
                
                elemento_eti_a    = False
                wBuscarEnPdf      = False
                wNroTablaSeleccionada   = -1
                wNroFilaSeleccionada    = -1
                elemento_estado   = False
                wConTabla = 0
                for table in self.fnRetornarElementoTabla( xpath_tab, driver ):
                    wConFila = 0
                    for tr in table.find_elements(By.XPATH,'tbody/tr'):
                        if  len(tr.find_elements(By.XPATH,'td')) >= int(re.findall(r'\[(\d+)\]', xpath_colum_busq)[0]):
                            elemento_td = tr.find_element(By.XPATH, xpath_colum_busq)
                            if not wClickEnlace :
                                for eti_a in elemento_td.find_elements(By.TAG_NAME, 'a'): 
                                    
                                    wdato_limpio = str(eti_a.text).strip().replace('-','').split(' ')[0]
                                    if not a.text[:3].isdigit():  
                                        elemento_eti_a = eti_a 
                                    else:
                                        for clave, valor in json_tbusqueda.items():
                                            if valor in wdato_limpio:
                                                elemento_eti_a = eti_a
                                                
                                                if xpath_colum_estado:
                                                    wNroTablaSeleccionada = wConTabla  
                                                    elemento_columna = tr.find_element(By.XPATH,xpath_colum_estado)
                                                    if len(elemento_columna.find_elements(By.TAG_NAME, 'a')):
                                                        wNroFilaSeleccionada = wConFila
                                                    #fin if 
                                                #fin if 
                                                
                                            #fin if 
                                        #fin for
                                    #fin if  
                                    
                                    if elemento_eti_a:
                                        break
                                    #fin if     
                                #fin for  
                            else:
                               
                                wdato_limpio = self.fnRetornarRadicadoLimpio( elemento_td.text )    
                                for clave, valor in json_tbusqueda.items():
                                    if valor in wdato_limpio:
                                        elemento_td = tr.find_element(By.XPATH, wClickEnlace)
                                        elemento_eti_a = elemento_td.find_element(By.TAG_NAME, 'a')
                                         
                                        if xpath_colum_estado: 
                                            # if utilizar_elemento_a:
                                            #     #variables 
                                            #     cadena = elemento_eti_a.get_attribute("href")
                                            #     posicion_codigo = cadena.find("/", cadena.find("/", cadena.find("/", cadena.find("/", cadena.find("/", cadena.find("/") + 1) + 1) + 1) + 1) + 1) + 1
                                            #     ubicacion_codigo = elemento_eti_a.get_attribute("href")[:posicion_codigo]
                                                
                                            #     elemento_tmp = tr.find_element(By.XPATH, xpath_colum_estado)
                                            #     elemento_estado = [ eti_a.text for eti_a in elemento_tmp.find_elements(By.TAG_NAME, 'a') if ubicacion_codigo  in eti_a.get_attribute('href')][0]
                                            # #fin if 
                                            
                                            wNroTablaSeleccionada = wConTabla  
                                            if wClickEnlace != xpath_colum_estado:
                                                elemento_columna = tr.find_element(By.XPATH,xpath_colum_estado)
                                                if len(elemento_columna.find_elements(By.TAG_NAME, 'a')):
                                                    wNroFilaSeleccionada = wConFila
                                                #fin if 
                                            #fin if 
                                        #fin if 
                                        
                                        break
                                    #fin if 
                                #fin for 
                            #fin if
                            
                            if elemento_eti_a:
                                break
                            #fin if 
                        #fin if
                        
                        wConFila+=1
                    #fin for
                    
                    if elemento_eti_a:
                        break
                    #fin if
                    
                    wConTabla+=1
                #fin for
                    
                if elemento_eti_a: 
                    
                    ## proceso de descarga de estado
                    if xpath_colum_estado:
                        wConTabla = 0
                        
                        # Seleccionamos el enlace de la etique A
                        # if not utilizar_elemento_a:
                        elemento_tabla   = self.fnRetornarElementoTabla( xpath_tab, driver )[wNroTablaSeleccionada]
                        if wNroFilaSeleccionada < 0:
                            for tr in  elemento_tabla.find_elements(By.TAG_NAME,'tr'):
                                elemento_columna = tr.find_element(By.XPATH,xpath_colum_estado)
                                if len(elemento_columna.find_elements(By.TAG_NAME, 'a')):
                                    elemento_eti_a   = elemento_columna.find_element(By.TAG_NAME, 'a')
                                    time.sleep(0.5)
                                    break
                                #fin if         
                            #fin if 
                        else:
                            elemento_fila    = elemento_tabla.find_elements(By.TAG_NAME,'tr')[wNroFilaSeleccionada]
                            elemento_columna = elemento_fila.find_element(By.XPATH,xpath_colum_estado)
                            elemento_eti_a   = elemento_columna.find_element(By.TAG_NAME, 'a')
                            time.sleep(0.5)       
                        #fin if 
                        # else:
                        #     elemento_eti_a = elemento_estado
                        # #fin if 
                                                    
                        # Presionamos en el enlace
                        elemento_eti_a.click()
                        time.sleep(2)
                        
                        # Validamos la descarga del documento
                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstadoDirecto( '01', elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver )  
                        
                        # Variables para el guardado en el drive
                        wNombreDocumento  = 'ES-' + datos['NumeroRadicacion'] + '.pdf'
                        wRutaDescargaDcto = wRutaDctoPdf +  wNombreDocumento
                        wRutaAlmacenamientoDriveDcto = self.app_config.wRutaAlmPdfDrive + wNombreDocumento
                        
                        if wExiste:
                            if wGuardarDrive:
                                #Guardando Estado
                                self.fnGuardarDocumento( wNombreDocumento, wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )  
                            #fin if
                        else:
                            self.app_general.fnImprimir( f"No se pudo descargar el estado pdf.",13)
                            self.app_log.error( f"Nro Radicación: {datos['NumeroRadicacion']} - No se pudo descargar el estado pdf.")
                            if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' ):
                                self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' )
                            #fin if  
                        #fin if 
                    #fin if  
                    
                    
                    # Moviendo scroll
                    elemento_eti_a.location_once_scrolled_into_view
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.5)
                    
                    # click
                    elemento_eti_a.click()
                    time.sleep(2)
                    
                    # Validando documento  
                    wNroDocumento = datos['NumeroRadicacion'] + ( '_'+str(wContador) if wContador > 0 else '' ) 
                    wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', elemento_eti_a, wNroDocumento, json_tbusqueda, datos ) 
                    
                    # Guardando Drive
                    if wExiste:
                        if wGuardarDrive:
                            wNombreDocumento  = 'EE-' + wNroDocumento + '.pdf'
                            #Guardando Providencia - auto
                            self.fnGuardarDocumento( wNombreDocumento, wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )   
                        #fin if 
                    #fin if 
                    
                #fin if 
            elif wProceso == 10:
                """ [ BOGOTA-020 ] 
                    datosProceso[0] => Campo a buscar fecha   
                    datosProceso[1] => Campo a buscar codigo
                    datosProceso[2] => Guardar al drive
                """  
                
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                # Variables segun parametro
                xpath_colum_busqFec  = datosProceso[0] 
                xpath_colum_busqCod  = datosProceso[1] 
                empieza_recorrer     = datosProceso[2] if len(datosProceso) > 2 else 0 
                wGuardarDrive        = datosProceso[3] if len(datosProceso) > 3 else False 
                buscarProvEst        = datosProceso[4] if len(datosProceso) > 4 else '01-00' 
                xpath_colum_estado   = datosProceso[5] if len(datosProceso) > 5 else False  #si este campo se encuentra lleno quiere decir que procedera a guardar el estado
                
                # Variables independientes
                elemento_eti_a       = False
                wBuscarEnPdf         = True 
                wArrEleBusCodIncidencia = []
                wNroTablaSeleccionada   = -1
                wNroFilaSeleccionada    = -1
                
                wConTabla = 0
                for table in self.fnRetornarElementoTabla( xpath_tab, driver ):
                    wConFila = empieza_recorrer
                    for tr in table.find_elements(By.TAG_NAME,'tr')[empieza_recorrer:]:
                        if  len(tr.find_elements(By.XPATH,'td')) >= int(re.findall(r'\[(\d+)\]', xpath_colum_busqFec)[0]):
                            if ( self.app_general.fnRetornarFechaLimpia( str(tr.find_element(By.XPATH, xpath_colum_busqFec).text).strip() ) in wArrFechaBusqueda ):
                                elemento_td = tr.find_element(By.XPATH, xpath_colum_busqCod)
                                for a in elemento_td.find_elements(By.TAG_NAME,'a'): 
                                    wDescripcion = (str(a.text).replace('Consulte aquí la providencia del proceso ','')).strip()
                                    if not wDescripcion[:3].isdigit(): 
                                        wArrEleBusCodIncidencia.append(a) 
                                        wBuscarEnPdf = True 
                                    else:     
                                        wdato_limpio = self.fnRetornarRadicadoLimpio( a.text )                                        
                                        for clave, valor in json_tbusqueda.items():
                                            if valor in wdato_limpio:
                                                wArrEleBusCodIncidencia.append(a) 
                                                
                                                if xpath_colum_estado:
                                                    wNroTablaSeleccionada = wConTabla  
                                                    elemento_columna = tr.find_element(By.XPATH,xpath_colum_estado)
                                                    if len(elemento_columna.find_elements(By.TAG_NAME, 'a')):
                                                        wNroFilaSeleccionada = wConFila
                                                    #fin if 
                                                #fin if                                                 
                                            #fin if 
                                        #fin for  
                                    #fin if   
                                
                                    if len(wArrEleBusCodIncidencia) > 0:
                                        break
                                    #fin if  
                                #fin for
                            
                                if len(wArrEleBusCodIncidencia) > 0:
                                    break
                                #fin if  
                            #fin if 
                        #fin for
                        
                        if len(wArrEleBusCodIncidencia) > 0:
                            break
                        #fin if  
                        wConFila+=1
                    #fin for
                    
                    if len(wArrEleBusCodIncidencia) > 0:
                        break
                    #fin if 
                    
                    wConTabla+=1
                #fin for
                
                if len(wArrEleBusCodIncidencia) > 0:
                    ## proceso de descarga de estado
                    if xpath_colum_estado:
                        wConTabla = 0
                        
                        # Seleccionamos el enlace de la etique A
                        elemento_tabla   = self.fnRetornarElementoTabla( xpath_tab, driver )[wNroTablaSeleccionada]
                        if wNroFilaSeleccionada < 0:
                            for tr in  elemento_tabla.find_elements(By.TAG_NAME,'tr'):
                                elemento_columna = tr.find_element(By.XPATH,xpath_colum_estado)
                                if len(elemento_columna.find_elements(By.TAG_NAME, 'a')):
                                    elemento_eti_a   = elemento_columna.find_element(By.TAG_NAME, 'a')
                                    time.sleep(0.5)
                                    break
                                #fin if         
                            #fin if 
                        else:
                            elemento_fila    = elemento_tabla.find_elements(By.TAG_NAME,'tr')[wNroFilaSeleccionada]
                            elemento_columna = elemento_fila.find_element(By.XPATH,xpath_colum_estado)
                            elemento_eti_a   = elemento_columna.find_element(By.TAG_NAME, 'a')
                            time.sleep(0.5)       
                        #fin if 
                                                    
                        # Presionamos en el enlace
                        elemento_eti_a.click()
                        time.sleep(2)
                        
                        # Validamos la descarga del documento
                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstadoDirecto( '01', elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver )  
                        
                        # Variables para el guardado en el drive
                        wNombreDocumento  = 'ES-' + datos['NumeroRadicacion'] + '.pdf'
                        wRutaDescargaDcto = wRutaDctoPdf +  wNombreDocumento
                        wRutaAlmacenamientoDriveDcto = self.app_config.wRutaAlmPdfDrive + wNombreDocumento
                        
                        if wExiste:
                            if wGuardarDrive:
                                self.fnGuardarDocumento( wNombreDocumento, wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )   
                            #fin if
                        else:
                            self.app_general.fnImprimir( f"No se pudo descargar el estado pdf.",13)
                            self.app_log.error( f"Nro Radicación: {datos['NumeroRadicacion']} - No se pudo descargar el estado pdf.")
                            if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' ):
                                self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' )
                            #fin if  
                        #fin if 
                    #fin if  
                    
                    ## proceso de descarga de providencia - autos - traslados
                    wContador = 0
                    for etiA in wArrEleBusCodIncidencia:
                        # Moviendo scroll
                        self.app_general.fnImprimir(f"Descargando Documento {dato_buscar}",13)
                        etiA.location_once_scrolled_into_view
                        driver.execute_script("window.scrollBy(0, -100);")
                        time.sleep(0.5)
                        
                        # click
                        etiA.click()
                        time.sleep(2)
                        
                        # Validando documento 
                        wNroDocumento = datos['NumeroRadicacion'] + ( '_'+str(wContador) if wContador > 0 else '' ) 
                        wRutaDctoPdf = self.app_config.wRutaDescPdfFin 
                        if wBuscarEnPdf:
                            if buscarProvEst.split('-')[0] == '01': 
                                wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                            elif buscarProvEst.split('-')[1] != '00':
                                wExiste, wRutaDctoPdf = self.fnSeleccionandoModeloEstado( buscarProvEst.split('-')[1], datos['NumeroRadicacion'], driver, datos, json_tbusqueda  ) 
                            #fin if
                        else:
                            wExiste = self.fnValidandoExisteDcto( wNroDocumento )
                        #fin if
                        
                        # Guardando Drive
                        if wExiste:
                            if wGuardarDrive:
                                wNombreDocumento = 'EE-' + wNroDocumento + '.pdf'
                                self.fnGuardarDocumento( wNombreDocumento, wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )   
                            #fin if  
                        #fin if 
                        wContador+=1
                    #fin for 
                #fin if 
            elif wProceso == 11: #Busqueda por numero de radicacion
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => Campo a buscar    
                    datosProceso[1] => GUardar al drive
                """ 
                
                self.app_general.fnImprimir(f"Buscando Nro de Radicado.",11)
                
                # Variables segun parametro
                xpath_colum_busq   = datosProceso[0] 
                wGuardarDrive      = datosProceso[1] if len(datosProceso) > 1 else False 
                xpath_colum_estado = datosProceso[2] if len(datosProceso) > 2 else False  #si este campo se encuentra lleno quiere decir que procedera a guardar el estado
                
                # Variables independientes
                elemento_eti_a          = False
                wArrEleBusCodIncidencia = [] 
                wExiste                 = False
                wRutaDctoPdf            = self.app_config.wRutaDescPdfFin
                wNroTablaSeleccionada   = -1
                wNroFilaSeleccionada    = -1

                try:
                    wConTabla = 0
                    for table in self.fnRetornarElementoTabla( xpath_tab, driver ):
                        wConFila = 0
                        for tr in table.find_elements(By.TAG_NAME,'tr'):
                            if str(tr.text).strip() == "" :
                                continue
                            #fin if 
                            
                            if 'td' in str(xpath_colum_busq):
                                elemento_td = tr.find_element(By.XPATH, xpath_colum_busq)
                            else:
                                elemento_td = tr.find_elements(By.TAG_NAME, 'td')[xpath_colum_busq]
                            #fin if 
                            
                            for eti_a in elemento_td.find_elements(By.TAG_NAME, 'a'):  
                                wdato_limpio = self.fnRetornarRadicadoLimpio( eti_a.text )
                                
                                for clave, valor in json_tbusqueda.items():
                                    if valor == wdato_limpio:
                                        wArrEleBusCodIncidencia.append(eti_a) 
                                        wNroTablaSeleccionada = wConTabla 
                                        
                                        if xpath_colum_estado:
                                            if 'td' in str(xpath_colum_busq):
                                                elemento_columna = tr.find_element(By.XPATH,xpath_colum_estado)
                                                if len(elemento_columna.find_elements(By.TAG_NAME, 'a')):
                                                    wNroFilaSeleccionada = wConFila
                                                #fin if
                                            #fin if 
                                        #fin if 
                                        
                                        break
                                    #fin if 
                                #fin for   
                                
                            #fin for 
                            
                            if len(wArrEleBusCodIncidencia) > 0:
                                break
                            #fin if 
                            
                            wConFila+=1
                        #fin for
                        
                        if len(wArrEleBusCodIncidencia) > 0:
                            break
                        #fin if
                        
                        wConTabla+=1
                    #fin for
                        
                    if len(wArrEleBusCodIncidencia) > 0:
                        
                        # PROCESO DE GUARDAR DOCUMENTO DE ESTADO
                        if xpath_colum_estado:
                            wConTabla = 0
                            
                            # Seleccionamos el enlace de la etique A
                            elemento_tabla   = self.fnRetornarElementoTabla( xpath_tab, driver )[wNroTablaSeleccionada]
                            if wNroFilaSeleccionada < 0:
                                for tr in  elemento_tabla.find_elements(By.TAG_NAME,'tr'):
                                    elemento_columna = tr.find_element(By.XPATH,xpath_colum_estado)
                                    if len(elemento_columna.find_elements(By.TAG_NAME, 'a')):
                                        elemento_eti_a   = elemento_columna.find_element(By.TAG_NAME, 'a')
                                        time.sleep(0.5)
                                        break
                                    #fin if         
                                #fin if 
                            else:
                                elemento_fila    = elemento_tabla.find_elements(By.TAG_NAME,'tr')[wNroFilaSeleccionada]
                                elemento_columna = elemento_fila.find_element(By.XPATH,xpath_colum_estado)
                                elemento_eti_a   = elemento_columna.find_element(By.TAG_NAME, 'a')
                                time.sleep(0.5)       
                            #fin if 
                                                        
                            # Presionamos en el enlace
                            elemento_eti_a.click()
                            time.sleep(2)
                            
                            # Validamos la descarga del documento
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstadoDirecto( '01', elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver )  
                            
                            # Variables para el guardado en el drive
                            wNombreDocumento  = 'ES-' + datos['NumeroRadicacion'] + '.pdf'
                            wRutaDescargaDcto = wRutaDctoPdf +  wNombreDocumento
                            wRutaAlmacenamientoDriveDcto = self.app_config.wRutaAlmPdfDrive + wNombreDocumento
                            
                            if wExiste:
                                if wGuardarDrive: 
                                    self.fnGuardarDocumento( wNombreDocumento, wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )    
                                #fin if
                            else:
                                self.app_general.fnImprimir( f"No se pudo descargar el estado pdf.",13)
                                self.app_log.error( f"Nro Radicación: {datos['NumeroRadicacion']} - No se pudo descargar el estado pdf.")
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' ):
                                    self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' )
                                #fin if  
                            #fin if 
                        #fin if  
                        
                        # PROCESO DE GUARDAR DOCUMENTO DE PROVIDENCIA - AUTO - TRASLADO
                        wContador = 0
                        for etiA in wArrEleBusCodIncidencia:
                            try:
                                # Moviendo scroll
                                etiA.location_once_scrolled_into_view
                                driver.execute_script("window.scrollBy(0, -100);")
                                time.sleep(0.5)
                                
                                # click
                                etiA.click()
                                time.sleep(1)
                                
                                # Validando documento 
                                wNroDocumento = datos['NumeroRadicacion'] + ( '_'+str(wContador) if wContador > 0 else '' ) 
                                wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', etiA, wNroDocumento, json_tbusqueda, datos ) 
                                
                                # Guardando Drive
                                if wExiste:
                                    if wGuardarDrive:
                                        wNombreDocumento  = 'EE-' + wNroDocumento + '.pdf'
                                        self.fnGuardarDocumento( wNombreDocumento, wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )     
                                    #fin if
                                else:
                                    self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                                    self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                                    if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' ):
                                        self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + wNroDocumento + '.pdf' )
                                    #fin if  
                                #fin if 
                            except:
                                pass
                            finally:
                                wContador+=1
                            #fin try
                        #fin for
                    else:
                        self.app_general.fnImprimir( f"{datos['NumeroRadicacion']} - No se encontro Numero de Radicado.", 11)
                        self.app_log.warning( f"{datos['NumeroRadicacion']} - No se encontro Numero de Radicado.")
                    #fin if 
                except:
                    pass
                finally:
                    print()
            elif wProceso == 12: #CALENDARIO SEGUN MES
                """ [ BOGOTA-006 ]   
                    datosProceso[0] => GUardar al drive
                    datosProceso[1] => Por estado o Providencia
                """ 
                
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                # Variables segun parametro
                wGuardarDrive = datosProceso[0] if len(datosProceso) > 0 else False 
                wPorEstadoPro = datosProceso[1] if len(datosProceso) > 1 else False 
                
                # Variables independientes
                wMesBuscar = str(datos['Mes']).upper()
                wDiaBuscar = str(datos['FechaIniciaTermino'].split('-')[-1]).strip()
                
                wArrEleBusCodIncidencia = []
                #PROCESO BUSCAR MES - Obtenemos la fila y la columna 
                wExiste   = False
                wExisteDia = False
                wArrBusqProvidencia = ['AUTO','PROVIDENCIA']
                wNroModelo          = wPorEstadoPro.split('-')[1] 
                wRutaDctoPdf        = self.app_config.wRutaDescPdfFin
                
                xpath_tab = xpath_tab.replace('/div/table/tbody/tr','')
                xpath_tab = xpath_tab.replace('/table/tbody/tr','')
                
                elemento_table = driver.find_element(By.XPATH, xpath_tab)
                
                # for tr in driver.find_elements(By.XPATH,xpath_tab): 
                for tr in elemento_table.find_elements(By.TAG_NAME,'tr'): 
                    for td in tr.find_elements(By.TAG_NAME, 'td'):
                        if str(td.text).strip().split(' ')[0].replace(';','').replace(',','').zfill(2) == wDiaBuscar:
                            if len(td.find_elements(By.TAG_NAME,'a')) > 0:
                                for elemento_a in td.find_elements(By.TAG_NAME,'a'):
                                    if 'ESTADO' not in wPorEstadoPro:
                                        for descProvidencia in wArrBusqProvidencia:
                                            if descProvidencia in elemento_a.text:
                                                wArrEleBusCodIncidencia.append(elemento_a)
                                                break
                                            #fin if 
                                        #fin for
                                        if wArrEleBusCodIncidencia:
                                            break
                                        #fin if
                                    else:
                                        wArrEleBusCodIncidencia.append(elemento_a)
                                        break
                                    #fin if 
                                #fin for 
                            #fin if 
                            
                            wExisteDia = True
                            break
                        #fin if 
                    #fin for
                    
                    if wExisteDia :
                        break
                    #fin if  
                #fin for 
                
                
                if len(wArrEleBusCodIncidencia)>0:
                    # Descargando Dcto Pdf
                    for eleA in wArrEleBusCodIncidencia:
                        self.app_general.fnImprimir(f"Descargando Documento {eleA.text}",13)
                        eleA.location_once_scrolled_into_view
                        driver.execute_script("window.scrollBy(0, -100);")
                        time.sleep(0.5)
                        eleA.click()
                        
                        if 'ESTADO' in wPorEstadoPro: 
                            wExiste, wExisteEnlace, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, eleA, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                        else:
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, eleA, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                        #fin if
                        
                        if wExiste == True and wExisteEnlace == True : # GUARDA LISTA DE ESTADO Y PROVIDENCIA
                            #Guardando Estado
                            self.fnGuardarDocumento( ('ES-' + datos['NumeroRadicacion'] + '.pdf') , wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive ) 
                            
                            #Guardando Providencia 
                            self.fnGuardarDocumento( ('EE-' + datos['NumeroRadicacion'] + '.pdf') , wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive ) 
                            
                        elif wExiste == True and wExisteEnlace == False: # GUARDA SOLO LISTA DE ESTADO
                            self.fnGuardarDocumento( ('ES-' + datos['NumeroRadicacion'] + '.pdf') , wRutaDctoPdf, self.app_config.wRutaAlmPdfDrive )                                
                        else:
                            self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {data_registro['NumeroRadicacion']} - en el documento pdf.",13)
                            self.app_log.error( f"No se encontro el Nro Radicación: {data_registro['NumeroRadicacion']} - en el documento pdf.")
                            if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + "ES-" + wNroDocumento + '.pdf' ):
                                self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + "ES-" + wNroDocumento + '.pdf' )
                            #fin if 
                        #fin if
                         
                    #fin for    
                else:
                    self.app_general.fnImprimir(f"ERROR: No se encontro providencia [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro enlace en la fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}] " )
                    return True
                #fin if  
                 
            elif wProceso == 13:
                """ [ BOGOTA-002 ] 
                    datosProceso[0] => columna donde se va a realizar la busqueda 
                    datosProceso[1] => Guarda en drive
                """ 
                self.app_general.fnImprimir(f"Buscando Nro Radicado",11)
                # Variables segun parametro
                xpath_colum_busq   = datosProceso[0]
                wGuardarDrive      = datosProceso[1] if len(datosProceso) > 1 else False  
                wModeloProvidencia = datosProceso[2] if len(datosProceso) > 2 else False  
                 
                # Variables independientes 
                existe_rd    = False
                wExiste      = False
                wRutaDctoPdf = self.app_config.wRutaDescPdfFin
                for tr in driver.find_elements(By.XPATH, xpath_tab):
                    elemento_eti_a = tr.find_element(By.XPATH, xpath_colum_busq)
                    if len(elemento_eti_a.find_elements(By.TAG_NAME, 'a')) == 0:
                        continue
                    #fin if 
                    
                    for a in elemento_eti_a.find_elements(By.TAG_NAME, 'a'):
                        dato_buscar    = str(a.text).strip().lower()
                        
                        if 'ver estado' in dato_buscar:
                            continue
                        #fin if 
                        
                        if not a.text[:3].isdigit(): 
                            self.app_general.fnImprimir(f"Descargando Documento {dato_buscar}",13)
                            a.location_once_scrolled_into_view
                            driver.execute_script("window.scrollBy(0, -100);")
                            time.sleep(0.5)
                            a.click()
                            time.sleep(2)
                            
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wModeloProvidencia, a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                            
                            if wExiste:
                                existe_rd = True
                                break
                            #fin if 
                        else: 
                            for clave, valor in json_tbusqueda.items():
                                if valor in dato_buscar:
                                            
                                    self.app_general.fnImprimir(f"Descargando Documento {dato_buscar}",13)
                                    a.location_once_scrolled_into_view
                                    driver.execute_script("window.scrollBy(0, -100);")
                                    time.sleep(0.5)
                                    a.click()
                                    time.sleep(2)
                                    
                                    wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wModeloProvidencia, a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                                    
                                    existe_rd = True
                                    break
                                #fin if 
                            #fin for  
                             
                        #fin if  
                        
                        if existe_rd:
                            break
                        #fin if 
                    #fin for
                    
                    if existe_rd :
                        break
                    #fin if 
                #fin for
                 
                if not wExiste:
                    self.app_general.fnImprimir(f"ERROR: No se encontro radicacion [{datos['NumeroRadicacion']}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro providencia [{json_tbusqueda['tbusq-1']}] " )
                    return True
                #fini if 
                
                # Descargando Dcto Pdf 
                if wExiste:                 
                    if wGuardarDrive:
                        self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                        
                        while True:
                            if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                self.app_general.fnEliminarArchivo( wRutaDctoPdf + datos['NumeroRadicacion']  + '.pdf' )
                                self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                break
                            #fin if 
                            time.sleep(1)
                        #fin while
                    #fin if
                else:
                    self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                    self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                    if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                        self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                    #fin if 
                #fin if 
                time.sleep(1)
            elif wProceso == 14: # 1 CALENDARIO  
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => Campo a buscar    
                    datosProceso[1] => GUardar al drive
                    datosProceso[2] => Por estado o Providencia
                """ 
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                
                # Variables segun parametro
                wGuardarDrive = datosProceso[0] if len(datosProceso) > 0 else False 
                wPorEstadoPro = datosProceso[1] if len(datosProceso) > 1 else False 
                
                # Variables independientes
                wMesBuscar     = str(datos['Mes']).upper()
                wDiaBuscar     = str(datos['FechaIniciaTermino'].split('-')[-1]).strip() 
                row            = 1 
                elemento_eti_a = False
                existe_dia     = False
                wExiste        = False
                wRutaDctoPdf   = self.app_config.wRutaDescPdfFin
                wNroModelo     = wPorEstadoPro.split('-')[1] 
                
                for tr in driver.find_elements(By.XPATH,xpath_tab):
                    col = 1
                    for td in tr.find_elements(By.TAG_NAME, 'td'):
                        if str(td.text).strip().zfill(2) == wDiaBuscar.zfill(2):
                            if len(td.find_elements(By.TAG_NAME,'a')) > 0 :
                                elemento_eti_a = td.find_element(By.TAG_NAME,'a')
                            #fin if  
                            
                            existe_dia = True
                            break 
                        #fin if 
                        col+=1
                    #fin if 
                    
                    if existe_dia :
                        break
                    #fin if 
                    row+=1
                #fin for 
                
                if elemento_eti_a:  
                    # Presionando Click
                    elemento_eti_a.click() 
                    time.sleep(2)
                    
                    if wPorEstadoPro:
                        if 'ESTADO' in wPorEstadoPro: 
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                        else:
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                        #fin if   
                    #fin if
                    
                    # Guardando Drive
                    if wExiste:
                        if wGuardarDrive:
                            self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                            
                            while True:
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                    self.app_general.fnEliminarArchivo( wRutaDctoPdf+ datos['NumeroRadicacion']  + '.pdf' )
                                    self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                    break
                                #fin if 
                                time.sleep(1)
                            #fin while
                        #fin if
                    else:
                        self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                        self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                        if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                            self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                        #fin if 
                    #fin if 
                else:
                    self.app_general.fnImprimir(f"ERROR: No se encontro fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro enlace en la fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}] " )
                    return True
                #fin if 
            elif wProceso == 15: # 1 CALENDARIO POR TABLA varia fila segun el mes  
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => Campo a buscar    
                    datosProceso[1] => GUardar al drive
                    datosProceso[2] => Por estado o Providencia
                """ 
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                
                # Variables segun parametro
                wGuardarDrive = datosProceso[0] if len(datosProceso) > 0 else False 
                wPorEstadoPro = datosProceso[1] if len(datosProceso) > 1 else False 
                 
                # Variables independientes
                wDiaBuscar     = str(datos['FechaIniciaTermino'].split('-')[-1]).strip()
                wMesBuscar     = str(datos['FechaIniciaTermino'].split('-')[1]).strip() 
                row            = 1 
                elemento_eti_a = False
                existe_dia     = False
                wExiste        = False
                wRutaDctoPdf   = self.app_config.wRutaDescPdfFin
                wNroModelo     = wPorEstadoPro.split('-')[1] 
                
                mes = 1
                for table in self.fnRetornarElementoTabla( xpath_tab, driver ):
                    try:
                        if str(mes).zfill(2) != wMesBuscar:
                            continue
                        #fin if 
                        
                        for tr in table.find_elements(By.TAG_NAME,"tr"): 
                            for td in tr.find_elements(By.TAG_NAME, 'td'):
                                if (td.get_attribute('class') != 'semana' and (str(td.text).strip().zfill(2) == wDiaBuscar.zfill(2))):
                                    if len(td.find_elements(By.TAG_NAME,'a')) > 0 :
                                        elemento_eti_a = td.find_element(By.TAG_NAME,'a')
                                    #fin if  
                                    
                                    existe_dia = True
                                    break 
                                #fin if  
                            #fin if 
                            
                            if existe_dia :
                                break
                            #fin if 
                        #fin for
                        
                        break
                    except:
                        pass
                    finally:
                        mes+=1
                    #fin try
                #fin for 
                
                if elemento_eti_a: 
                    
                    # Presionando Click 
                    self.app_general.fnImprimir(f"Descargando Documento ",13)
                    elemento_eti_a.location_once_scrolled_into_view
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.5)
                    elemento_eti_a.click()
                    time.sleep(2)
                    
                    if wPorEstadoPro:
                        if 'ESTADO' in wPorEstadoPro: 
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                        else:
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                        #fin if
                    #fin if 
                    
                    # Guardando Drive
                    if wExiste:
                        if wGuardarDrive:
                            self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                            
                            while True:
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                    self.app_general.fnEliminarArchivo( wRutaDctoPdf+ datos['NumeroRadicacion']  + '.pdf' )
                                    self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                    break
                                #fin if 
                                time.sleep(1)
                            #fin while
                        #fin if
                    else:
                        self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                        self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                        if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                            self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                        #fin if 
                    #fin if 
                else:
                    self.app_general.fnImprimir(f"ERROR: No se encontro fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro enlace en la fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}] " )
                    return True
                #fin if 
            elif wProceso == 16: # 1 CALENDARIO POR TABLA varia fila segun el mes  
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => Campo a buscar    
                    datosProceso[1] => GUardar al drive
                    datosProceso[2] => Por estado o Providencia
                """  
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                
                # Variables segun parametro
                wGuardarDrive = datosProceso[0] if len(datosProceso) > 0 else False 
                wPorEstadoPro = datosProceso[1] if len(datosProceso) > 1 else False 
                 
                # Variables independientes
                wDiaBuscar     = str(datos['FechaIniciaTermino'].split('-')[-1]).strip() 
                wNroMesBuscar  = str(datos['FechaIniciaTermino'].split('-')[1]).strip() 
                wMesBuscar     = str(datos['Mes']).strip().upper()
                elemento_eti_a = False  
                posiciones     = { '01' : [0,7],'02' : [8,15], '03' : [0,7], '04' : [8,15], '05' : [0,7], '06' : [8,15], '07' : [0,7], '08' : [8,15], '09' : [0,7], '10' : [8,15], '11' : [0,7], '12' : [8,15] }
                existe_dia     = False
                wExiste        = False
                wRutaDctoPdf   = self.app_config.wRutaDescPdfFin
                wNroModelo     = wPorEstadoPro.split('-')[1] 
 
                # una vez encontrado el mes deberia solo sumar maximo 8 filas
                elemento_mes_encontrado = [ (index, tr) for index, tr in enumerate(driver.find_elements(By.XPATH, xpath_tab)) if wMesBuscar in str(tr.text).strip().upper() ]
                
                if len(elemento_mes_encontrado) > 0:
                    for fila in range( elemento_mes_encontrado[0][0]+1, elemento_mes_encontrado[0][0]+8):
                        tr = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[4]/div[1]/div[2]/div/div[2]/div/div/div/div[1]/table/tbody/tr['+str(fila)+']')
                        for td in range(posiciones[wNroMesBuscar][0], posiciones[wNroMesBuscar][1]):
                            elemento_td = tr.find_element(By.XPATH,'td['+str(td+1)+']') 
                            if (elemento_td.get_attribute('class') != 'semana' and (str(elemento_td.text).strip().zfill(2) == wDiaBuscar.zfill(2))):
                                if len(elemento_td.find_elements(By.TAG_NAME,'a')) > 0 :
                                    elemento_eti_a = elemento_td.find_element(By.TAG_NAME,'a')
                                #fin if  
                                existe_dia = True
                                break 
                            #fin if  
                        #fin for
                        
                        if existe_dia: 
                            break
                        #fin if 
                    #fin for 
                #fin if 
                 
                
                if elemento_eti_a:
                    
                    # Presionando Click 
                    self.app_general.fnImprimir(f"Descargando Documento ",13)
                    elemento_eti_a.location_once_scrolled_into_view
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.5)
                    elemento_eti_a.click()
                    time.sleep(2)
                     
                    if wPorEstadoPro:
                        if 'ESTADO' in wPorEstadoPro: 
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                        else:
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                        #fin if
                    #fin if 
                    
                    # Guardando Drive
                    if wExiste:
                        if wGuardarDrive:
                            self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                            
                            while True:
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                    self.app_general.fnEliminarArchivo( wRutaDctoPdf+ datos['NumeroRadicacion']  + '.pdf' )
                                    self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                    break
                                #fin if 
                                time.sleep(1)
                            #fin while
                        #fin if
                    else:
                        self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                        self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                        if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                            self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                        #fin if 
                    #fin if 
                else:
                    self.app_general.fnImprimir(f"ERROR: No se encontro fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro enlace en la fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}] " )
                    return True
                #fin if 
            elif wProceso == 17: # 1 CALENDARIO POR TABLA varia fila segun el mes  
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => Campo a buscar    
                    datosProceso[1] => GUardar al drive
                    datosProceso[2] => Por estado o Providencia
                """ 
                
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                
                # Variables segun parametro
                wGuardarDrive = datosProceso[0] if len(datosProceso) > 0 else False 
                wPorEstadoPro = datosProceso[1] if len(datosProceso) > 1 else False 
                 
                # Variables independientes
                wDiaBuscar     = str(datos['FechaIniciaTermino'].split('-')[-1]).strip() 
                wNroMesBuscar  = str(datos['FechaIniciaTermino'].split('-')[1]).strip() 
                wMesBuscar     = str(datos['Mes']).strip().upper()
                elemento_eti_a = False 
                existe_dia     = False
                wExiste        = False
                wRutaDctoPdf   = self.app_config.wRutaDescPdfFin
                wNroModelo     = wPorEstadoPro.split('-')[1] 
                 
                fila = 1
                for tr in driver.find_elements(By.XPATH, xpath_tab ):
                    try:
                        columna = 1
                        for td in tr.find_elements(By.TAG_NAME,'td'):
                            if str(td.text).strip().split(' ')[0].zfill(2) == wDiaBuscar:
                                if len(td.find_elements(By.TAG_NAME,'a')) > 0 :
                                    elemento_eti_a = td.find_element(By.TAG_NAME,'a')
                                #fin if  
                                existe_dia = True 
                                break
                            #fin if
                            columna+=1 
                        #fin for
                        
                        if existe_dia:
                            break
                        #fin if
                    except:
                        pass
                    finally:
                        fila+=1
                    #fin try
                #fin for
                
                if elemento_eti_a:  
                    # Presionando Click 
                    self.app_general.fnImprimir(f"Descargando Documento ",13)
                    elemento_eti_a.location_once_scrolled_into_view
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.5)
                    elemento_eti_a.click()
                    time.sleep(1)
                    
                    if wPorEstadoPro:
                        if 'ESTADO' in wPorEstadoPro: 
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                        else:
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                        #fin if
                    #fin if 
                    
                    # Guardando Drive
                    if wExiste:
                        if wGuardarDrive:
                            self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                            
                            while True:
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                    self.app_general.fnEliminarArchivo( wRutaDctoPdf+ datos['NumeroRadicacion']  + '.pdf' )
                                    self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                    break
                                #fin if 
                                time.sleep(1)
                            #fin while
                        #fin if
                    else:
                        self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                        self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                        if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                            self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                        #fin if 
                    #fin if 
                else:
                    self.app_general.fnImprimir(f"ERROR: No se encontro fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro enlace en la fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}] " )
                    return True
                #fin if 
            elif wProceso == 18: # 1 CALENDARIO POR TABLA varia fila segun el mes  
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => Campo a buscar    
                    datosProceso[1] => GUardar al drive
                    datosProceso[2] => Por estado o Providencia
                """ 
                
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                
                # Variables segun parametro
                wGuardarDrive = datosProceso[0] if len(datosProceso) > 0 else False 
                wPorEstadoPro = datosProceso[1] if len(datosProceso) > 1 else False 
                 
                # Variables independientes
                wDiaBuscar    = str(datos['FechaIniciaTermino'].split('-')[-1]).strip() 
                wNroMesBuscar = str(datos['FechaIniciaTermino'].split('-')[1]).strip() 
                wMesBuscar    = str(datos['Mes']).strip().upper()
                elemento_eti_a = False 
                existe_dia     = False
                wExiste        = False
                wRutaDctoPdf   = self.app_config.wRutaDescPdfFin
                wNroModelo     = wPorEstadoPro.split('-')[1] 
                 
                fila = 1
                for tr in driver.find_elements(By.XPATH, xpath_tab ):
                    try:
                        columna = 1
                        for td in tr.find_elements(By.TAG_NAME,'td'):
                            if str(td.text).strip().split(' ')[0].zfill(2) == wDiaBuscar:
                                if len(td.find_elements(By.TAG_NAME,'a')) > 0 :
                                    elemento_eti_a = td.find_element(By.TAG_NAME,'a')
                                #fin if  
                                existe_dia = True
                                break
                            #fin if
                            columna+=1 
                        #fin for
                        
                        if existe_dia:
                            break
                        #fin if
                    except:
                        pass
                    finally:
                        fila+=1
                    #fin try
                #fin for
                
                if elemento_eti_a:  
                    # Presionando Click 
                    self.app_general.fnImprimir(f"Descargando Documento ",13)
                    elemento_eti_a.location_once_scrolled_into_view
                    driver.execute_script("window.scrollBy(0, -100);")
                    time.sleep(0.5)
                    elemento_eti_a.click()
                    
                    if wPorEstadoPro:
                        if 'ESTADO' in wPorEstadoPro: 
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                        else:
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                        #fin if
                    #fin if 
                    
                    # Guardando Drive
                    if wExiste:
                        if wGuardarDrive:
                            self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                            
                            while True:
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                    self.app_general.fnEliminarArchivo( wRutaDctoPdf+ datos['NumeroRadicacion']  + '.pdf' )
                                    self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                    break
                                #fin if 
                                time.sleep(1)
                            #fin while
                        #fin if
                    else:
                        self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                        self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                        if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                            self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                        #fin if 
                    #fin if 
                else:
                    self.app_general.fnImprimir(f"ERROR: No se encontro fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro enlace en la fecha [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}] " )
                    return True
                #fin if 
            elif wProceso == 19: # MAS DE 1 CALENDARIO EN UNA SOLA COLUMNA  
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => GUardar al drive
                    datosProceso[1] => Por estado o Providencia 
                """ 
                
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                
                # Variables segun parametro
                wGuardarDrive = datosProceso[0] if len(datosProceso) > 0 else False 
                wPorEstadoPro = datosProceso[1] if len(datosProceso) > 1 else False 
                
                # Variables independientes
                wMesBuscar   = str(datos['Mes']).upper()
                wDiaBuscar   = str(datos['FechaIniciaTermino'].split('-')[-1]).strip() 
                xpath_tab   += '['+str(int(str(datos['FechaIniciaTermino'].split('-')[1]).strip()))+']'   
                wExisteDia   = False
                wExiste      = False
                wRutaDctoPdf = self.app_config.wRutaDescPdfFin
                wNroModelo   = wPorEstadoPro.split('-')[1] 
                
                for tr in driver.find_elements(By.XPATH, xpath_tab +'/table/tbody/tr')[1:]:
                    for td in tr.find_elements(By.TAG_NAME,'td'):
                        if str(td.text).strip() == wDiaBuscar and td.get_attribute('class') != 'wno':
                            if len(td.find_elements(By.TAG_NAME,'a')) > 0:
                                elemento_eti_a = td.find_element(By.TAG_NAME,'a')
                                elemento_eti_a.click() 
                                time.sleep(2)
                                
                                
                                if wPorEstadoPro:
                                    if 'ESTADO' in wPorEstadoPro: 
                                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                                    else:
                                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                                    #fin if
                                #fin if 
                                                
                                # Guardando Drive
                                if wExiste:
                                    if wGuardarDrive:
                                        self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                                        
                                        while True:
                                            if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                                self.app_general.fnEliminarArchivo( wRutaDctoPdf+ datos['NumeroRadicacion']  + '.pdf' )
                                                self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                                break
                                            #fin if 
                                            time.sleep(1)
                                        #fin while
                                    #fin if  
                                    break
                                else: 
                                    self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                                    self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                                    if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                                        self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                                    #fin if 
                                #fin if 
                            else:
                                self.app_general.fnImprimir("Se encontro Fecha, pero no cuenta con enlace para descarga.")
                            #fin if 
                            wExisteDia = True
                            break
                        #fin if 
                    #fin for
                    if wExisteDia:
                        break
                    #fin if
                #fin for 
            elif wProceso == 20: #es similar al 3
                """ [ BOGOTA-004 ]
                    datosProceso[0] => columna busqueda
                    datosProceso[1] => columna click
                    datosProceso[2] => Guardar en drive 
                """ 
                # Variables
                xpath_colum_busq  = datosProceso[0]
                xpath_colum_click = datosProceso[1]
                wGuardarDrive     = datosProceso[2] if len(datosProceso) > 2 else False 
                elemento          = []
                existe            = False 
                
                ## Descargar segun Numero de radicación
                self.app_general.fnImprimir(f"Buscando Nro Radicado",11)
                
                for table in self.fnRetornarElementoTabla( xpath_tab, driver ):
                    for tr in table.find_elements(By.XPATH,'tbody/tr'):
                        if  len(tr.find_elements(By.XPATH,'td')) < int(re.findall(r'\[(\d+)\]', xpath_colum_busq)[0]):
                            continue
                        #fin if 
                        
                        elemento_td = tr.find_element(By.XPATH, xpath_colum_busq)
                        # buscando a nivel de radicación
                        for eti_a in elemento_td.find_elements(By.TAG_NAME, 'a'):
                            if (not eti_a.text[:3].isdigit() or len(str(eti_a.text).strip()) == 4) :
                                continue
                            #fin if 
                            
                            # caso1 -> -01234    -> RES: AÑO-01234
                            if str(eti_a.text).strip()[0] == '-':
                                wdato_limpio = datos['Anio'] + str(eti_a.text).strip()
                            #fin if
                            
                            elif str(eti_a.text).strip().index("-") == 4:
                                wdato = str(eti_a.text).strip().split(' ')[0]
                                wdato_limpio = wdato.replace(wdato.split('-')[-1],wdato.split('-')[-1].zfill(5))
                            # caso3 -> radicado completo
                            else:
                                wdato_limpio = str(eti_a.text).strip().split(' ')[0]
                            #fin if 
                             
                            for clave, valor in json_tbusqueda.items():
                                if valor in wdato_limpio:
                                    elemento.append(eti_a)
                                    break
                                #fin if 
                            #fin for
                            
                            if len(elemento) > 0:
                                break
                            #fin if   
                        #fin for 
                        
                        if len(elemento) > 0:
                            break
                        #fin if   
                    #fin for  
                    
                    if len( elemento ) > 0: 
                        try:
                            elemento[0].click()
                            existe = True
                        except:
                            pass
                        #fin try
                         
                        break
                    #fin if
                #fin for
                
                if existe:
                    wExiste = self.fnValidandoExisteDcto( datos['NumeroRadicacion'] )
                    
                    if wExiste:
                        if wGuardarDrive:
                            self.app_general.fnCopiarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                            
                            while True:
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                    self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion']  + '.pdf' )
                                    self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                    break
                                #fin if 
                                time.sleep(1)
                            #fin while
                        #fin if 
                    #fin if
                else:
                    self.app_general.fnImprimir(f"No se encontro Nro de Radicado.",13)
                #fin if
            elif wProceso == 21:
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => GUardar al drive
                    datosProceso[1] => Por estado o Providencia 
                """ 
                
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                
                # Variables segun parametro
                wGuardarDrive = datosProceso[0] if len(datosProceso) > 0 else False  
                wPorEstadoPro = datosProceso[1] if len(datosProceso) > 1 else False 
                meses_dict    = datosProceso[2] if len(datosProceso) > 2 else False 
                
                # Variables independientes
                wDiaBuscar   = str(datos['FechaIniciaTermino'].split('-')[-1]).strip()  
                xpath_tab    = xpath_tab.replace('FIL', str(meses_dict[datos['Mes'].lower()][0]) ).replace('COL', str(meses_dict[datos['Mes'].lower()][1]) ) 
                wExiste      = False
                wRutaDctoPdf = self.app_config.wRutaDescPdfFin
                wNroModelo   = wPorEstadoPro.split('-')[1] 
                
                for tr in driver.find_elements(By.XPATH, xpath_tab +'/table/tbody/tr')[1:]:
                    for td in tr.find_elements(By.TAG_NAME,'td'):
                        if str(td.text).strip() == wDiaBuscar and td.get_attribute('class') != 'wno':
                            if len(td.find_elements(By.TAG_NAME,'a')) > 0:
                                elemento_eti_a = td.find_element(By.TAG_NAME,'a')
                                elemento_eti_a.click()
                                time.sleep(2)
                                
                                if wPorEstadoPro:
                                    if 'ESTADO' in wPorEstadoPro: 
                                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                                    else:
                                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                                    #fin if 
                                #fin if 
                                                
                                # Guardando Drive
                                if wExiste:
                                    if wGuardarDrive:
                                        self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                                        
                                        while True:
                                            if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                                self.app_general.fnEliminarArchivo( wRutaDctoPdf+ datos['NumeroRadicacion']  + '.pdf' )
                                                self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                                break
                                            #fin if 
                                            time.sleep(1)
                                        #fin while
                                    #fin if   
                                else: 
                                    self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                                    self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                                    if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                                        self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                                    #fin if 
                                #fin if 
                            else:
                                self.app_general.fnImprimir("Se encontro Fecha, pero no cuenta con enlace para descarga.")
                            #fin if 
                        #fin if 
                            
                        break
                    #fin for
                #fin for 
            elif wProceso == 22: #Recorriendo por filas las providencias 
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => columna donde se va a realizar la busqueda
                    datosProceso[1] => Guarda en drive
                """  
                xpath_colum_click = datosProceso[0] 
                wGuardarDrive     = datosProceso[1] if len(datosProceso) > 1 else False 
                empieza           = datosProceso[2] if len(datosProceso) > 2 else 1 
                tipo_dcto         = datosProceso[3] if len(datosProceso) > 3 else 'providencia' 
                wExiste           = False
                wBuscarPdf        = False
                wRutaDctoPdf      = '' 
                
                
                for table in self.fnRetornarElementoTabla( xpath_tab, driver ):
                    for tr in table.find_elements(By.TAG_NAME,'tr')[empieza:]: 
                        if len(tr.find_elements(By.XPATH,'td')) >= int(re.findall(r'\[(\d+)\]', xpath_colum_click)[0]) :
                            elemento_td = tr.find_elements(By.XPATH, xpath_colum_click) 
                            if len( elemento_td.find_elements(By.TAG_NAME, 'a') ) > 0:
                                for elemento_a in elemento_td.find_element(By.TAG_NAME, 'a'):
                                    if tipo_dcto == 'providencia':
                                        if (('estado' in str(elemento_a.get_attribute('href')).lower()) or ('/est+' in str(elemento_a.get_attribute('href')).lower)) :
                                            continue
                                        #fin if
                                    #fin if 
                                    
                                    elemento_a.click()
                                    time.sleep(1)
                                    
                                    if tipo_dcto == 'providencia':
                                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', elemento_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                                    else:
                                        wNroModelo = tipo_dcto.split("-")[1] #estado-01
                                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, elemento_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                                    #fin if 
                                    
                                    if wExiste:
                                        break
                                    #fin if 
                                #fin for 
                                
                                if wExiste:
                                    break
                                #fin if 
                            #fin if 
                        #fin if 
                    #fin for
                    if wExiste:
                        break
                    #fin if 
                #fin for
                 
                if not wExiste:
                    self.app_general.fnImprimir(f"ERROR: No se encontro providencia [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}]",13)
                    self.app_log.error( f"NRO RADICACION: {datos['NumeroRadicacion']} - MSJ: No se encontro providencia [{self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}] " )
                    return True
                #fini if 
                  
                if wExiste:
                    if wGuardarDrive:
                        self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                        
                        while True:
                            if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                self.app_general.fnEliminarArchivo( wRutaDctoPdf + datos['NumeroRadicacion']  + '.pdf' )
                                self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                break
                            #fin if 
                            time.sleep(1)
                        #fin while
                    #fin if  
                #fin if 
            elif wProceso == 23:
                """ [ BOGOTA-006 ] 
                    datosProceso[0] => GUardar al drive
                    datosProceso[1] => Por estado o Providencia 
                """  
                self.app_general.fnImprimir(f"Buscando Fecha {self.app_general.fnConvertirFecha(datos['FechaIniciaTermino'],'-','/')}",11)
                
                # Variables segun parametro
                wGuardarDrive = datosProceso[0] if len(datosProceso) > 0 else False  
                wPorEstadoPro = datosProceso[1] if len(datosProceso) > 1 else False 
                meses_dict    = datosProceso[2] if len(datosProceso) > 2 else False 
                
                # Variables independientes
                wDiaBuscar   = str(datos['FechaIniciaTermino'].split('-')[-1]).strip()  
                xpath_tab    = xpath_tab.replace('FIL', str(meses_dict[datos['Mes'].lower()][0]) ).replace('COL', str(meses_dict[datos['Mes'].lower()][1]) )
                wExiste      = False
                wRutaDctoPdf = self.app_config.wRutaDescPdfFin
                wNroModelo   = wPorEstadoPro.split('-')[1] 
                
                for td in range(meses_dict[datos['Mes'].lower()][1], meses_dict[datos['Mes'].lower()][1] + 7):
                    elemento_td = driver.find_element(By.XPATH, xpath_tab + '/td['+str(td)+']' ) 
                    if str(elemento_td.text).strip() == wDiaBuscar and td.get_attribute('class') != 'wno':
                        if len(elemento_td.find_elements(By.TAG_NAME,'a')) > 0:
                            elemento_eti_a = elemento_td.find_element(By.TAG_NAME,'a')
                            elemento_eti_a.click()
                            time.sleep(2)
                             
                            if wPorEstadoPro:
                                if 'ESTADO' in wPorEstadoPro: 
                                    wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                                else:
                                    wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, wNroModelo, elemento_eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                                #fin if
                            #fin if 
                            
                            if wExiste:
                                if wGuardarDrive:
                                    self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                                    
                                    while True:
                                        if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                            self.app_general.fnEliminarArchivo( wRutaDctoPdf+ datos['NumeroRadicacion']  + '.pdf' )
                                            self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                            break
                                        #fin if 
                                        time.sleep(1)
                                    #fin while
                                #fin if 
                            else: 
                                self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                                self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                                    self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                                #fin if 
                            #fin if
                        #fin if 
                        
                        break
                    #fin if 
                #fin for
            elif wProceso == 24:   
                """ 
                    datosProceso[0] => columna donde se va a realizar la busqueda
                    datosProceso[1] => columna donde se va a realizar el click 
                    datosProceso[2] => Separador para la fecha
                    datosProceso[3] => Numero la cual desea que empiece el for a buscar 
                    datosProceso[4] => True guardar directo al drive, False no 
                """
                
                #'/',0,0,False, False, 1, False
                wColBusqueda  = datosProceso[0] if len( datosProceso ) > 0 else 0      # Columna donde se encuentra la fecha a buscar
                wColEnlace    = datosProceso[1] if len( datosProceso ) > 1 else 0      # Columna donde se encuentra el enlace a buscar
                wModeloDctos  = datosProceso[2] if len( datosProceso ) > 2 else 0      # Nro de modelos de documetos ( Estado y Providencias )
                wSeparador    = datosProceso[3] if len( datosProceso ) > 3 else '/'    # Separador 
                wInicioFor    = datosProceso[4] if len( datosProceso ) > 4 else 0      # Inicio de For
                wGuardarDrive = datosProceso[5] if len( datosProceso ) > 5 else False  # Guardar en la carpeta del Drive 

                # Descargando Segun fecha
                self.app_general.fnImprimir(f"Descargando Documento",9)
                continuar, nro_estado = self.fnDescargarPdfPorFechaProc24( driver, datos, xpath_tab, wColBusqueda, wColEnlace, wModeloDctos, wSeparador, wInicioFor,  wGuardarDrive ) 
                if continuar: wContinuar = True
            elif wProceso == 25:
                wGuardarDrive  = datosProceso[0] if len( datosProceso ) > 0 else False      # Guardar en el drive
                
                elemento_tmp = driver.find_element(By.XPATH, xpath_tab)
                elemento_eti_p = elemento_tmp.find_elements(By.TAG_NAME,'p')
                
                for eti_p in elemento_eti_p:
                    eti_a = eti_p.find_element(By.TAG_NAME, 'a')
                    eti_a.click()
                    
                    wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                              
                    if wExiste : 
                        if wGuardarDrive:
                            self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                            
                            while True:
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                    self.app_general.fnEliminarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf' )
                                    self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                    break
                                #fin if 
                                time.sleep(1)
                            #fin while
                        #fin if 
                        
                        break
                    else:
                        self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                        self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                        if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                            self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                        #fin if 
                    #fin if
                #fin for 
            elif wProceso == 26:
                wGuardarDrive  = datosProceso[0] if len( datosProceso ) > 0 else 0      # Guardar en el drive
                
                elemento_tmp = driver.find_element(By.XPATH, xpath_tab)
                elemento_eti_a = elemento_tmp.find_elements(By.TAG_NAME,'a')
                wArrExisteRadicado = []
                for eti_a in elemento_eti_a:
                    wdato = eti_a.text
                    for clave, valor in json_tbusqueda.items():
                        if valor in wdato:
                            wArrExisteRadicado.append(eti_a) 
                            eti_a.click() 
                            time.sleep(2)
                            
                            wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                                    
                            if wExiste : 
                                if wGuardarDrive:
                                    self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                                    
                                    while True:
                                        if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                            self.app_general.fnEliminarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf' )
                                            self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                            break
                                        #fin if 
                                        time.sleep(1)
                                    #fin while
                                #fin if 
                                
                                break
                            else:
                                self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                                self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                                if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                                    self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                                #fin if 
                            #fin if
                            
                            break # por el momento solo obtener el primero
                        #fin if 
                    #fin for  
                    
                    if len(wArrExisteRadicado) > 0:
                        break
                    # eti_a = eti_p.find_element(By.TAG_NAME, 'a')
                    # eti_a.click() 
                #fin for
            elif wProceso == 27:   # MODELO ARTICLE
                elemento_article = driver.find_element(By.XPATH, xpath_tab.replace('MES', datos['FechaIniciaTermino'].split('-')[2] ))
                eti_a = elemento_article.find_element(By.TAG_NAME,'a')
                
                eti_a.click()
                time.sleep(2)
                
                wExiste, wRutaDctoPdf = self.fnProcesoDescargaDcto( driver, '01', eti_a, datos['NumeroRadicacion'], json_tbusqueda, datos ) 
                                    
                if wExiste : 
                    if wGuardarDrive:
                        self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                        
                        while True:
                            if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                self.app_general.fnEliminarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf' )
                                self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                break
                            #fin if 
                            time.sleep(1)
                        #fin while
                    #fin if  
                else:
                    self.app_general.fnImprimir( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.",13)
                    self.app_log.error( f"No se encontro el Nro Radicación: {datos['NumeroRadicacion']} - en el documento pdf.")
                    if self.app_general.fnExisteArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' ):
                        self.app_general.fnEliminarArchivo( self.app_config.wRutaDescPdfFin + datos['NumeroRadicacion'] + '.pdf' )
                    #fin if 
                #fin if
            elif wProceso == 28:   # MODELO LI
                click_estado  = datosProceso[0] if len( datosProceso ) > 0 else 0      # Tipo de Documento y nro
                
                for eti_li in driver.find_elements(By.XPATH, xpath_tab):
                    wFechaEtiqueta = str(eti_li.text).split('(')[0].replace('Traslado','').strip()
                    if self.app_general.fnRetornarFechaLimpia( wFechaEtiqueta ) in self.fnTipoBusquedaFecha( datos ):
                        eti_a = eti_li.find_element(By.TAG_NAME,'a')
                        eti_a.click()
                        time.sleep(2)
                        
                        wExiste, wRutaDctoPdf = self.fnProcesoDescargaDctoEstado( click_estado.split('-')[1], eti_a,  datos['NumeroRadicacion'], json_tbusqueda, datos, driver ) 
                             
                        if wExiste : 
                            if wGuardarDrive:
                                if wRutaDctoPdf != '':
                                    self.app_general.fnCopiarArchivo( wRutaDctoPdf + datos['NumeroRadicacion'] + '.pdf', self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf' )
                                    
                                    while True:
                                        if self.app_general.fnExisteArchivo( self.app_config.wRutaAlmPdfDrive + datos['NumeroRadicacion'] + '.pdf'  ):
                                            self.app_general.fnEliminarArchivo( wRutaDctoPdf + wNroDocumento + '.pdf' )
                                            self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                                            break
                                        #fin if 
                                        time.sleep(1)
                                    #fin while
                                #fin if 
                            #fin if 
                        #fin if 
                        
                        break
                    #fin if 
                #FIN FOR
                  
            #FIN IF  
        except:
            self.app_log.error( f"ERROR fnProcesos: {datos['NumeroRadicacion']} - NroProceso: {wProceso} - MSJ ERROR: {sys.exc_info()[1]}" )
            wContinuar = True
            pass
        finally:
            time.sleep(1)
        #fin try
        
        return wContinuar
    ######## fin Procesos
    
    ############ Modelo de Estados
    def fnBuscarEstadoModelo01( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaDocumentoPdf = self.app_config.wRutaDescPdfFin 
        wRutaArchivo      = self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf'
        wNroRadicacion    = str(wNombreDocumento).split('_')[0]
        wExisteRadicado   = False 
        wExisteEnlace     = False 
        wArrDatoBusqueda  = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
        #fin if
        
        if wExisteEnlace:
            wRutaDocumentoPdf = self.app_config.wRutaDescPdfSel
        #fin if
        
        return wExisteRadicado, wExisteEnlace, wRutaDocumentoPdf
    ######## fnBuscarEstadoModelo01
    
    def fnBuscarEstadoModelo03( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDcto        = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )  
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado :
            ObjPdfreader = PdfReader(wRutaArchivo)

            init_Page = 1
            num_pages = len(ObjPdfreader.pages)
            wExisteRadicado = False
                                
            while init_Page <= num_pages:
                
                # en paraqmetro de Pages debe de cambiar 1, 2, 3, 4, 5, 6 
                dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
                time.sleep(1) 
                
                # get information of tables 
                for item in dfs[ 1 if init_Page == 1 else 0].values.tolist()[0:]:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                    
                    #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                    wNroRadicado = '' if pd.isna(item[0]) else item[0].replace('\r',' ') 
                    
                    if wNroRadicado in wArrDatoBusqueda:
                        wClase         = '' if pd.isna(item[1]) else item[1].replace('\r',' ')
                        wDemandante    = '' if pd.isna(item[2]) else item[2].replace('\r',' ')
                        wDemandado     = '' if pd.isna(item[3]) else item[3].replace('\r',' ')
                        wFechaAuto     = '' if pd.isna(item[4]) else item[4].replace('\r',' ')
                        wAutoAnotacion = '' if pd.isna(item[5]) else item[5].replace('\r',' ')  
                        
                        wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion )
                        
                        if wExisteEnlace:
                            wRutaDcto = self.app_config.wRutaDescPdfSel
                        #fin if 
                        
                        wExisteRadicado = True
                        break
                    #fin if  
                #fin for 
                
                if wExisteRadicado: 
                    break
                #fin if
                
                init_Page+=1
            #fin while 
        #fin if     
        
        return wExisteRadicado, wExisteEnlace,  wRutaDcto
    ######## fnBuscarEstadoModelo03 
    
    def fnBuscarEstadoModelo06( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDcto        = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )  
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11) 
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado :
            ObjPdfreader = PdfReader(wRutaArchivo) 
            init_Page = 1
            num_pages = len(ObjPdfreader.pages)
            wExisteRadicado   = False
                                
            while init_Page <= num_pages:
                    
                dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
                time.sleep(1)
                
                if init_Page == 1:
                    wNroRadicacion = str(dfs[1].columns[1]) + '-' + str(dfs[1].columns[2])
                    wClase         = dfs[1].columns[3]
                    wDemandante    = dfs[1].columns[4]
                    wDemandado     = dfs[1].columns[6]
                    wFechaAuto     = dfs[1].columns[7] #13 DE DICIEMBRE 23
                #fin if 
                
                if wNroRadicacion in wArrDatoBusqueda:
                    # wExiste = True
                    wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion )
                    
                    if wExisteEnlace:
                        wRutaDcto = self.app_config.wRutaDescPdfSel
                    #fin if 
                    wExisteRadicado = True 
                #fin if 
                
                if not wExisteRadicado: 
                    conta = 0
                    for item in dfs[ 1 if init_Page == 1 else 0].values.tolist()[0:]:
                        
                        #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                        wNroRadicacion = '' if pd.isna(item[0]) else item[0].replace('\r',' ') 
                        
                        if wNroRadicacion in wArrDatoBusqueda:
                            wNroRadicacion = '' if pd.isna(item[1]) else str(item[1]).replace('\r',' ') + '-' + ( '' if pd.isna(item[2]) else str(item[2]).replace('\r',' ') )
                            wClase         = '' if pd.isna(item[3]) else item[3].replace('\r',' ')
                            wDemandante    = '' if pd.isna(item[4]) else item[4].replace('\r',' ')
                            wDemandado     = '' if pd.isna(item[6]) else item[6].replace('\r',' ')
                            wFechaAuto     = '' if pd.isna(item[7]) else item[7].replace('\r',' ') #13 DE DICIEMBRE 23
                            wAutoAnotacion = ''
                            
                            # wExiste = True
                            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion )
                            if wExisteEnlace:
                                wRutaDcto = self.app_config.wRutaDescPdfSel
                            #fin if

                            wExisteRadicado = True
                            break
                        #fin if 
                        
                        conta+=1
                    #fin for
                #fin if 
                
                if wExisteRadicado:
                    break
                #fin if 
                
                init_Page+=1
            #fin while 
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDcto
    ######## fnBuscarEstadoModelo06
    
    def fnBuscarEstadoModelo07( self, wNombreDocumento, driver, datos, json_tbusqueda):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda )
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo07
    
    def fnBuscarEstadoModelo18( self, wNombreDocumento, driver, datos, json_tbusqueda):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda )
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo18
    
    def fnBuscarEstadoModelo19( self, wNombreDocumento, driver, datos, json_tbusqueda):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda )
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo19
    
    def fnBuscarEstadoModelo08( self, wNombreDocumento ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        if wExisteRadicado:
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if 
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo08 
    
    def fnBuscarEstadoModelo16( self, wNombreDocumento ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        if wExisteRadicado:
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if 
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo16 
    
    def fnBuscarEstadoModelo11( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )  
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11) 
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado :
            ObjPdfreader = PdfReader(wRutaArchivo) 
            init_Page = 1
            num_pages = len(ObjPdfreader.pages)
            wExisteRadicado   = False
                                
            while init_Page <= num_pages: 
                dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
                time.sleep(1)
                
                for item in dfs[0].values.tolist()[0:]:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', ''
                    wColumna01 = '' if pd.isna(item[0]) else item[0].replace('\r',' ')
                    
                    if wColumna01!= '':
                        wNroRadicado = wColumna01.replace(wColumna01.split(' ')[0],'') 
                    #fin if 
                        
                    if wNroRadicado in wArrDatoBusqueda:
                        wClase         =  ('' if pd.isna(item[1]) else item[1].replace('\r',' ')) if dfs[0].columns[1] == 'ASUNTO' else wColumna01.replace(wColumna01.split(' ')[0],'').strip()
                        wDemandante    =  ('' if pd.isna(item[2]) else item[2].replace('\r',' ')) if dfs[0].columns[1] == 'ASUNTO' else ('' if pd.isna(item[1]) else item[1].replace('\r',' '))
                        wDemandado     =  ('' if pd.isna(item[3]) else item[3].replace('\r',' ')) if dfs[0].columns[1] == 'ASUNTO' else ('' if pd.isna(item[2]) else item[2].replace('\r',' '))
                        wFechaAuto     =  ('' if pd.isna(item[4]) else item[4].replace('\r',' ')) if dfs[0].columns[1] == 'ASUNTO' else ('' if pd.isna(item[3]) else item[3].replace('\r',' ')) #dd/mm/YYYY
                        wAutoAnotacion =  ('' if pd.isna(item[6]) else item[6].replace('\r',' ')) if dfs[0].columns[1] == 'ASUNTO' else ('' if pd.isna(item[4]) else item[4].replace('\r',' ')) #Decision
                    
                    
                        wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
                        if wExisteEnlace:
                            wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                        #fin if
                        wExisteRadicado = True

                        break 
                    #fin if
                    
                    conta+=1
                #fin for
                
                if wExisteRadicado: 
                    break
                #fin if
                
                init_Page+=1
            #fin whilev
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo11 
    
    def fnBuscarEstadoModelo12( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )  
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11) 
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado :
            ObjPdfreader = PdfReader(wRutaArchivo)

            init_Page = 1
            num_pages = len(ObjPdfreader.pages)
            wExisteRadicado = False
                                
            while init_Page <= num_pages: 
                dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
                time.sleep(1)
                
                wPosicion = 0 if dfs[0].values.tolist()[0:][0][0] == 'Número' else 1
                for item in dfs[wPosicion].values.tolist()[0:]:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 

                    wNroRadicado = '' if pd.isna(item[1]) else item[1].replace('\r',' ')
                        
                    if wNroRadicado in wArrDatoBusqueda:
                            
                        wClase         = '' if pd.isna(item[0]) else item[0].replace('\r',' ')
                        wDemandante    = '' if pd.isna(item[2]) else item[2].replace('\r',' ')
                        wDemandado     = '' if pd.isna(item[3]) else item[3].replace('\r',' ')
                        wFechaAuto     = '' if pd.isna(item[4]) else item[4].replace('\r',' ') #28 de febrero de 2023
                        wAutoAnotacion = '' if pd.isna(item[6]) else item[6].replace('\r',' ') #Decision
                        
                        wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
                        if wExisteEnlace:
                            wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                        #fin if 
                        wExisteRadicado = True
                        
                        break
                    #fin if 
                #fin for
                
                if wExisteRadicado: 
                    break
                #fin if
                
                init_Page+=1
            #fin while
            
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo12
    
    def fnBuscarEstadoModelo20( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDcto        = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )  
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11) 
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado :
            ObjPdfreader = PdfReader(wRutaArchivo) 
            init_Page = 1
            num_pages = len(ObjPdfreader.pages)
            wExisteRadicado = False
                                
            while init_Page <= num_pages: 
                dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
                time.sleep(1)
                
                wPosicion = 0 if dfs[0].values.tolist()[0:][0][0] == 'Número' else 1
                for item in dfs[wPosicion].values.tolist()[0:]:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', ''
                    wNroRadicado = '' if pd.isna(item[0]) else item[0].replace('\r',' ')

                    if wNroRadicado not in ( 'Número', 'N°' ): 
                        wNroRadicado   = wNroRadicado  if wPosicion == 0 else ( '' if pd.isna(item[1]) else item[1].replace('\r',' ') )
                    #fin if 
                        
                    if wNroRadicado in wArrDatoBusqueda:
                        
                        wClase         =  ('' if pd.isna(item[1]) else item[1].replace('\r',' ')) if wPosicion == 0 else ( '' if pd.isna(item[2]) else item[2].replace('\r',' ') )
                        wDemandante    =  ('' if pd.isna(item[2]) else item[2].replace('\r',' ')) if wPosicion == 0 else ( '' if pd.isna(item[3]) else item[3].replace('\r',' ') )
                        wDemandado     =  ('' if pd.isna(item[3]) else item[3].replace('\r',' ')) if wPosicion == 0 else ( '' if pd.isna(item[4]) else item[4].replace('\r',' ') )
                        wFechaAuto     =  ('' if pd.isna(item[4]) else item[4].replace('\r',' ')) if wPosicion == 0 else ( '' if pd.isna(item[5]) else item[5].replace('\r',' ') )
                        wAutoAnotacion =  ('' if pd.isna(item[6]) else item[6].replace('\r',' ')) if wPosicion == 0 else ( '' if pd.isna(item[7]) else item[7].replace('\r',' ') )
                        
                        wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
                        if wExisteEnlace:
                            wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                        #fin if 
                        wExisteRadicado = True
                        break
                    #fin if 
                #fin for
                
                if wExisteRadicado: 
                    break
                #fin if
                
                init_Page+=1
            #fin while
            
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo20 
    
    def fnBuscarEstadoModelo22( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado = False 
        wExisteEnlace   = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda =  self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            #No se puede leer la tabla
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if 
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo22 
    
    def fnBuscarEstadoModelo23( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado :
            ObjPdfreader = PdfReader(wRutaArchivo)
            
            init_Page = 1
            num_pages = len(ObjPdfreader.pages)
            wExisteRadicado = False
                                
            while init_Page <= num_pages: 
                dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
                time.sleep(1) 
                
                for item in dfs[0].values.tolist()[0:]:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                    
                    #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                    wNroRadicado = '' if pd.isna(item[0]) else item[0].replace('\r',' ') 
                    
                    if wNroRadicado in wArrDatoBusqueda:
                    #     wClase         = '' if pd.isna(item[1]) else item[1].replace('\r',' ')
                    #     wDemandante    = '' if pd.isna(item[2]) else item[2].replace('\r',' ')
                    #     wDemandado     = '' if pd.isna(item[3]) else item[3].replace('\r',' ')
                    #     wFechaAuto     = '' if pd.isna(item[4]) else item[4].replace('\r',' ')
                    #     wAutoAnotacion = '' if pd.isna(item[4]) else item[4].replace('\r',' ') 
                        
                        wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion )
                        if wExisteEnlace:
                            wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                        #fin if 
                        wExisteRadicado = True
                        break
                    #fin if  
                #fin for 
                
                if wExisteRadicado: 
                    break
                #fin if
                
                init_Page+=1
            #fin while 
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo23
    
    def fnBuscarEstadoModelo25( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado :
            ObjPdfreader = PdfReader(wRutaArchivo) 
            init_Page = 1
            num_pages = len(ObjPdfreader.pages)
            wExisteRadicado = False
                                
            while init_Page <= num_pages: 
                dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
                time.sleep(1) 
                
                for item in dfs[0].values.tolist()[0:]:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                    
                    #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                    wNroRadicado = '' if pd.isna(item[1]) else item[1].replace('\r',' ')
                    
                    if wNroRadicado in wArrDatoBusqueda: 
                        wClase         = '' if pd.isna(item[3]) else item[3].replace('\r',' ')
                        wDemandante    = '' if pd.isna(item[5]) else item[5].replace('\r',' ')
                        wDemandado     = '' if pd.isna(item[7]) else item[7].replace('\r',' ')
                        wFechaAuto     = '' if pd.isna(item[9]) else item[9].replace('\r',' ')
                        wAutoAnotacion = '' if pd.isna(item[11]) else item[11].replace('\r',' ') 
                        
                        wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion )
                        if wExisteEnlace:
                            wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                        #fin if 
                        wExisteRadicado = True
                        break
                    #fin if  
                #fin for
                    
                if wExisteRadicado: 
                    break
                #fin if
                
                init_Page+=1
            #fin while 
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo25 
    
    def fnBuscarEstadoModelo26( self, wNombreDocumento, driver, datos, json_tbusqueda):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)

        ObjPdfreader = PdfReader(wRutaArchivo)

        # Get the number of pages in the PDF file
        init_Page = 1
        num_pages = len(ObjPdfreader.pages) 
                            
        while init_Page <= num_pages: 
            dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
            time.sleep(1) 
                
            for item in dfs[0].values.tolist()[0:]:
                wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                
                #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                wNroRadicado = '' if pd.isna(item[0]) else item[0].replace('\r',' ') 
                
                if wNroRadicado in wArrDatoBusqueda: 
                    # No se puede leer auto anotación 
                    
                    # Descargando unico enlace 
                    wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda )
                    if wExisteEnlace:
                        wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                    #fin if

                    wExisteRadicado = True
                    break
                #fin if 
                
            #fin for 
            
            if wExisteRadicado: 
                break
            #fin if
            
            init_Page+=1
        #fin while  
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo26 
    
    def fnBuscarEstadoModelo27( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDcto        = self.app_config.wRutaDescPdfSel
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
    
        ####### Proceso de busqueda de URL 
        ObjPdfreader = PdfReader(wRutaArchivo)

        # Get the number of pages in the PDF file
        init_Page = 1
        num_pages = len(ObjPdfreader.pages) 
                            
        while init_Page <= num_pages: 
            dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
            time.sleep(1) 
                
            for item in dfs[0].values.tolist()[0:]:
                wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                
                #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                wNroRadicado =  '' if pd.isna(item[0]) else item[0].replace('\r',' ') 
                
                if wNroRadicado in wArrDatoBusqueda:  
                    
                    wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion )
                    wExisteRadicado = True
                    break
                #fin if  
            #fin for
            
            
            if wExisteRadicado: 
                break
            #fin if
            
            init_Page+=1
        #fin while  
        
        return wExisteRadicado, wExisteEnlace, wRutaDcto
    ######## fnBuscarEstadoModelo27 
    
    def fnBuscarEstadoModelo28( self, wNombreDocumento, driver, datos, json_tbusqueda):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)

        ObjPdfreader = PdfReader(wRutaArchivo)

        init_Page = 1
        num_pages = len(ObjPdfreader.pages) 
                            
        while init_Page <= num_pages: 
            dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
            time.sleep(1) 
                
            for item in dfs[0].values.tolist()[0:]:
                wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                
                #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                wNroRadicado = '' if pd.isna(item[1]) else item[1].replace('\r',' ') 
                
                if wNroRadicado in wArrDatoBusqueda: 
                    wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda )
                    if wExisteEnlace:
                        wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                    #fin if
                    wExisteRadicado = True
                    break
                #fin if
            #fin for
            
            if wExisteRadicado: 
                break
            #fin if
            
            init_Page+=1
        #fin while  
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo28 
    
    def fnBuscarEstadoModelo29( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda =  self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            #No se puede leer la tabla
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo29 
    
    def fnBuscarEstadoModelo30( self, wNombreDocumento ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado = False 
        wExisteEnlace   = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfSel
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)

        ObjPdfreader = PdfReader(wRutaArchivo)

        init_Page    = 1
        num_pages    = len(ObjPdfreader.pages) 
                            
        while init_Page <= num_pages: 
            dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
            time.sleep(1) 
                
            for item in dfs[0].values.tolist()[0:]:
                try:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                    
                    #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                    wNroRadicado = '' if pd.isna(item[2]) else item[2].replace('\r',' ').replace('.',' ').strip()
                    
                    if wNroRadicado in wArrDatoBusqueda: 
                        # 
                        wClase         = '' if pd.isna(item[3]) else item[3].replace('\r',' ')
                        wDemandante    = '' if pd.isna(item[4]) else item[4].replace('\r',' ')
                        wDemandado     = '' if pd.isna(item[5]) else item[5].replace('\r',' ')
                        wFechaAuto     = '' if pd.isna(item[6]) else item[6].replace('\r',' ')
                        wAutoAnotacion = ''
                        
                        wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
                        if wExisteEnlace:
                            wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                        #fin if
                        wExisteRadicado = True
                        break
                    #fin if 
                except:
                    pass
                #fin try
            #fin for
            
            if wExisteRadicado: 
                break
            #fin if
            
            init_Page+=1
        #fin while  
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo30 
    
    def fnBuscarEstadoModelo31( self, wNombreDocumento, driver, datos, json_tbusqueda ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado = False 
        wExisteEnlace   = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfSel
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)

        ObjPdfreader = PdfReader(wRutaArchivo)

        # Get the number of pages in the PDF file
        init_Page    = 1
        num_pages    = len(ObjPdfreader.pages) 
                            
        while init_Page <= num_pages: 
            dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
            time.sleep(1) 
                
            for item in dfs[0].values.tolist()[0:]:
                try:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                    
                    #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                    wNroRadicado = '' if pd.isna(item[1]) else item[1].replace('\r',' ').replace('.',' ').strip()
                    
                    if wNroRadicado in wArrDatoBusqueda: 
                        
                        # Columnas desordenadas. 
                        wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda )
                        if wExisteEnlace:
                            wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                        #fin if
                        wExisteRadicado = True
                        break
                    #fin if 
                except:
                    pass
                #fin try
            #fin for
            
            if wExisteRadicado: 
                break
            #fin if
            
            init_Page+=1
        #fin while  
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo31 
     
    def fnBuscarEstadoModelo32( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo32
     
    def fnBuscarEstadoModelo34( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado = False 
        wExisteEnlace   = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo34
    
    def fnBuscarEstadoModelo37( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo    = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion  = str(wNombreDocumento).split('_')[0]
        wExisteRadicado = False 
        wExisteEnlace   = False
        wRutaDctoPdf    = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion )  
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo37
    
    def fnBuscarEstadoModelo38( self, wNombreDocumento, driver, datos, json_tbusqueda ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' + wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False
        wRutaDctoPdf    = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        if wExisteRadicado:
            wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda )
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if 
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo38 
    
    def fnBuscarEstadoModelo40( self, wNombreDocumento, driver, datos, json_tbusqueda ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        if wExisteRadicado:
            wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda )
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if 
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo40 
    
    def fnBuscarEstadoModelo41( self, wNombreDocumento, driver, datos, json_tbusqueda ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        if wExisteRadicado:
            wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( driver, datos, wRutaArchivo, wNroRadicacion, json_tbusqueda )
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if 
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo41
      
    def fnBuscarEstadoModelo42( self, wNombreDocumento ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado = False 
        wExisteEnlace   = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)

        ObjPdfreader = PdfReader(wRutaArchivo)

        # Get the number of pages in the PDF file
        init_Page    = 1
        num_pages    = len(ObjPdfreader.pages) 
                            
        while init_Page <= num_pages: 
            dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
            time.sleep(1) 
                
            for item in dfs[0].values.tolist()[0:]:
                try:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                    
                    #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                    wNroRadicado = '' if pd.isna(item[0]) else item[0].replace('\r',' ').replace('.',' ').strip()
                    
                    for radicado in wArrDatoBusqueda:
                        if radicado in wNroRadicado : 
                            # Columnas desordenadas. 
                            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
                            if wExisteEnlace:
                                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                            #fin if
                            wExisteRadicado = True
                            break
                        #fin if
                    #fin for
                    
                    if wExisteRadicado:
                        break
                    #fin if 
                except:
                    pass
                #fin try
            #fin for
            
            if wExisteRadicado: 
                break
            #fin if
            
            init_Page+=1
        #fin while  
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo42 
      
    def fnBuscarEstadoModelo43( self, wNombreDocumento ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)

        ObjPdfreader = PdfReader(wRutaArchivo)

        # Get the number of pages in the PDF file
        init_Page    = 1
        num_pages    = len(ObjPdfreader.pages) 
                            
        while init_Page <= num_pages: 
            dfs = tabula.read_pdf(wRutaArchivo, encoding='latin-1' , pages=init_Page,relative_area=True, relative_columns=True )                                                                        
            time.sleep(1) 
                
            for item in dfs[0].values.tolist()[0:]:
                try:
                    wNroRadicado, wClase, wDemandante, wDemandado, wFechaAuto, wAutoAnotacion = '', '', '', '', '', '' 
                    
                    #SOLO CUANDO NO ES LA CABECERA REALIZA EL PROCESO
                    wNroRadicado = '' if pd.isna(item[0]) else item[0].replace('\r',' ').replace('.',' ').strip()
                    
                    for radicado in wArrDatoBusqueda:
                        if radicado in wNroRadicado :  
                            wActuacion = '' if pd.isna(item[1]) else item[1].replace('\r',' ').replace('.',' ').strip()
                    
                            # Columnas desordenadas. 
                            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
                            if wExisteEnlace:
                                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
                            #fin if
                            wExisteRadicado = True
                            break
                        #fin if
                    #fin for
                    
                    if wExisteRadicado:
                        break
                    #fin if 
                except:
                    pass
                #fin try
            #fin for
            
            if wExisteRadicado: 
                break
            #fin if
            
            init_Page+=1
        #fin while  
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo43 
      
    def fnBuscarEstadoModelo47( self, wNombreDocumento ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda =  self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            #No se puede leer la tabla
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion )
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if 
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo47 
     
    def fnBuscarEstadoModelo52( self, wNombreDocumento ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0] 
        wExisteRadicado  = False
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        if wExisteRadicado:
            wExisteEnlace, wRutaDctoPdf = self.fnSeleccionandoEnlace1raPagina( wRutaArchivo, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if 
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo52 - ref mod 38
     
    def fnBuscarEstadoModelo57( self, wNombreDocumento ):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0] 
        wExisteRadicado  = False
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        if wExisteRadicado:
            wExisteEnlace, wRutaDctoPdf = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if 
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo57 - ref mod 38
     
    def fnBuscarEstadoModelo113( self, wNombreDocumento ): 
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0] 
        wExisteRadicado  = False
        wExisteEnlace    = False
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion ) 
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        if wExisteRadicado:
            wExisteEnlace, wRutaDctoPdf = self.fnPresionandoClickEnlacePdf( wRutaArchivo, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if 
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo113 - ref mod 38
    
    def fnBuscarEstadoModelo39( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, self.app_config.wRutaDescPdfSel
    ######## fnBuscarEstadoModelo39
    
    def fnBuscarEstadoModelo69( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo69
    
    def fnBuscarEstadoModelo70( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo70
    
    def fnBuscarEstadoModelo72( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo72
    
    def fnBuscarEstadoModelo88( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo88
    
    def fnBuscarEstadoModelo91( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo91
    
    def fnBuscarEstadoModelo95( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo95
    
    def fnBuscarEstadoModelo102( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo102
    
    def fnBuscarEstadoModelo98( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo98
    
    def fnBuscarEstadoModelo104( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo104
    
    def fnBuscarEstadoModelo105( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo105
    
    def fnBuscarEstadoModelo107( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion )
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo107
    
    def fnBuscarEstadoModelo112( self, wNombreDocumento):
        #Abrimos pdf 
        wRutaArchivo     = self.app_config.wRutaDescPdfFin + 'ES-' +wNombreDocumento + '.pdf'
        wNroRadicacion   = str(wNombreDocumento).split('_')[0]
        wExisteRadicado  = False 
        wExisteEnlace    = False 
        wRutaDctoPdf     = self.app_config.wRutaDescPdfFin
        wArrDatoBusqueda = self.fnRadicadosBusquedaPdf( wNroRadicacion )
        
        self.app_general.fnImprimir("Leyendo Listado de estados Pdf ",11)
        texto_pdf = extract_text(wRutaArchivo)
        
        for datoBusqueda in wArrDatoBusqueda:
            if datoBusqueda in str(texto_pdf).replace("Verbal","").replace("  "," ").replace("  "," "):
                wExisteRadicado = True
                break
            #fin if
        #fin for
        
        ####### Proceso de busqueda de URL
        if wExisteRadicado : 
            wExisteEnlace = self.fnPresionandoClickEnlacePdf( wNombreDocumento, wNroRadicacion ) 
            if wExisteEnlace:
                wRutaDctoPdf = self.app_config.wRutaDescPdfSel
            #fin if
        #fin if     
        
        return wExisteRadicado, wExisteEnlace, wRutaDctoPdf
    ######## fnBuscarEstadoModelo112
    
    ############ Fin
    
    ############ Modelo de Providencia
    def fnBuscarProvidenciaModelo01( self, wNombreDocumento, razon_social = '', wRutaDescarga = '' ):
        try:
            # variables
            wExiste = False
            wNroRadicacion    = str(wNombreDocumento).split('_')[0]
            wNombreDocumento  = 'EE-' + wNombreDocumento
            wRutaArchivo      = (self.app_config.wRutaDescPdfFin if wRutaDescarga == '' else self.app_config.wRutaDescPdfFinV2 ) +  wNombreDocumento + '.pdf'
            wRutaArchivoFinal = self.app_config.wRutaDescPdfSel + wNombreDocumento + '.pdf'
            # wRutaArchivoDrive = self.app_config.wRutaAlmPdfDrive + wNroRadicacion + '.pdf'
            
            self.app_general.fnImprimir("Leyendo Listado de Autos Pdf",11)
            if self.app_general.fnExisteArchivo( wRutaArchivo ): 
                wArrDiffNroRadicacion = self.fnRetornarRadicadoBusqueda( wNroRadicacion )
                if razon_social != '':
                    wArrDiffNroRadicacion.append( razon_social )
                #fini f
                
                # Abre el archivo original y crea un escritor para el nuevo PDF
                with open(wRutaArchivo, 'rb') as file:
                    pdf_reader = PdfReader(file)
                    pdf_writer = PdfWriter()
                    num_pages  = len(pdf_reader.pages)
                    wNroRadicacionEncontrado = ''

                    # Crear un nuevo PDF con las páginas deseadas
                    contador = 0
                    for page_num in range(num_pages):
                        try:
                            page = pdf_reader.pages[page_num]
                            wTextoExtraido =  page.extract_text().replace('  ',' ').replace('  ',' ').replace(' – ','-').replace(' -','-').replace('– ','-')
                            if not wExiste:
                                for buscar in wArrDiffNroRadicacion:
                                    if buscar in wTextoExtraido:
                                        pdf_writer.add_page(page)
                                        wExiste = True
                                        wNroRadicacionEncontrado = buscar 
                                        
                                        break
                                    #fin if
                                #fin try 
                            #fin if 
                            
                            if (wExiste == True and contador == 0):
                                contador+=1
                                continue
                            #fin if 
                            
                            if wExiste: 
                                if ( 
                                    'REPÚBLICA DE COLOMBIA'  in wTextoExtraido or
                                    'Rad. No.'  in wTextoExtraido or
                                    'EXPEDIENTE'  in wTextoExtraido or
                                    'SOLICITUD No'  in wTextoExtraido or
                                    'Rad.'  in wTextoExtraido or
                                    'RAD'   in wTextoExtraido or
                                    'Exp.'   in wTextoExtraido or
                                    'REF'  in wTextoExtraido or
                                    'PROCESO'  in wTextoExtraido or
                                    'Radicación'   in wTextoExtraido or
                                    'en contra de'   in wTextoExtraido or
                                    'Responsabilidad Civil No.'   in wTextoExtraido or
                                    'Declarativo No.'   in wTextoExtraido or
                                    'Proceso ejecutivo No.' in wTextoExtraido or
                                    'Aprehensión y Entrega N°' in wTextoExtraido or
                                    'Ejecutivo quirografario No.' in wTextoExtraido or
                                    'Ejecutivo Singular No.'   in wTextoExtraido or
                                    'Impugnación Actas de Asamblea No.'   in wTextoExtraido or
                                    'Radicado: Ejecutivo Singular No.'   in wTextoExtraido or
                                    'Pertenencia No.'   in wTextoExtraido or
                                    'Expropiación No.'   in wTextoExtraido or
                                    'Rendición Provocada de Cuentas No.'   in wTextoExtraido or
                                    'Disolución, Nulidad y Liquidación No.'   in wTextoExtraido or
                                    'Restitución No.'   in wTextoExtraido or
                                    'Verbal No.'   in wTextoExtraido or
                                    'Acción de Tutela N°'   in wTextoExtraido or
                                    # 'Disolución, Nulidad y Liquidación No.'   in wTextoExtraido or
                                    # 'Disolución, Nulidad y Liquidación No.'   in wTextoExtraido or
                                    'Radicado'  in wTextoExtraido.replace('Radic ado','Radicado').replace('Radicad o','Radicado').replace('Radica do','Radicado').replace('Radi cado','Radicado') 
                                ): 
                                    if wNroRadicacionEncontrado not in wTextoExtraido:
                                        break
                                    else:
                                        pdf_writer.add_page(page)
                                    #fin if 
                                else:
                                    pdf_writer.add_page(page)
                                #fin if 
                            #fin if 
                        except:
                            pass
                        #fin try
                    #fin for 

                    # Guardar el nuevo PDF si se encontraron páginas que cumplen con el criterio
                    if wExiste:
                        self.app_general.fnImprimir("Se encontro Nro de Radicación",13)
                        with open(wRutaArchivoFinal, 'wb') as new_file:
                            pdf_writer.write(new_file)
                        #fin with
                        
                        # self.app_general.fnCopiarArchivo( wRutaArchivoFinal, wRutaArchivoDrive)
                    else:
                        self.app_general.fnImprimir("No se encontraron páginas que cumplan con el criterio especificado.", 13)
                    #fin if
                #fin with
            else:
                self.app_general.fnImprimir("No se pudo encontrar el archivo")
            #fin if
        except:
            pass
        return wExiste,  self.app_config.wRutaDescPdfSel
    #fnBuscarProvidenciaModelo01
    ############ Fin
    
    def fnGuardarDocumento( self, wNombreDocumento, wRutaDctoPdf, wRutaAlmPdfDrive, wNombreDocumentoSalida = '' ):
        #Guardando Documento estado o providencia 
        
        wNombreDocumentoSalida = wNombreDocumento if wNombreDocumentoSalida == '' else wNombreDocumentoSalida
        wRutaDctoDescarga = wRutaDctoPdf + wNombreDocumento
        wRutaAlmacenamientoDcto = wRutaAlmPdfDrive + wNombreDocumentoSalida
        
        self.app_general.fnCopiarArchivo( wRutaDctoDescarga, wRutaAlmacenamientoDcto )
        
        while True:
            if self.app_general.fnExisteArchivo( wRutaAlmacenamientoDcto  ):
                self.app_general.fnEliminarArchivo( wRutaDctoDescarga )
                self.app_general.fnImprimir(f"Se subio al drive exitosamente",13)
                break
            #fin if 
            time.sleep(1)
        #fin while
    #fin if 
#fin clase
